# ============================================================
# app.py - Ứng dụng Quản Lý Công Việc
# Sử dụng: Streamlit + Google Sheets (gspread) + Cloudinary + FPDF
# ============================================================

import streamlit as st
import gspread
from google.oauth2.service_account import Credentials
import cloudinary
import cloudinary.uploader
import uuid
from fpdf import FPDF
import pandas as pd
from datetime import datetime, timedelta, date
import requests
import tempfile
import os
import json
import hashlib
import hmac as _hmac_mod
import threading
import base64 as _b64_mod

# ============================================================
# TẢI FONT UNICODE HỖ TRỢ TIẾNG VIỆT
# ============================================================
FONT_DIR   = os.path.join(os.path.dirname(__file__), "fonts")
FONT_REGULAR = os.path.join(FONT_DIR, "DejaVuSans.ttf")
FONT_BOLD    = os.path.join(FONT_DIR, "DejaVuSans-Bold.ttf")

def _tai_font_neu_chua_co():
    """
    Tự động tải font DejaVuSans về thư mục fonts/ nếu chưa có.
    Dùng jsDelivr CDN để đảm bảo tải đúng file binary TTF.
    """
    os.makedirs(FONT_DIR, exist_ok=True)
    urls = {
        FONT_REGULAR: "https://cdn.jsdelivr.net/npm/dejavu-fonts-ttf@2.37.3/ttf/DejaVuSans.ttf",
        FONT_BOLD:    "https://cdn.jsdelivr.net/npm/dejavu-fonts-ttf@2.37.3/ttf/DejaVuSans-Bold.ttf",
    }
    for duong_dan, url in urls.items():
        if not os.path.exists(duong_dan):
            phan_hoi = requests.get(url, timeout=30)
            phan_hoi.raise_for_status()
            with open(duong_dan, "wb") as f:
                f.write(phan_hoi.content)

# Tải font ngay khi khởi động (chỉ chạy 1 lần vì Streamlit cache module)
_tai_font_neu_chua_co()

# ============================================================
# CẤU HÌNH TRANG ỨNG DỤNG
# ============================================================
st.set_page_config(
    page_title="Quản Lý Công Việc",
    page_icon="📋",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# CSS ẩn toolbar + avatar badge trong iframe app
st.markdown(
    """
    <style>
    /* ẩn toolbar Streamlit */
    [data-testid="stToolbar"]      { display: none !important; visibility: hidden !important; }
    [data-testid="stDecoration"]   { display: none !important; visibility: hidden !important; }
    /* ẩn status widget / avatar badge */
    [data-testid="stStatusWidget"] { display: none !important; visibility: hidden !important; }
    [data-testid="stBottom"]       { display: none !important; visibility: hidden !important; }
    [data-testid="stBottomBlockContainer"] { display: none !important; visibility: hidden !important; }
    /* ẩn theo class name (phòng trường hợp testid thay đổi) */
    [class*="viewerBadge"]         { display: none !important; visibility: hidden !important; }
    [class*="ViewerBadge"]         { display: none !important; visibility: hidden !important; }
    [class*="StatusWidget"]        { display: none !important; visibility: hidden !important; }
    [class*="statusWidget"]        { display: none !important; visibility: hidden !important; }
    [class*="createdBy"]           { display: none !important; visibility: hidden !important; }
    [class*="CreatedBy"]           { display: none !important; visibility: hidden !important; }
    [class*="deployButton"]        { display: none !important; visibility: hidden !important; }
    [class*="manage"]              { display: none !important; visibility: hidden !important; }
    footer                         { display: none !important; visibility: hidden !important; }
    /* ẩn ảnh tròn cứng */
    img[style*="border-radius: 50"] { display: none !important; visibility: hidden !important; }
    img[style*="border-radius:50"]  { display: none !important; visibility: hidden !important; }
    /* ẩn fixed bottom-right badge/avatar */
    iframe[style*="position: fixed"] { display: none !important; }
    iframe[style*="position:fixed"]  { display: none !important; }
    div[style*="position: fixed"][style*="bottom"] { display: none !important; }
    div[style*="position:fixed"][style*="bottom"]  { display: none !important; }
    [class*="badge"], [class*="Badge"], [class*="avatar"], [class*="Avatar"] {
        display: none !important; visibility: hidden !important;
    }
    /* ẩn app creator avatar / viewer badge (bottom-right) */
    [data-testid="appCreatorAvatar"]          { display: none !important; }
    img[data-testid="appCreatorAvatar"]        { display: none !important; width:0 !important; height:0 !important; }
    div:has(> img[data-testid="appCreatorAvatar"])   { display: none !important; }
    div:has(img[data-testid="appCreatorAvatar"])      { display: none !important; }
    a:has(img[data-testid="appCreatorAvatar"])        { display: none !important; }
    [class*="viewerBadge"]                    { display: none !important; }
    [class*="profileContainer"]               { display: none !important; }
    [class*="profilePreview"]                 { display: none !important; }
    [class*="profileImage"]                   { display: none !important; }
    a[href*="share.streamlit.io/user"]        { display: none !important; }
    a[href*="streamlit.io/cloud"]             { display: none !important; }
    /* Giảm font placeholder toàn app */
    input::placeholder, textarea::placeholder {
        font-size: 0.78rem !important;
    }
    /* Selectbox: cho text đã chọn xuống hàng, giữ nguyên input */
    /* Control wrapper */
    [data-testid="stSelectbox"] [data-baseweb="select"] > div {
        height: auto !important;
        min-height: 38px !important;
    }
    /* Search input trong selectbox: luôn ltr để con trỏ ở cuối khi gõ */
    [data-testid="stSelectbox"] [data-baseweb="select"] input,
    [data-testid="stSelectbox"] input {
        direction: ltr !important;
        text-align: left !important;
    }
    /* Selected value: div trực tiếp bên trong ValueContainer không phải input */
    /* Baseweb structure: select > div(control) > div(ValueContainer) > div(singleValue) + div(input) */
    [data-testid="stSelectbox"] [data-baseweb="select"] > div > div > div:not([data-testid]) {
        font-size: 0.72rem !important;
        white-space: nowrap !important;
        overflow: hidden !important;
        text-overflow: ellipsis !important;
        max-width: calc(100% - 28px) !important;
        direction: rtl !important;
        text-align: left !important;
    }
    /* Class-based fallback */
    [data-testid="stSelectbox"] div[class*="singleValue"],
    [data-testid="stSelectbox"] div[class*="SingleValue"],
    [data-testid="stSelectbox"] div[class*="single-value"] {
        font-size: 0.72rem !important;
        white-space: nowrap !important;
        overflow: hidden !important;
        text-overflow: ellipsis !important;
        max-width: calc(100% - 28px) !important;
        direction: rtl !important;
        text-align: left !important;
    }
    /* Dropdown options — cho xuống hàng thay vì cắt ... */
    [data-baseweb="menu"] li,
    [data-baseweb="menu"] [role="option"],
    [data-baseweb="popover"] li,
    [data-baseweb="popover"] [role="option"],
    [role="listbox"] [role="option"],
    ul[data-baseweb="menu"] li,
    [data-baseweb="select-dropdown"] li,
    [data-baseweb="select-dropdown"] [role="option"] {
        white-space: normal !important;
        word-break: break-word !important;
        overflow: visible !important;
        text-overflow: unset !important;
        font-size: 0.72rem !important;
        line-height: 1.3 !important;
        height: auto !important;
        min-height: 32px !important;
        max-height: none !important;
        padding-top: 5px !important;
        padding-bottom: 5px !important;
    }
    /* Tất cả con cháu bên trong option: override bất kỳ inline style nào */
    [data-baseweb="menu"] li *,
    [data-baseweb="menu"] [role="option"] *,
    [role="listbox"] [role="option"] * {
        white-space: normal !important;
        word-break: break-word !important;
        overflow: visible !important;
        text-overflow: unset !important;
        line-height: 1.3 !important;
        max-width: 100% !important;
        font-size: 0.72rem !important;
    }
    /* Span/div text bên trong option cũng cần wrap */
    [data-baseweb="menu"] li span,
    [data-baseweb="menu"] li div,
    [role="listbox"] [role="option"] span,
    [role="listbox"] [role="option"] div {
        white-space: normal !important;
        word-break: break-word !important;
        overflow: visible !important;
        text-overflow: unset !important;
    }
    /* Thêm padding-bottom trên mobile để nội dung không bị avatar badge che khuất */
    @media (max-width: 768px) {
        .main .block-container {
            padding-bottom: 2rem !important;
        }
        section[data-testid="stMain"] > div {
            padding-bottom: 2rem !important;
        }
    }
    </style>
    """,
    unsafe_allow_html=True,
)

# ── Lightbox overlay — click ảnh để xem full size + zoom ──────────────────────
st.markdown("""
<style>
#__lb_ov {
    display:none; position:fixed; inset:0;
    background:rgba(0,0,0,.92); z-index:999999;
    flex-direction:column; align-items:center; justify-content:center;
}
#__lb_ov.open { display:flex; }
#__lb_img_el {
    max-width:90vw; max-height:80vh; object-fit:contain;
    border-radius:6px; box-shadow:0 8px 40px rgba(0,0,0,.7);
    transition:transform .15s; transform-origin:center;
    user-select:none; cursor:default; touch-action:none;
}
#__lb_ctrl { display:flex; gap:12px; margin-top:14px; }
.lb-btn {
    background:rgba(255,255,255,.18); border:none; color:#fff;
    font-size:20px; width:44px; height:44px; border-radius:50%;
    cursor:pointer; display:flex; align-items:center; justify-content:center;
    transition:background .15s;
}
.lb-btn:hover { background:rgba(255,255,255,.38); }
#__lb_cap { color:#bbb; font-size:13px; margin-top:10px; }
img[data-lb] { cursor:zoom-in !important; }
</style>
<div id="__lb_ov" onclick="if(event.target===this)window.__lbClose()">
  <img id="__lb_img_el" src="" onclick="event.stopPropagation()">
  <div id="__lb_ctrl">
    <button class="lb-btn" onclick="window.__lbZoom(1.25)"  title="Phóng to">+</button>
    <button class="lb-btn" onclick="window.__lbZoom(0.8)"   title="Thu nhỏ">−</button>
    <button class="lb-btn" onclick="window.__lbRotate(-90)" title="Xoay trái">⟲</button>
    <button class="lb-btn" onclick="window.__lbRotate(90)"  title="Xoay phải">⟳</button>
    <button class="lb-btn" onclick="window.__lbReset()"     title="Đặt lại">↺</button>
    <button class="lb-btn" onclick="window.__lbClose()"     title="Đóng">✕</button>
  </div>
  <div id="__lb_cap"></div>
</div>
""", unsafe_allow_html=True)

# JS inject CSS vào tất cả frames có thể truy cập
import streamlit.components.v1 as components
components.html(
    """
    <script>
    var HIDE = [
        '[data-testid="stStatusWidget"]',
        '[data-testid="stBottom"]',
        '[data-testid="stBottomBlockContainer"]',
        '[data-testid="stDeployButton"]',
        '[data-testid="manage-app-button"]',
        '[data-testid="stToolbar"]',
        '[data-testid="appCreatorAvatar"]',
        '[class*="StatusWidget"]',
        '[class*="statusWidget"]',
        '[class*="viewerBadge"]',
        '[class*="ViewerBadge"]',
        '[class*="deployButton"]',
        '[class*="createdBy"]',
        '[class*="CreatedBy"]',
        '[class*="profileContainer"]',
        '[class*="profilePreview"]',
        '[class*="profileImage"]',
        'a[href*="share.streamlit.io/user"]',
        'footer',
        'img[style*="border-radius: 50"]',
        'img[style*="border-radius:50"]',
        '[class*="badge"]',
        '[class*="Badge"]',
        '[class*="avatar"]',
        '[class*="Avatar"]',
        'iframe[style*="position: fixed"]',
        'iframe[style*="position:fixed"]',
    ];
    var CSS_RULES = HIDE.map(function(s){ return s+'{display:none!important;visibility:hidden!important}'; }).join('');

    // CSS cho dropdown options wrap 2 hàng (inject vào body vì portal render ngoài DOM chính)
    var DROPDOWN_CSS = [
        '[data-baseweb="menu"] li { white-space:normal!important; word-break:break-word!important; overflow:visible!important; text-overflow:unset!important; height:auto!important; min-height:32px!important; max-height:none!important; padding:5px 10px!important; line-height:1.3!important; font-size:0.72rem!important; }',
        '[data-baseweb="menu"] li * { white-space:normal!important; word-break:break-word!important; overflow:visible!important; text-overflow:unset!important; line-height:1.3!important; max-width:100%!important; font-size:0.72rem!important; }',  
        '[data-baseweb="menu"] li span, [data-baseweb="menu"] li div { white-space:normal!important; word-break:break-word!important; overflow:visible!important; text-overflow:unset!important; }',
        '[role="option"] { white-space:normal!important; word-break:break-word!important; height:auto!important; max-height:none!important; overflow:visible!important; }',
        '[role="option"] * { white-space:normal!important; word-break:break-word!important; overflow:visible!important; text-overflow:unset!important; }',
        '[role="listbox"] { overflow-y:auto!important; }',
        '[data-testid="stSelectbox"] [data-baseweb="select"] [class*="singleValue"] { font-size:0.72rem!important; white-space:nowrap!important; overflow:hidden!important; text-overflow:ellipsis!important; max-width:calc(100% - 28px)!important; }',
        '[data-testid="stSelectbox"] [data-baseweb="select"] [class*="SingleValue"] { font-size:0.72rem!important; white-space:nowrap!important; overflow:hidden!important; text-overflow:ellipsis!important; max-width:calc(100% - 28px)!important; }',
    ].join('');

    function injectCSS(doc) {
        try {
            var id = '__st_hide_v2';
            var existing = doc.getElementById(id);
            if (existing) return;
            var s = doc.createElement('style');
            s.id = id;
            s.textContent = CSS_RULES + DROPDOWN_CSS;
            (doc.head || doc.documentElement).appendChild(s);
        } catch(e) {}
    }

    // Kiểm tra element có nằm trong row công việc con không (marker class "dlgcv...")
    function isInsideCvCmRow(el) {
        try {
            var block = el.closest('[data-testid="stHorizontalBlock"]');
            if (!block) return false;
            return !!block.querySelector('[class*="dlgcv"]');
        } catch(e) { return false; }
    }

    // Fix selectbox NV trong công việc con: bỏ overflow/clip, direction ltr
    function fixCvCmSelectboxes(doc) {
        try {
            doc.querySelectorAll('[data-testid="stHorizontalBlock"]').forEach(function(block) {
                if (!block.querySelector('[class*="dlgcv"]')) return;
                // Tìm tất cả selectbox trong row này
                block.querySelectorAll('[data-testid="stSelectbox"] [data-baseweb="select"]').forEach(function(sel) {
                    // Fix toàn bộ children trực tiếp của baseweb select
                    sel.querySelectorAll('div').forEach(function(div) {
                        div.style.setProperty('overflow', 'visible', 'important');
                        div.style.setProperty('text-overflow', 'clip', 'important');
                        div.style.setProperty('max-width', 'none', 'important');
                        div.style.setProperty('white-space', 'nowrap', 'important');
                        div.style.setProperty('direction', 'ltr', 'important');
                    });
                });
            });
        } catch(e) {}
    }

    // Set font-size nhỏ trực tiếp lên singleValue div (override baseweb inline style)
    function fixSelectedValue(doc) {
        try {
            var sels = [
                '[data-testid="stSelectbox"] [data-baseweb="select"] [class*="singleValue"]',
                '[data-testid="stSelectbox"] [data-baseweb="select"] [class*="SingleValue"]',
            ];
            sels.forEach(function(sel) {
                doc.querySelectorAll(sel).forEach(function(el) {
                    // Bỏ qua NV selectbox trong rows công việc con (hàng 1 layout)
                    if (isInsideCvCmRow(el)) return;
                    // Nhận diện selectbox nhân viên qua wrapper class hoặc label text
                    var stBox = el.closest('[data-testid="stSelectbox"]');
                    var isNvBox = false;
                    if (stBox) {
                        // Kiểm tra wrapper div cv-nv-frag- hoặc cv-nv-row-
                        var wrap = stBox.parentElement;
                        while (wrap) {
                            if (wrap.className && typeof wrap.className === 'string' &&
                                (wrap.className.indexOf('cv-nv-frag-') !== -1 ||
                                 wrap.className.indexOf('cv-nv-row-') !== -1)) {
                                isNvBox = true; break;
                            }
                            if (wrap === doc.body) break;
                            wrap = wrap.parentElement;
                        }
                        // Fallback: kiểm tra label text
                        if (!isNvBox) {
                            var lbl = stBox.querySelector('label');
                            if (lbl && lbl.textContent && lbl.textContent.indexOf('Nh\u00e2n vi\u00ean') !== -1) {
                                isNvBox = true;
                            }
                        }
                    }
                    if (isNvBox) {
                        el.style.setProperty('direction', 'ltr', 'important');
                        el.style.setProperty('overflow', 'visible', 'important');
                        el.style.setProperty('text-overflow', 'clip', 'important');
                        el.style.setProperty('max-width', 'none', 'important');
                        el.style.setProperty('white-space', 'nowrap', 'important');
                        el.style.setProperty('font-size', '0.9rem', 'important');
                        return;
                    }
                    el.style.setProperty('font-size', '0.72rem', 'important');
                    el.style.setProperty('white-space', 'nowrap', 'important');
                    el.style.setProperty('overflow', 'hidden', 'important');
                    el.style.setProperty('text-overflow', 'ellipsis', 'important');
                    el.style.setProperty('max-width', 'calc(100% - 28px)', 'important');
                });
            });
        } catch(e) {}
    }

    // Fix inline styles trên dropdown option items (baseweb dùng inline style vượt qua CSS !important)
    // Chỉ fix trực tiếp li/option và direct children — không đụng input, không querySelectorAll('*')
    function fixDropdownOptions(doc) {
        try {
            // Không chạy khi user đang gõ/xóa trong search input của selectbox
            try {
                var ae = doc.activeElement;
                if (ae && ae.tagName && ae.tagName.toLowerCase() === 'input') return;
                // Kiểm tra window.parent nếu đây là iframe doc
                var pae = window.parent && window.parent.document && window.parent.document.activeElement;
                if (pae && pae.tagName && pae.tagName.toLowerCase() === 'input') return;
            } catch(e) {}
            var sels = [
                '[data-baseweb="menu"] li',
                '[data-baseweb="menu"] [role="option"]',
                '[role="listbox"] [role="option"]',
            ];
            sels.forEach(function(sel) {
                doc.querySelectorAll(sel).forEach(function(el) {
                    el.style.setProperty('white-space', 'normal', 'important');
                    el.style.setProperty('word-break', 'break-word', 'important');
                    el.style.setProperty('overflow', 'visible', 'important');
                    el.style.setProperty('text-overflow', 'unset', 'important');
                    el.style.setProperty('height', 'auto', 'important');
                    el.style.setProperty('max-height', 'none', 'important');
                    el.style.setProperty('font-size', '0.72rem', 'important');
                    el.style.setProperty('line-height', '1.3', 'important');
                    el.style.setProperty('min-height', '28px', 'important');
                    // Chỉ fix direct children (không phải input)
                    for (var i = 0; i < el.children.length; i++) {
                        var child = el.children[i];
                        if (child.tagName && child.tagName.toLowerCase() !== 'input') {
                            child.style.setProperty('white-space', 'normal', 'important');
                            child.style.setProperty('word-break', 'break-word', 'important');
                            child.style.setProperty('overflow', 'visible', 'important');
                            child.style.setProperty('text-overflow', 'unset', 'important');
                            child.style.setProperty('max-width', '100%', 'important');
                            child.style.setProperty('font-size', '0.72rem', 'important');
                        }
                    }
                });
            });
        } catch(e) {}
    }

    var _VI_MONTHS_MAP = {
        'January':'Tháng 1','February':'Tháng 2','March':'Tháng 3',
        'April':'Tháng 4','May':'Tháng 5','June':'Tháng 6',
        'July':'Tháng 7','August':'Tháng 8','September':'Tháng 9',
        'October':'Tháng 10','November':'Tháng 11','December':'Tháng 12'
    };
    function vietHoaCalendar(doc) {
        try {
            // 1. Thay option text trong <select> (month select của react-datepicker)
            doc.querySelectorAll('select').forEach(function(sel) {
                sel.querySelectorAll('option').forEach(function(opt) {
                    var t = (opt.text || '').trim();
                    if (_VI_MONTHS_MAP[t]) opt.text = _VI_MONTHS_MAP[t];
                });
            });
            // 2. TreeWalker: thay bất kỳ text node nào chứa đúng tên tháng tiếng Anh
            try {
                var walk = doc.createTreeWalker(doc.documentElement, 4, null, false);
                var n;
                while ((n = walk.nextNode())) {
                    var val = n.nodeValue;
                    if (!val) continue;
                    var trimmed = val.trim();
                    if (_VI_MONTHS_MAP[trimmed]) {
                        n.nodeValue = val.replace(trimmed, _VI_MONTHS_MAP[trimmed]);
                    }
                }
            } catch(e2) {}
        } catch(e) {}
    }

    var _VI_TEXT = [
        ['Drag and drop file here', 'Kéo thả hoặc chọn file'],
        ['Drop file here', 'Thả file vào đây'],
        ['Limit 200MB per file', 'Tối đa 200MB mỗi file'],
        ['Browse files', 'Chọn file'],
    ];
    function vietHoaUploader(doc) {
        try {
            _VI_TEXT.forEach(function(pair) {
                doc.querySelectorAll('span, p, small, div, button').forEach(function(el) {
                    if (el.childNodes.length === 1 && el.childNodes[0].nodeType === 3) {
                        var t = el.childNodes[0].nodeValue;
                        if (t && t.trim() === pair[0]) {
                            el.childNodes[0].nodeValue = pair[1];
                        }
                    }
                });
            });
        } catch(e) {}
    }

    function hideElements(doc) {
        try {
            HIDE.forEach(function(sel) {
                try {
                    doc.querySelectorAll(sel).forEach(function(el) {
                        el.style.setProperty('display', 'none', 'important');
                        el.style.setProperty('visibility', 'hidden', 'important');
                    });
                } catch(e) {}
            });
        } catch(e) {}
    }

    function removeAvatarFromDOM(doc) {
        try {
            // Xóa hẳn element khỏi DOM thay vì chỉ ẩn
            var avatar = doc.querySelector('img[data-testid="appCreatorAvatar"]');
            if (avatar) {
                // Leo lên DOM tìm container ngoài cùng (_profileContainer)
                var el = avatar;
                for (var i = 0; i < 6; i++) {
                    var parent = el.parentElement;
                    if (!parent || parent === doc.body) break;
                    if (parent.className && typeof parent.className === 'string' &&
                        parent.className.indexOf('profileContainer') !== -1) {
                        parent.remove(); return;
                    }
                    el = parent;
                }
                // Fallback: xóa thẳng phần tử leo lên 3 cấp
                var up = avatar.parentElement && avatar.parentElement.parentElement && avatar.parentElement.parentElement.parentElement;
                if (up) up.remove(); else avatar.remove();
            }
        } catch(e) {}
    }

    // Thử inject vào window.top và tất cả ancestors
    var frames = [];
    try { frames.push(window.document); } catch(e) {}
    try { frames.push(window.parent.document); } catch(e) {}
    try { frames.push(window.parent.parent.document); } catch(e) {}
    try { frames.push(window.top.document); } catch(e) {}

    frames.forEach(function(doc) {
        injectCSS(doc);
        hideElements(doc);
        removeAvatarFromDOM(doc);
        vietHoaUploader(doc);
        vietHoaCalendar(doc);
        fixSelectedValue(doc);
        fixCvCmSelectboxes(doc);
    });

    // Chạy sớm sau khi DOM sẵn sàng
    setTimeout(function() {
        frames.forEach(function(doc) { fixSelectedValue(doc); fixCvCmSelectboxes(doc); });
    }, 100);
    setTimeout(function() {
        frames.forEach(function(doc) { fixSelectedValue(doc); fixCvCmSelectboxes(doc); });
    }, 500);

    // MutationObserver để bắt các element được thêm sau
    function observeDoc(doc) {
        try {
            var _fixTimer = null;
            var _menuPresent = false; // track xem menu đã có trong DOM chưa
            var obs = new MutationObserver(function(mutations) {
                injectCSS(doc);
                hideElements(doc);
                removeAvatarFromDOM(doc);
                vietHoaUploader(doc);
                vietHoaCalendar(doc);
                // Chỉ fix khi menu MỚI xuất hiện (không có → có)
                // Không chạy lại khi user gõ/xóa bên trong menu đang mở
                var menuNow = !!(doc.querySelector('[data-baseweb="menu"]') || doc.querySelector('[role="listbox"]'));
                if (menuNow && !_menuPresent) {
                    // Menu vừa mở ra → fix một lần sau 30ms
                    _menuPresent = true;
                    clearTimeout(_fixTimer);
                    _fixTimer = setTimeout(function() {
                        fixDropdownOptions(doc);
                        fixSelectedValue(doc);
                    }, 30);
                } else if (!menuNow && _menuPresent) {
                    _menuPresent = false;
                    // Menu vừa đóng → update selected value display
                    setTimeout(function() { fixSelectedValue(doc); fixCvCmSelectboxes(doc); }, 50);
                } else if (!menuNow) {
                    _menuPresent = false;
                }
            });
            obs.observe(doc.documentElement, { childList: true, subtree: true });
        } catch(e) {}
    }

    frames.forEach(function(doc) { observeDoc(doc); });

    // ── Lightbox JS — inject vào parent window vì st.markdown() strip <script> ──
    (function() {
      try {
        var W = window.parent || window;
        if (W.__lbReady) return;  // đã inject rồi
        W.__lbReady = true;
        var _sc = 1, _rot = 0;
        function _apply() {
          var D = W.document;
          var img = D.getElementById('__lb_img_el');
          if (img) img.style.transform = 'scale('+_sc+') rotate('+_rot+'deg)';
        }
        W.__lbOpen = function(src, cap) {
          var D = W.document;
          var ov = D.getElementById('__lb_ov');
          var img = D.getElementById('__lb_img_el');
          var capEl = D.getElementById('__lb_cap');
          if (!ov || !img) return;
          img.src = src; _sc = 1; _rot = 0; _apply();
          if (capEl) capEl.textContent = cap || '';
          ov.classList.add('open');
          D.body.style.overflow = 'hidden';
        };
        W.__lbClose = function() {
          var D = W.document;
          var ov = D.getElementById('__lb_ov');
          if (ov) ov.classList.remove('open');
          D.body.style.overflow = '';
        };
        W.__lbZoom = function(f) {
          _sc = Math.max(0.2, Math.min(_sc * f, 10)); _apply();
        };
        W.__lbRotate = function(deg) {
          _rot = (_rot + deg + 360) % 360; _apply();
        };
        W.__lbReset = function() {
          _sc = 1; _rot = 0; _apply();
        };
        W.document.addEventListener('keydown', function(e) {
          if (e.key === 'Escape') W.__lbClose();
        });
        W.document.addEventListener('wheel', function(e) {
          var ov = W.document.getElementById('__lb_ov');
          if (!ov || !ov.classList.contains('open')) return;
          e.preventDefault();
          W.__lbZoom(e.deltaY < 0 ? 1.1 : 0.909);
        }, {passive: false, capture: true});
        var _pt = 0;
        W.document.addEventListener('touchstart', function(e) {
          var ov = W.document.getElementById('__lb_ov');
          if (!ov || !ov.classList.contains('open')) return;
          if (e.touches.length === 2)
            _pt = Math.hypot(e.touches[0].clientX-e.touches[1].clientX, e.touches[0].clientY-e.touches[1].clientY);
        }, {passive: true});
        W.document.addEventListener('touchmove', function(e) {
          var ov = W.document.getElementById('__lb_ov');
          if (!ov || !ov.classList.contains('open')) return;
          if (e.touches.length === 2 && _pt > 0) {
            e.preventDefault();
            var ct = Math.hypot(e.touches[0].clientX-e.touches[1].clientX, e.touches[0].clientY-e.touches[1].clientY);
            W.__lbZoom(ct / _pt); _pt = ct;
          }
        }, {passive: false});
      } catch(e) {}
    })();

    // Chạy lại sau 1s và 3s phòng element load chậm
    [1000, 3000, 6000].forEach(function(delay) {
        setTimeout(function() {
            frames.forEach(function(doc) {
                injectCSS(doc);
                hideElements(doc);
                fixDropdownOptions(doc);
                fixSelectedValue(doc);
                fixCvCmSelectboxes(doc);
                vietHoaCalendar(doc);
            });
        }, delay);
    });
    </script>
    """,
    height=1,
)


# Token secret — dùng để ký HMAC cho session token
_TOKEN_SECRET = st.secrets.get("TOKEN_SECRET", "tm_default_secret_xk9p_2026")


def _create_session_token(data: dict) -> str:
    """Mã hoá user data vào token (base64 JSON + HMAC). Không cần server storage."""
    payload = json.dumps(data, ensure_ascii=False, separators=(',', ':'))
    b64 = _b64_mod.urlsafe_b64encode(payload.encode()).decode()
    sig = _hmac_mod.new(_TOKEN_SECRET.encode(), b64.encode(), "sha256").hexdigest()[:24]
    return f"{b64}.{sig}"


def _verify_session_token(token: str):
    """Xác minh và giải mã token. Trả về user data hoặc None nếu không hợp lệ."""
    try:
        b64, sig = token.rsplit(".", 1)
        expected = _hmac_mod.new(_TOKEN_SECRET.encode(), b64.encode(), "sha256").hexdigest()[:24]
        if not _hmac_mod.compare_digest(sig, expected):
            return None
        return json.loads(_b64_mod.urlsafe_b64decode(b64 + "==").decode())
    except Exception:
        return None


def _session_store():
    """Backward compat wrapper — không còn dùng server storage."""
    class _Noop:
        def get(self, k, d=None): return d
        def pop(self, k, *a): return a[0] if a else None
        def __setitem__(self, k, v): pass
    return _Noop()



# ============================================================
# KẾT NỐI GOOGLE SHEETS
# ============================================================
def _xoa_cache_va_thong_bao(loi: Exception):
    """Xóa cache kết nối GSheets để lần sau tự reconnect, trả về thông báo lỗi."""
    for fn in [
        "ket_noi_google_sheets", "_lay_bang_tinh",
        "lay_sheet", "lay_sheet_cong_ty", "lay_sheet_users",
    ]:
        try:
            globals()[fn].clear()
        except Exception:
            pass
    return f"🔌 Mất kết nối mạng — Google Sheets không thể truy cập. Chi tiết: {type(loi).__name__}: {loi}"


@st.cache_resource
def ket_noi_google_sheets():
    """
    Kết nối với Google Sheets sử dụng Service Account Credentials.
    Dùng @st.cache_resource để chỉ kết nối một lần duy nhất.
    """
    pham_vi = [
        "https://spreadsheets.google.com/feeds",
        "https://www.googleapis.com/auth/drive"
    ]
    # Lấy thông tin xác thực từ Streamlit Secrets (file .streamlit/secrets.toml)
    thong_tin_xac_thuc = st.secrets["gcp_service_account"]
    chung_chi = Credentials.from_service_account_info(thong_tin_xac_thuc, scopes=pham_vi)
    khach_hang = gspread.authorize(chung_chi)
    return khach_hang


@st.cache_resource
def _lay_bang_tinh():
    """Trả về đối tượng Spreadsheet, chỉ gọi open() một lần duy nhất."""
    return ket_noi_google_sheets().open("QuanLyCongViec")


@st.cache_resource
def lay_sheet():
    """
    Mở file 'QuanLyCongViec' trên Google Drive và trả về worksheet 'Tasks'.
    Dùng @st.cache_resource để chỉ mở một lần, tái sử dụng đối tượng sheet.
    """
    try:
        return _lay_bang_tinh().worksheet("Tasks")
    except Exception as e:
        raise ConnectionError(_xoa_cache_va_thong_bao(e)) from e


def _lay_sheet_fresh():
    """Lấy worksheet Tasks mới (không cache) — dùng cho các thao tác ghi."""
    try:
        pham_vi = [
            "https://spreadsheets.google.com/feeds",
            "https://www.googleapis.com/auth/drive"
        ]
        thong_tin = st.secrets["gcp_service_account"]
        chung_chi = Credentials.from_service_account_info(thong_tin, scopes=pham_vi)
        gc = gspread.authorize(chung_chi)
        return gc.open("QuanLyCongViec").worksheet("Tasks")
    except Exception:
        # Fallback về cached sheet nếu kết nối mới thất bại
        return lay_sheet()


# ============================================================
# GOOGLE DRIVE (lưu trữ ảnh nghiệm thu) — dùng requests thay httplib2
# ============================================================
GDRIVE_FOLDER_ID = "1-o1gny8zFKQ5NejyUjeinLVdmDpKfWgH"
_DRIVE_SCOPES = ["https://www.googleapis.com/auth/drive"]
_SUBFOLDER_ID_CACHE: dict = {}  # ma_so → Drive folder_id


@st.cache_resource
def _lay_drive_session():
    """Tạo AuthorizedSession — cache lại để tái sử dụng, không tạo mới mỗi lần."""
    from google.oauth2.service_account import Credentials as SACredentials
    from google.auth.transport.requests import AuthorizedSession
    try:
        info = dict(st.secrets["gdrive_service_account"])
        if "private_key" in info:
            info["private_key"] = info["private_key"].replace("\\n", "\n")
        creds = SACredentials.from_service_account_info(info, scopes=_DRIVE_SCOPES)
    except Exception:
        _base = os.path.dirname(os.path.abspath(__file__))
        creds = SACredentials.from_service_account_file(
            os.path.join(_base, "gdrive_key.json"), scopes=_DRIVE_SCOPES
        )
    return AuthorizedSession(creds)


def _get_or_create_drive_subfolder(ma_so: str) -> str:
    """Lấy hoặc tạo subfolder tên mã số trong GDRIVE_FOLDER_ID. Cache kết quả trong session."""
    if ma_so in _SUBFOLDER_ID_CACHE:
        return _SUBFOLDER_ID_CACHE[ma_so]
    session = _lay_drive_session()
    # Tìm folder đã tồn tại
    r = session.get(
        "https://www.googleapis.com/drive/v3/files",
        params={
            "q": f"name='{ma_so}' and '{GDRIVE_FOLDER_ID}' in parents"
                 " and mimeType='application/vnd.google-apps.folder' and trashed=false",
            "fields": "files(id)",
            "supportsAllDrives": "true",
            "includeItemsFromAllDrives": "true",
        },
    )
    files = r.json().get("files", [])
    if files:
        folder_id = files[0]["id"]
    else:
        import json as _json
        r2 = session.post(
            "https://www.googleapis.com/drive/v3/files?supportsAllDrives=true&fields=id",
            headers={"Content-Type": "application/json; charset=UTF-8"},
            data=_json.dumps({
                "name": ma_so,
                "mimeType": "application/vnd.google-apps.folder",
                "parents": [GDRIVE_FOLDER_ID],
            }),
        )
        folder_id = r2.json().get("id", GDRIVE_FOLDER_ID)
    _SUBFOLDER_ID_CACHE[ma_so] = folder_id
    return folder_id


def _lay_ma_so_tu_task_id(task_id) -> str:
    """Tra cứu mã số từ task_id (dùng cached dataframe). Trả chuỗi rỗng nếu không tìm thấy."""
    try:
        df = lay_danh_sach_cong_viec()
        tid = int(float(str(task_id)))
        row = df[df["ID"] == tid]
        if not row.empty:
            ma = str(row.iloc[0].get("Mã Số", "")).strip()
            if ma and ma not in ("", "nan"):
                return ma
    except Exception:
        pass
    return ""


def _compress_image(content: bytes, mime: str, max_px: int = 1200, quality: int = 82) -> bytes:
    """Giảm kích thước ảnh xuống max_px cạnh dài, giữ tỉ lệ. Bỏ qua nếu đã nhỏ."""
    try:
        from PIL import Image, ImageOps
        import io as _io
        img = Image.open(_io.BytesIO(content))
        img = ImageOps.exif_transpose(img)
        w, h = img.size
        if max(w, h) > max_px:
            ratio = max_px / max(w, h)
            img = img.resize((int(w * ratio), int(h * ratio)), Image.LANCZOS)
        if img.mode in ("RGBA", "P"):
            img = img.convert("RGB")
        buf = _io.BytesIO()
        img.save(buf, format="JPEG", quality=quality, optimize=True)
        compressed = buf.getvalue()
        # Chỉ dùng bản nén nếu nhỏ hơn bản gốc
        return compressed if len(compressed) < len(content) else content
    except Exception:
        return content


def tai_anh_len_drive(file_anh, max_px: int = 1200, quality: int = 82, ma_so: str = "", task_id: str = "") -> str:
    """Upload ảnh lên Google Drive qua requests, set public, trả về thumbnail URL.
    Nếu ma_so được cung cấp, ảnh sẽ được lưu vào subfolder tên mã số trong GDRIVE_FOLDER_ID.
    task_id và ma_so được lưu vào Drive file properties để trace recovery nếu sheet write thất bại.
    """
    import io, json as _json
    session = _lay_drive_session()
    try:
        file_anh.seek(0)
    except Exception:
        pass
    content = file_anh.read()
    if not content:
        raise ValueError("File ảnh rỗng hoặc không hợp lệ — vui lòng chọn lại ảnh.")
    mime = getattr(file_anh, "type", "image/jpeg")
    name = getattr(file_anh, "name", "image.jpg")

    # Compress ảnh trước khi upload (bỏ qua video)
    if mime.startswith("image/"):
        content = _compress_image(content, mime, max_px=max_px, quality=quality)
        name = name.rsplit(".", 1)[0] + ".jpg"
        mime = "image/jpeg"

    # Chọn folder đích: subfolder theo mã số nếu có, fallback về root
    try:
        parent_id = _get_or_create_drive_subfolder(str(ma_so)) if ma_so else GDRIVE_FOLDER_ID
    except Exception:
        parent_id = GDRIVE_FOLDER_ID

    # Lưu task_id và ma_so vào Drive file properties để có thể trace nếu sheet write thất bại
    props = {}
    if task_id:
        props["task_id"] = str(task_id)
    if ma_so:
        props["ma_so"] = str(ma_so)

    # Multipart upload trực tiếp — không qua httplib2
    meta: dict = {"name": name, "parents": [parent_id]}
    if props:
        meta["properties"] = props
    metadata = _json.dumps(meta)
    resp = session.post(
        "https://www.googleapis.com/upload/drive/v3/files"
        "?uploadType=multipart&supportsAllDrives=true&fields=id",
        files={
            "metadata": ("metadata", metadata, "application/json; charset=UTF-8"),
            "file": (name, io.BytesIO(content), mime),
        },
        timeout=60,
    )
    if not resp.ok:
        st.error(f"Drive upload lỗi {resp.status_code}: {resp.text[:500]}")
        st.stop()
    file_id = resp.json()["id"]

    # Set public read — chạy nền để không block UI
    def _set_public(fid):
        try:
            session.post(
                f"https://www.googleapis.com/drive/v3/files/{fid}/permissions"
                "?supportsAllDrives=true",
                json={"type": "anyone", "role": "reader"},
                timeout=30,
            )
        except Exception:
            pass
    threading.Thread(target=_set_public, args=(file_id,), daemon=True).start()

    # thumbnail URL — lưu vào DB, dùng file_id để fetch khi hiển thị
    return f"https://drive.google.com/thumbnail?id={file_id}&sz=w800"


@st.cache_data(ttl=3600)
def _lay_bytes_anh_drive(file_id: str) -> bytes:
    """Fetch bytes ảnh từ Drive qua service account (cache 1 giờ)."""
    session = _lay_drive_session()
    resp = session.get(
        f"https://www.googleapis.com/drive/v3/files/{file_id}"
        f"?alt=media&supportsAllDrives=true",
        timeout=30,
    )
    resp.raise_for_status()
    return resp.content


def _hien_thi_anh_drive(url: str, **kwargs):
    """Hiển thị ảnh Drive: fetch bytes qua service account, click để xem full size."""
    import re, base64
    caption      = kwargs.get("caption", "")
    width_full   = kwargs.get("use_container_width", False)
    img_style    = (
        "width:100%;border-radius:4px;cursor:zoom-in;"
        if width_full else
        "max-width:100%;border-radius:4px;cursor:zoom-in;"
    )
    m = re.search(r"[?&]id=([^&]+)", url)
    if m:
        try:
            data = _lay_bytes_anh_drive(m.group(1))
            mime = "image/png" if data[:4] == b'\x89PNG' else "image/jpeg"
            b64  = base64.b64encode(data).decode()
            cap_js = (caption or "").replace("\\", "\\\\").replace("'", "\\'")
            st.markdown(
                f'<img src="data:{mime};base64,{b64}" '
                f'style="{img_style}" data-lb="1" '
                f'onclick="window.__lbOpen(this.src,\'{cap_js}\')">',
                unsafe_allow_html=True,
            )
            if caption:
                st.caption(caption)
            return
        except Exception:
            pass
    st.image(url, **kwargs)


def cau_hinh_cloudinary():
    """Giữ lại để tương thích — không dùng nữa."""
    pass


def tai_anh_len_cloudinary(file_anh, max_px: int = 1200, quality: int = 82, ma_so: str = "", task_id: str = "") -> str:
    """Wrapper gọi sang Drive (giữ tên cũ để không cần đổi call sites)."""
    return tai_anh_len_drive(file_anh, max_px=max_px, quality=quality, ma_so=ma_so, task_id=task_id)


def _tai_media_len_drive(file_obj, ma_so: str = "", task_id: str = "") -> str:
    """Upload ảnh hoặc video lên Drive. Video → trả view URL; ảnh → thumbnail URL."""
    import re as _re
    mime = getattr(file_obj, "type", "")
    # Ảnh: compress 900px/75 thay vì 1200px/82 để upload nhanh hơn
    url = tai_anh_len_drive(file_obj, max_px=600, quality=65, ma_so=ma_so, task_id=task_id)
    if mime.startswith("video/"):
        m = _re.search(r"[?&]id=([^&]+)", url)
        if m:
            return f"https://drive.google.com/file/d/{m.group(1)}/view"
    return url


# ============================================================
# CÁC HÀM THAO TÁC DỮ LIỆU GOOGLE SHEETS
# ============================================================
@st.cache_resource
def lay_sheet_cong_ty():
    """
    Mở sheet 'Companies' — nếu chưa tồn tại thì tự tạo mới.
    Cấu trúc: ID | Cong_Ty | Created_Date
    Dùng @st.cache_resource để tái sử dụng đối tượng sheet.
    """
    try:
        bang_tinh = _lay_bang_tinh()
        try:
            sheet = bang_tinh.worksheet("Companies")
        except gspread.exceptions.WorksheetNotFound:
            sheet = bang_tinh.add_worksheet(title="Companies", rows=200, cols=7)
            sheet.append_row(["ID", "Tên Công Ty", "Địa Chỉ", "Mã Khách Hàng", "Mã Số Thuế", "Ngày Tạo"])
            return sheet
        # Migration: thêm các cột mới nếu chưa có
        headers = sheet.row_values(1)
        if "Địa Chỉ" not in headers:
            sheet.insert_cols([["Địa Chỉ"]], col=3)
            headers = sheet.row_values(1)
        if "Mã Khách Hàng" not in headers:
            # Chèn trước "Ngày Tạo"
            ngay_col = headers.index("Ngày Tạo") + 1 if "Ngày Tạo" in headers else len(headers)
            sheet.insert_cols([["Mã Khách Hàng"]], col=ngay_col)
            headers = sheet.row_values(1)
        if "Mã Số Thuế" not in headers:
            ngay_col = headers.index("Ngày Tạo") + 1 if "Ngày Tạo" in headers else len(headers)
            sheet.insert_cols([["Mã Số Thuế"]], col=ngay_col)
        return sheet
    except (ConnectionError, OSError, Exception) as e:
        raise ConnectionError(_xoa_cache_va_thong_bao(e)) from e


@st.cache_data(ttl=60)
def lay_danh_sach_cong_ty() -> pd.DataFrame:
    """Lấy toàn bộ danh sách công ty từ sheet 'Companies'."""
    _COT = ["ID", "Tên Công Ty", "Địa Chỉ", "Mã Khách Hàng", "Mã Số Thuế", "Ngày Tạo"]
    try:
        sheet = lay_sheet_cong_ty()
    except ConnectionError:
        return pd.DataFrame(columns=_COT)
    try:
        allvals = sheet.get_all_values()
    except Exception as e:
        _xoa_cache_va_thong_bao(e)
        return pd.DataFrame(columns=_COT)
    if len(allvals) <= 1:
        return pd.DataFrame(columns=_COT)
    # Dùng header thực tế từ sheet
    actual_headers = [v.strip() for v in allvals[0]]
    rows = []
    for r in allvals[1:]:
        padded = (r + [""] * len(actual_headers))[:len(actual_headers)]
        if any(padded):
            rows.append(padded)
    df = pd.DataFrame(rows, columns=actual_headers)
    df.rename(columns={"Cong_Ty": "Tên Công Ty", "Created_Date": "Ngày Tạo"}, inplace=True)
    # Đảm bảo đủ cột theo thứ tự chuẩn
    for c in _COT:
        if c not in df.columns:
            df[c] = ""
    return df[_COT]


def them_cong_ty(ten_cong_ty: str, dia_chi: str = "", ma_kh: str = "", ma_so_thue: str = "") -> int:
    """
    Thêm công ty mới vào sheet 'Companies'.

    Args:
        ten_cong_ty: Tên công ty khách hàng
        dia_chi: Địa chỉ công ty
        ma_kh: Mã khách hàng
        ma_so_thue: Mã số thuế

    Returns:
        int: ID vừa tạo
    """
    sheet = lay_sheet_cong_ty()
    id_moi = len(sheet.col_values(1))  # cột A: header + các hàng dữ liệu
    ngay_tao = datetime.now().strftime("%Y-%m-%d")
    sheet.append_row([id_moi, ten_cong_ty, dia_chi, ma_kh, ma_so_thue, ngay_tao])
    lay_danh_sach_cong_ty.clear()
    lay_ten_cac_cong_ty.clear()
    return id_moi


def sua_cong_ty(record_id: str, ten_moi: str, dia_chi: str = "", ma_kh: str = "", ma_so_thue: str = ""):
    sheet = lay_sheet_cong_ty()
    o_tim = sheet.find(str(record_id), in_column=1)
    if o_tim is None:
        raise ValueError(f"Không tìm thấy ID={record_id}.")
    sheet.update_cell(o_tim.row, 2, ten_moi)
    sheet.update_cell(o_tim.row, 3, dia_chi)
    sheet.update_cell(o_tim.row, 4, ma_kh)
    sheet.update_cell(o_tim.row, 5, ma_so_thue)
    lay_danh_sach_cong_ty.clear()
    lay_ten_cac_cong_ty.clear()


def xoa_cong_ty(record_id: str):
    sheet = lay_sheet_cong_ty()
    o_tim = sheet.find(str(record_id), in_column=1)
    if o_tim is None:
        raise ValueError(f"Không tìm thấy ID={record_id}.")
    sheet.delete_rows(o_tim.row)
    lay_danh_sach_cong_ty.clear()
    lay_ten_cac_cong_ty.clear()


@st.cache_data(ttl=60)
def lay_ten_cac_cong_ty() -> list:
    """Trả về danh sách tên công ty để hiển thị trong selectbox."""
    df = lay_danh_sach_cong_ty()
    if df.empty:
        return []
    return df["Tên Công Ty"].dropna().tolist()


# ============================================================
# LOẠI MÁY
# ============================================================
def _lay_sheet_don_gian(ten_sheet: str, tieu_de: list) -> object:
    """Helper chung: mở worksheet (tự tạo nếu chưa có), dùng spreadsheet đã câu hình."""
    try:
        bang_tinh = _lay_bang_tinh()
        try:
            return bang_tinh.worksheet(ten_sheet)
        except gspread.exceptions.WorksheetNotFound:
            sheet = bang_tinh.add_worksheet(title=ten_sheet, rows=500, cols=len(tieu_de) + 1)
            sheet.append_row(tieu_de)
            return sheet
    except gspread.exceptions.WorksheetNotFound:
        raise
    except Exception as e:
        raise ConnectionError(_xoa_cache_va_thong_bao(e)) from e


def _lay_df_don_gian(ten_sheet: str, tieu_de: list) -> pd.DataFrame:
    """Helper chung: đọc toàn bộ dữ liệu về DataFrame. Trả về rỗng nếu mất mạng."""
    try:
        sheet   = _lay_sheet_don_gian(ten_sheet, tieu_de)
        allvals = sheet.get_all_values()
    except (ConnectionError, OSError, Exception) as e:
        _xoa_cache_va_thong_bao(e)
        return pd.DataFrame(columns=tieu_de)
    n = len(tieu_de)
    # Tự phục hồi header nếu bị xóa nhầm
    if allvals:
        first_row = [v.strip() for v in allvals[0][:n]]
        expected  = [v.strip() for v in tieu_de[:n]]
        if first_row != expected:
            try:
                sheet.insert_row(tieu_de, index=1)
                allvals = [tieu_de] + allvals
            except Exception:
                pass
    if len(allvals) <= 1:
        return pd.DataFrame(columns=tieu_de)
    headers = [v.strip() for v in allvals[0][:n]]
    rows = [r[:n] for r in allvals[1:] if any(r[:n])]
    return pd.DataFrame(rows, columns=headers)


def _them_hang_don_gian(ten_sheet: str, tieu_de: list, hang: list) -> int:
    """Helper chung: thêm 1 hàng, trả về ID mới. Raise ConnectionError nếu mất mạng."""
    sheet  = _lay_sheet_don_gian(ten_sheet, tieu_de)
    id_moi = len(sheet.col_values(1))  # số hàng hiện tại (kể cả header) = ID tiếp theo
    sheet.append_row([id_moi] + hang)
    return id_moi


def _sua_hang_don_gian(ten_sheet: str, tieu_de: list, record_id: str, ten_cot: str, gia_tri_moi: str):
    """Helper chung: tìm hàng theo ID (cột 1) rồi cập nhật cột ten_cot."""
    sheet = _lay_sheet_don_gian(ten_sheet, tieu_de)
    o_tim = sheet.find(str(record_id), in_column=1)
    if o_tim is None:
        raise ValueError(f"Không tìm thấy ID={record_id} trong sheet {ten_sheet}.")
    col_idx = tieu_de.index(ten_cot) + 1
    sheet.update_cell(o_tim.row, col_idx, gia_tri_moi)


def _xoa_hang_don_gian(ten_sheet: str, tieu_de: list, record_id: str):
    """Helper chung: tìm hàng theo ID (cột 1) rồi xóa."""
    if not str(record_id).strip():
        raise ValueError("ID rỗng — không thể xóa. Vui lòng kiểm tra dữ liệu trong sheet.")
    sheet = _lay_sheet_don_gian(ten_sheet, tieu_de)
    o_tim = sheet.find(str(record_id), in_column=1)
    if o_tim is None:
        raise ValueError(f"Không tìm thấy ID={record_id} trong sheet {ten_sheet}.")
    # Bảo vệ: không xóa header row
    if o_tim.row <= 1:
        raise ValueError(f"ID={record_id} trùng với header — bỏ qua để tránh mất dữ liệu.")
    sheet.delete_rows(o_tim.row)


# ============================================================
# HỆ THỐNG THÔNG BÁO (NOTIFICATIONS)
# ============================================================
_TB_SHEET   = "Notifications"
_TB_HEADERS = ["ID", "Nguoi_Nhan", "Noi_Dung", "Task_ID", "Loai", "Thoi_Gian", "Da_Doc"]


def them_thong_bao(nguoi_nhan: str, noi_dung: str, task_id: int = 0, loai: str = "general"):
    """Thêm 1 thông báo cho người nhận cụ thể vào sheet Notifications."""
    try:
        sheet     = _lay_sheet_don_gian(_TB_SHEET, _TB_HEADERS)
        id_moi    = len(sheet.col_values(1))
        thoi_gian = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        sheet.append_row([id_moi, nguoi_nhan, noi_dung, str(task_id), loai, thoi_gian, "0"])
    except Exception:
        pass


def them_thong_bao_tat_ca(noi_dung: str, task_id: int = 0, loai: str = "general", tru_nguoi: str = ""):
    """Gửi thông báo cho tất cả nhân viên trong hệ thống (trừ người tạo nếu có).
    Dùng append_rows() để ghi một lần thay vì N lần riêng lẻ.
    """
    try:
        df_users = lay_danh_sach_users()
        sheet = _lay_sheet_don_gian(_TB_SHEET, _TB_HEADERS)
        id_bat_dau = len(sheet.col_values(1))  # 1 API call để lấy offset ID
        thoi_gian = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        hang_moi = []
        for i, (_, row) in enumerate(df_users.iterrows()):
            ten = str(row.get("HoTen", "")).strip()
            if ten and ten != tru_nguoi:
                hang_moi.append([id_bat_dau + len(hang_moi), ten, noi_dung, str(task_id), loai, thoi_gian, "0"])
        if hang_moi:
            sheet.append_rows(hang_moi, value_input_option="RAW")
    except Exception:
        pass


def lay_thong_bao_nguoi_dung(ho_ten: str) -> pd.DataFrame:
    """Lấy danh sách thông báo của user trong 2 tuần gần nhất, mới nhất trước."""
    try:
        df = _lay_df_don_gian(_TB_SHEET, _TB_HEADERS)
        if df.empty:
            return df
        df = df[df["Nguoi_Nhan"] == ho_ten].copy()
        # Lọc chỉ thông báo trong 2 tuần gần nhất
        try:
            df["Thoi_Gian"] = pd.to_datetime(df["Thoi_Gian"], errors="coerce")
            nguong = datetime.now() - timedelta(weeks=2)
            df = df[df["Thoi_Gian"] >= nguong]
            df["Thoi_Gian"] = df["Thoi_Gian"].dt.strftime("%Y-%m-%d %H:%M:%S")
        except Exception:
            pass
        return df.sort_values("Thoi_Gian", ascending=False).reset_index(drop=True)
    except Exception:
        return pd.DataFrame(columns=_TB_HEADERS)


def xoa_thong_bao_cu():
    """Xóa các thông báo cũ hơn 2 tuần khỏi sheet (chạy định kỳ)."""
    try:
        sheet = _lay_sheet_don_gian(_TB_SHEET, _TB_HEADERS)
        rows = sheet.get_all_values()
        if len(rows) <= 1:
            return
        headers = rows[0]
        idx_time = headers.index("Thoi_Gian")
        nguong = datetime.now() - timedelta(weeks=2)
        # Tìm các hàng cần xóa (từ dưới lên để không lệch index)
        rows_to_delete = []
        for i, row in enumerate(rows[1:], start=2):
            try:
                tg = datetime.strptime(row[idx_time], "%Y-%m-%d %H:%M:%S")
                if tg < nguong:
                    rows_to_delete.append(i)
            except Exception:
                pass
        # Xóa từ dưới lên
        for r in reversed(rows_to_delete):
            sheet.delete_rows(r)
    except Exception:
        pass


def dem_chua_doc(ho_ten: str) -> int:
    """Đếm số thông báo chưa đọc của user."""
    try:
        df = lay_thong_bao_nguoi_dung(ho_ten)
        if df.empty:
            return 0
        return int((df["Da_Doc"] == "0").sum())
    except Exception:
        return 0


def danh_dau_da_doc_tat_ca(ho_ten: str):
    """Đánh dấu tất cả thông báo của user là đã đọc (batch update)."""
    try:
        sheet = _lay_sheet_don_gian(_TB_SHEET, _TB_HEADERS)
        rows  = sheet.get_all_values()
        if len(rows) <= 1:
            return
        headers   = rows[0]
        idx_nhan  = headers.index("Nguoi_Nhan") + 1
        idx_doc   = headers.index("Da_Doc") + 1
        col_letter = chr(64 + idx_doc)
        batch = []
        for i, row in enumerate(rows[1:], start=2):
            nhan = row[idx_nhan - 1] if len(row) >= idx_nhan else ""
            doc  = row[idx_doc  - 1] if len(row) >= idx_doc  else "0"
            if nhan == ho_ten and doc == "0":
                batch.append({"range": f"{col_letter}{i}", "values": [["1"]]})
        if batch:
            sheet.batch_update(batch)
    except Exception:
        pass


# Danh sách trạng thái mặc định (theo quy trình thực tế)
_DS_TRANG_THAI_MAC_DINH = [
    "Đang Kiểm Tra",
    "Đã Phê Duyệt",
    "Đã Báo Giá",
    "Có Đơn",
    "Chờ Giao",
    "Đã Hoàn Thành - Giao Máy",
]

# ---- Loại Máy ----
_TIEUDE_LOAI_MAY = ["ID", "Tên Loại Máy", "Ngày Tạo"]

@st.cache_data(ttl=60)
def lay_danh_sach_loai_may() -> pd.DataFrame:
    return _lay_df_don_gian("LoaiMay", _TIEUDE_LOAI_MAY)

def them_loai_may(ten: str) -> int:
    result = _them_hang_don_gian("LoaiMay", _TIEUDE_LOAI_MAY,
                                 [ten, datetime.now().strftime("%Y-%m-%d")])
    lay_danh_sach_loai_may.clear()
    lay_ten_cac_loai_may.clear()
    return result

def sua_loai_may(record_id: str, ten_moi: str):
    _sua_hang_don_gian("LoaiMay", _TIEUDE_LOAI_MAY, record_id, "Tên Loại Máy", ten_moi)
    lay_danh_sach_loai_may.clear()
    lay_ten_cac_loai_may.clear()

def xoa_loai_may(record_id: str):
    _xoa_hang_don_gian("LoaiMay", _TIEUDE_LOAI_MAY, record_id)
    lay_danh_sach_loai_may.clear()
    lay_ten_cac_loai_may.clear()

@st.cache_data(ttl=60)
def lay_ten_cac_loai_may() -> list:
    df = lay_danh_sach_loai_may()
    return df["Tên Loại Máy"].dropna().tolist() if not df.empty else []


# ---- Tình Trạng ----
_TIEUDE_TINH_TRANG = ["ID", "Tên Tình Trạng", "Ngày Tạo"]

@st.cache_data(ttl=60)
def lay_danh_sach_tinh_trang() -> pd.DataFrame:
    return _lay_df_don_gian("TinhTrang", _TIEUDE_TINH_TRANG)

def them_tinh_trang(ten: str) -> int:
    result = _them_hang_don_gian("TinhTrang", _TIEUDE_TINH_TRANG,
                                 [ten, datetime.now().strftime("%Y-%m-%d")])
    lay_danh_sach_tinh_trang.clear()
    lay_ten_cac_tinh_trang.clear()
    return result

def sua_tinh_trang(record_id: str, ten_moi: str):
    _sua_hang_don_gian("TinhTrang", _TIEUDE_TINH_TRANG, record_id, "Tên Tình Trạng", ten_moi)
    lay_danh_sach_tinh_trang.clear()
    lay_ten_cac_tinh_trang.clear()

def xoa_tinh_trang(record_id: str):
    _xoa_hang_don_gian("TinhTrang", _TIEUDE_TINH_TRANG, record_id)
    lay_danh_sach_tinh_trang.clear()
    lay_ten_cac_tinh_trang.clear()

@st.cache_data(ttl=60)
def lay_ten_cac_tinh_trang() -> list:
    df = lay_danh_sach_tinh_trang()
    return df["Tên Tình Trạng"].dropna().tolist() if not df.empty else []


# ---- Trạng Thái Công Việc (manual list) ----
_TIEUDE_TRANG_THAI = ["ID", "Tên Trạng Thái", "Ngày Tạo"]

@st.cache_data(ttl=60)
def lay_danh_sach_trang_thai_custom() -> pd.DataFrame:
    return _lay_df_don_gian("TrangThai", _TIEUDE_TRANG_THAI)

def them_trang_thai_custom(ten: str) -> int:
    result = _them_hang_don_gian("TrangThai", _TIEUDE_TRANG_THAI,
                                 [ten, datetime.now().strftime("%Y-%m-%d")])
    lay_danh_sach_trang_thai_custom.clear()
    lay_ten_cac_trang_thai.clear()
    return result

def sua_trang_thai_custom(record_id: str, ten_moi: str):
    _sua_hang_don_gian("TrangThai", _TIEUDE_TRANG_THAI, record_id, "Tên Trạng Thái", ten_moi)
    lay_danh_sach_trang_thai_custom.clear()
    lay_ten_cac_trang_thai.clear()

def xoa_trang_thai_custom(record_id: str):
    _xoa_hang_don_gian("TrangThai", _TIEUDE_TRANG_THAI, record_id)
    lay_danh_sach_trang_thai_custom.clear()
    lay_ten_cac_trang_thai.clear()

def _seed_trang_thai_mac_dinh():
    """Seed các trạng thái mặc định vào TrangThai sheet nếu còn trống."""
    df = lay_danh_sach_trang_thai_custom()
    if df.empty:
        for ten in _DS_TRANG_THAI_MAC_DINH:
            them_trang_thai_custom(ten)

@st.cache_data(ttl=60)
def lay_ten_cac_trang_thai() -> list:
    df = lay_danh_sach_trang_thai_custom()
    return df["Tên Trạng Thái"].dropna().tolist() if not df.empty else _DS_TRANG_THAI_MAC_DINH


# ---- Công Đoạn ----
_TIEUDE_CONG_DOAN = ["ID", "Tên Công Đoạn", "Ngày Tạo"]

@st.cache_data(ttl=60)
def lay_danh_sach_cong_doan() -> pd.DataFrame:
    return _lay_df_don_gian("CongDoan", _TIEUDE_CONG_DOAN)

def them_cong_doan(ten: str) -> int:
    result = _them_hang_don_gian("CongDoan", _TIEUDE_CONG_DOAN,
                                 [ten, datetime.now().strftime("%Y-%m-%d")])
    lay_danh_sach_cong_doan.clear()
    lay_ten_cac_cong_doan.clear()
    return result

def sua_cong_doan(record_id: str, ten_moi: str):
    _sua_hang_don_gian("CongDoan", _TIEUDE_CONG_DOAN, record_id, "Tên Công Đoạn", ten_moi)
    lay_danh_sach_cong_doan.clear()
    lay_ten_cac_cong_doan.clear()

def xoa_cong_doan_item(record_id: str):
    _xoa_hang_don_gian("CongDoan", _TIEUDE_CONG_DOAN, record_id)
    lay_danh_sach_cong_doan.clear()
    lay_ten_cac_cong_doan.clear()

@st.cache_data(ttl=60)
def lay_ten_cac_cong_doan() -> list:
    df = lay_danh_sach_cong_doan()
    return df["Tên Công Đoạn"].dropna().tolist() if not df.empty else []


# ---- Người Thực Hiện Công Đoạn ----
_TIEUDE_NGUOI_CD = ["ID", "Họ Tên", "Công Đoạn", "Ngày Tạo"]

@st.cache_data(ttl=60)
def lay_danh_sach_nguoi_cong_doan() -> pd.DataFrame:
    return _lay_df_don_gian("NguoiCongDoan", _TIEUDE_NGUOI_CD)

def them_nguoi_cong_doan(ho_ten: str, cong_doan: str) -> int:
    result = _them_hang_don_gian("NguoiCongDoan", _TIEUDE_NGUOI_CD,
                                 [ho_ten, cong_doan, datetime.now().strftime("%Y-%m-%d")])
    lay_danh_sach_nguoi_cong_doan.clear()
    return result


_TASK_HEADERS = [
    "ID", "Công Ty", "Công Số", "Năm", "Tên Công Việc", "Mô Tả",
    "Nhân Viên", "Trạng Thái", "Ngày Tạo", "Hạn Hoàn Thành",
    "Link Ảnh", "Người Phê Duyệt", "Checklist", "Công Việc Con", "Công Đoạn",
    "Loại Máy", "Tình Trạng",
    "Công Suất", "Số Cực", "Mã Số", "Số PO Nội Bộ", "Số PO KH/HĐ", "Số Báo Giá",
    "Ngày Kết Thúc", "Ảnh Đo Lường", "Ngày Xóa"
]


@st.cache_data(ttl=60)
def _fetch_tasks_cached() -> pd.DataFrame:
    """Fetch raw tasks from Google Sheets — raises on error (not cached on failure)."""
    _N = len(_TASK_HEADERS)
    sheet   = lay_sheet()           # raises on failure → not cached
    allvals = sheet.get_all_values()  # raises on failure → not cached
    if len(allvals) <= 1:
        raise RuntimeError("Tasks sheet returned no data — possible stale connection")
    rows = [r[:_N] for r in allvals[1:] if any(r[:_N])]
    if not rows:
        raise RuntimeError("Tasks sheet has header but no data rows — possible stale connection")
    rows_padded = [r + [""] * (_N - len(r)) for r in rows]
    try:
        df = pd.DataFrame(rows_padded, columns=_TASK_HEADERS) if rows_padded else pd.DataFrame(columns=_TASK_HEADERS)
    except Exception:
        df = pd.DataFrame([dict(zip(_TASK_HEADERS, r)) for r in rows_padded]) if rows_padded else pd.DataFrame(columns=_TASK_HEADERS)
    df.rename(columns={
        "Cong_Ty": "Công Ty", "Cong_So": "Công Số", "Nam": "Năm",
        "Task_Name": "Tên Công Việc", "Description": "Mô Tả",
        "Assigned_To": "Nhân Viên", "Status": "Trạng Thái",
        "Created_Date": "Ngày Tạo", "Deadline": "Hạn Hoàn Thành",
        "Image_URL": "Link Ảnh",
    }, inplace=True)
    if "Trạng Thái" in df.columns:
        df["Trạng Thái"] = df["Trạng Thái"].replace({
            "Todo": "Chờ Làm", "Doing": "Đang Làm", "Done": "Hoàn Thành",
        })
    for col in ["Công Ty", "Công Số", "Năm", "Người Phê Duyệt", "Checklist", "Công Việc Con"]:
        if col not in df.columns:
            df[col] = ""
    return df


def _lay_tat_ca_tasks() -> pd.DataFrame:
    """Lấy TẤT CẢ task (kể cả đã xóa mềm) — dùng nội bộ."""
    try:
        return _fetch_tasks_cached()
    except Exception as _e:
        import traceback
        print(f"[_lay_tat_ca_tasks lần 1 lỗi] {_e}")
        try:
            lay_sheet.clear()
            _lay_bang_tinh.clear()
            _fetch_tasks_cached.clear()
            return _fetch_tasks_cached()
        except Exception as _e2:
            print(f"[_lay_tat_ca_tasks lần 2 lỗi] {_e2}\n{traceback.format_exc()}")
            return pd.DataFrame(columns=_TASK_HEADERS)


def lay_danh_sach_cong_viec() -> pd.DataFrame:
    """
    Lấy danh sách công việc CHƯA bị xóa từ sheet 'Tasks'.
    Tasks đã xóa mềm (có Ngày Xóa) được lọc ra.
    """
    df = _lay_tat_ca_tasks()
    if "Ngày Xóa" in df.columns:
        df = df[df["Ngày Xóa"].fillna("").astype(str).str.strip() == ""].copy()
    return df


# Cho phép gọi lay_danh_sach_cong_viec.clear() như trước
lay_danh_sach_cong_viec.clear = _fetch_tasks_cached.clear


def them_cong_viec(ten_task: str, mo_ta: str, nguoi_duoc_giao: str, deadline: str,
                   cong_ty: str = "", cong_so: str = "", nam: str = "",
                   trang_thai: str = "Chờ Làm", nguoi_phe_duyet: str = "",
                   checklist: list = None, cong_viec_con: list = None,
                   cong_doan: str = "", loai_may: str = "", tinh_trang: str = "",
                   cong_suat: str = "", so_cuc: str = "", ma_so: str = "",
                   so_po_noi_bo: str = "", so_po_kh: str = "", so_bao_gia: str = "",
                   ngay_ket_thuc: str = "") -> int:
    """
    Thêm một công việc mới vào cuối sheet.

    Cấu trúc hàng:
    A:ID | B:Cong_Ty | C:Cong_So | D:Nam | E:Task_Name | F:Description
    G:Assigned_To | H:Status | I:Created_Date | J:Deadline | K:Image_URL
    L:Nguoi_Phe_Duyet | M:Checklist (JSON) | N:Cong_Viec_Con (JSON) | O:Cong_Doan
    P:Loai_May | Q:Tinh_Trang | R:Cong_Suat | S:So_Cuc | T:Ma_So
    U:So_PO_Noi_Bo | V:So_PO_KH | W:So_Bao_Gia
    """
    sheet = _lay_sheet_fresh()  # Luôn dùng connection mới cho thao tác ghi
    # Đọc col A 1 lần duy nhất — dùng để tính cả ID mới lẫn vị trí hàng mới
    _col_a = sheet.col_values(1)          # 1 API call thay vì 2
    _ids_hien_co = _col_a[1:]            # bỏ header
    _so_ids = set()
    for x in _ids_hien_co:
        try:
            _so_ids.add(int(float(str(x).strip())))
        except (ValueError, TypeError):
            pass
    id_moi = (max(_so_ids) + 1) if _so_ids else 1
    # Đảm bảo ID chưa tồn tại (tránh race condition)
    while id_moi in _so_ids:
        id_moi += 1
    _dong_moi = len(_col_a) + 1          # row tiếp theo
    ngay_tao = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    hang_moi = [
        id_moi,                           # A: ID
        cong_ty,                          # B: Công Ty
        cong_so,                          # C: Công Số
        nam,                              # D: Năm
        ten_task,                         # E: Tên Công Việc
        mo_ta,                            # F: Mô Tả
        nguoi_duoc_giao,                  # G: Nhân Viên
        trang_thai,                       # H: Trạng Thái
        ngay_tao,                         # I: Ngày Tạo
        deadline,                         # J: Hạn Hoàn Thành
        "",                               # K: Link Ảnh (để trống)
        nguoi_phe_duyet,                  # L: Người Phê Duyệt
        json.dumps(checklist or [], ensure_ascii=False),      # M: Checklist
        json.dumps(cong_viec_con or [], ensure_ascii=False),  # N: Công Việc Con
        cong_doan,                                            # O: Công Đoạn
        loai_may,                                             # P: Loại Máy
        tinh_trang,                                           # Q: Tình Trạng
        cong_suat,                                            # R: Công Suất
        so_cuc,                                               # S: Số Cực
        ma_so,                                                # T: Mã Số
        so_po_noi_bo,                                         # U: Số PO Nội Bộ
        so_po_kh,                                             # V: Số PO KH/HĐ
        so_bao_gia,                                           # W: Số Báo Giá
        ngay_ket_thuc,                                        # X: Ngày Kết Thúc
        "{}",                                                 # Y: Ảnh Đo Lường (JSON)
    ]

    sheet.update(f"A{_dong_moi}:Y{_dong_moi}", [hang_moi])
    # Chỉ clear data cache — KHÔNG clear lay_sheet/_lay_bang_tinh để tránh
    # force reconnect làm chậm hoặc lỗi lần đọc tiếp theo
    lay_danh_sach_cong_viec.clear()

    # ── Gửi thông báo cho tất cả người dùng ──────────────────────────────
    _tb_task = (
        f"📋 Task mới #{id_moi}: {ten_task}"
        f" | Giao: {nguoi_duoc_giao}"
        + (f" | Công ty: {cong_ty}" if cong_ty else "")
    )
    them_thong_bao_tat_ca(_tb_task, task_id=id_moi, loai="task_moi")
    # Thông báo riêng cho từng người được giao công việc con
    for _cv in (cong_viec_con or []):
        _nv_cv = str(_cv.get("nhan_vien", "")).strip()
        _ten_cv = str(_cv.get("ten", "")).strip()
        if _nv_cv:
            them_thong_bao(
                _nv_cv,
                f"🔧 Bạn được giao việc: {_ten_cv} (Task #{id_moi}: {ten_task})",
                task_id=id_moi,
                loai="cong_viec_con",
            )

    return id_moi


def xoa_cong_viec(task_id: int):
    """Xóa mềm — chuyển task vào thùng rác (đặt Ngày Xóa = hôm nay)."""
    try:
        ngay_xoa = datetime.now().strftime("%Y-%m-%d")
        cap_nhat_nhieu_truong_task(int(task_id), {"Ngày Xóa": ngay_xoa})
    except Exception:
        pass


def lay_cong_viec_da_xoa() -> pd.DataFrame:
    """Lấy các task đang ở thùng rác (có Ngày Xóa)."""
    df = _lay_tat_ca_tasks()
    if "Ngày Xóa" not in df.columns:
        return pd.DataFrame(columns=_TASK_HEADERS)
    return df[df["Ngày Xóa"].fillna("").astype(str).str.strip() != ""].copy()


def khoi_phuc_cong_viec(task_id: int):
    """Khôi phục task từ thùng rác (xóa Ngày Xóa)."""
    cap_nhat_nhieu_truong_task(int(task_id), {"Ngày Xóa": ""})


def xoa_vinh_vien_cong_viec(task_id: int):
    """Xóa vĩnh viễn task khỏi sheet (không thể khôi phục)."""
    try:
        sheet = _lay_sheet_fresh()
        o_tim = sheet.find(str(task_id), in_column=1)
        if o_tim:
            sheet.delete_rows(o_tim.row)
            _fetch_tasks_cached.clear()
            _ROW_CACHE.pop(int(task_id), None)
    except Exception:
        pass


def don_dep_thung_rac():
    """Xóa vĩnh viễn các task đã ở thùng rác > 7 ngày (chạy tự động khi khởi động)."""
    import datetime as _dt
    try:
        df_xoa = lay_cong_viec_da_xoa()
        if df_xoa.empty:
            return
        nguong = _dt.date.today() - _dt.timedelta(days=7)
        to_delete = []
        for _, row in df_xoa.iterrows():
            ngay_xoa_str = str(row.get("Ngày Xóa", "")).strip()
            if not ngay_xoa_str:
                continue
            try:
                if _dt.date.fromisoformat(ngay_xoa_str[:10]) <= nguong:
                    to_delete.append(int(row["ID"]))
            except Exception:
                pass
        for tid in to_delete:
            xoa_vinh_vien_cong_viec(tid)
        if to_delete:
            _fetch_tasks_cached.clear()
    except Exception:
        pass


@st.dialog("🗑️ Xác nhận xóa công việc")
def _dialog_xac_nhan_xoa(task_id, ten_cv: str):
    try:
        task_id = int(float(task_id))
    except (ValueError, TypeError):
        st.error("❌ Không xác định được ID công việc.")
        return
    st.warning(f"Chuyển công việc vào **Thùng Rác**:\n\n**{ten_cv}**?", icon="🗑️")
    st.caption("Công việc sẽ tự xóa vĩnh viễn sau **7 ngày** nếu không khôi phục.")
    col_ok, col_cancel = st.columns(2)
    with col_ok:
        if st.button("🗑️ Vào thùng rác", type="primary", use_container_width=True):
            with st.spinner("Đang xử lý..."):
                xoa_cong_viec(task_id)
            lay_danh_sach_cong_viec.clear()
            st.rerun(scope="app")
    with col_cancel:
        if st.button("Hủy", use_container_width=True):
            st.rerun(scope="app")


def cap_nhat_trang_thai(task_id: int, trang_thai_moi: str):
    """
    Cập nhật trạng thái (Status) của công việc theo ID.
    """
    try:
        sheet = _lay_sheet_fresh()
        o_tim = sheet.find(str(task_id), in_column=1)
        if o_tim:
            sheet.update_cell(o_tim.row, 8, trang_thai_moi)
            lay_danh_sach_cong_viec.clear()
    except Exception:
        pass


def cap_nhat_ngay_ket_thuc(task_id: int, ngay_ket_thuc: str):
    """
    Cập nhật Ngày Kết Thúc (cột X, cột 24) của công việc theo ID.
    """
    try:
        sheet = _lay_sheet_fresh()
        o_tim = sheet.find(str(task_id), in_column=1)
        if o_tim:
            sheet.update_cell(o_tim.row, 24, ngay_ket_thuc)
            lay_danh_sach_cong_viec.clear()
    except Exception:
        pass


def cap_nhat_han_hoan_thanh(task_id: int, han: str):
    """
    Cập nhật Hạn Hoàn Thành (cột J, cột 10) của công việc theo ID.
    """
    try:
        sheet = _lay_sheet_fresh()
        o_tim = sheet.find(str(task_id), in_column=1)
        if o_tim:
            sheet.update_cell(o_tim.row, 10, han)
            lay_danh_sach_cong_viec.clear()
    except Exception:
        pass


def doc_anh_do_luong(gia_tri: str) -> dict:
    """Parse JSON ảnh đo lường từ cột Y. Trả về dict {label: [url,...]}."""
    if not gia_tri or str(gia_tri).strip() in ("", "{}"):
        return {}
    try:
        return json.loads(str(gia_tri).strip())
    except Exception:
        return {}


# Cache row number theo task_id — row không thay đổi nên dùng module-level dict
_ROW_CACHE: dict[int, int] = {}


def cap_nhat_anh_do_luong(task_id: int, anh_dict: dict):
    """Ghi toàn bộ dict ảnh đo lường (cột Y = col 25) lên Google Sheets — background thread."""
    payload = json.dumps(anh_dict, ensure_ascii=False)
    task_id = int(task_id)

    def _write():
        try:
            sheet = lay_sheet()  # dùng cached connection, không tạo mới mỗi lần
            row = _ROW_CACHE.get(task_id)
            if not row:
                o_tim = sheet.find(str(task_id), in_column=1)
                if not o_tim:
                    return
                row = o_tim.row
                _ROW_CACHE[task_id] = row
            sheet.update_cell(row, 25, payload)
            lay_danh_sach_cong_viec.clear()
        except Exception:
            # Nếu cached sheet lỗi, thử lại với fresh connection
            try:
                sheet = _lay_sheet_fresh()
                _ROW_CACHE.pop(task_id, None)  # xóa cache để find lại
                o_tim = sheet.find(str(task_id), in_column=1)
                if o_tim:
                    _ROW_CACHE[task_id] = o_tim.row
                    sheet.update_cell(o_tim.row, 25, payload)
                    lay_danh_sach_cong_viec.clear()
            except Exception:
                pass

    threading.Thread(target=_write, daemon=False).start()


def doc_danh_sach_anh(gia_tri: str) -> list:
    """
    Parse danh sách URL ảnh từ chuỗi lưu trong Google Sheets.
    Hỗ trợ cả dạng JSON array '["url1","url2"]' và URL đơn cũ.
    """
    if not gia_tri:
        return []
    gia_tri = str(gia_tri).strip()
    if gia_tri.startswith("["):
        try:
            danh_sach = json.loads(gia_tri)
            return [u for u in danh_sach if u]
        except Exception:
            return []
    return [gia_tri] if gia_tri else []


def cap_nhat_url_anh(task_id: int, url_anh: str):
    """
    Gắn thêm URL ảnh mới vào danh sách ảnh của task (lưu dạng JSON array).
    Cột K lưu JSON array URLs. Cột L lưu =IMAGE() của ảnh đầu tiên.

    Args:
        task_id: ID công việc cần cập nhật
        url_anh: URL ảnh mới trên Cloudinary
    """
    try:
        sheet = _lay_sheet_fresh()
        o_tim = sheet.find(str(task_id), in_column=1)
        if o_tim:
            so_hang = o_tim.row
            # Lấy danh sách ảnh hiện có
            gia_tri_cu = sheet.cell(so_hang, 11).value or ""
            ds_anh = doc_danh_sach_anh(gia_tri_cu)
            if url_anh not in ds_anh:
                ds_anh.append(url_anh)
            # Lưu JSON array vào cột K (11)
            sheet.update_cell(so_hang, 11, json.dumps(ds_anh, ensure_ascii=False))
            # Cột L (12): =IMAGE() của ảnh đầu tiên để xem trong Google Sheets
            sheet.update(
                [[f'=IMAGE("{ds_anh[0]}")']],
                f"L{so_hang}",
                value_input_option="USER_ENTERED"
            )
            lay_danh_sach_cong_viec.clear()
    except Exception:
        pass


def xoa_url_anh(task_id: int, url_xoa: str):
    """
    Xoá một URL ảnh khỏi danh sách ảnh của task.
    """
    try:
        sheet = _lay_sheet_fresh()
        o_tim = sheet.find(str(task_id), in_column=1)
        if o_tim:
            so_hang = o_tim.row
            gia_tri_cu = sheet.cell(so_hang, 11).value or ""
            ds_anh = doc_danh_sach_anh(gia_tri_cu)
            ds_anh = [u for u in ds_anh if u != url_xoa]
            sheet.update_cell(so_hang, 11, json.dumps(ds_anh, ensure_ascii=False) if ds_anh else "")
            if ds_anh:
                sheet.update(
                    [[f'=IMAGE("{ds_anh[0]}")']],
                    f"L{so_hang}",
                    value_input_option="USER_ENTERED"
                )
            else:
                sheet.update_cell(so_hang, 12, "")
            lay_danh_sach_cong_viec.clear()
    except Exception:
        pass


def cap_nhat_checklist(task_id: int, checklist: list):
    """Lưu danh sách checklist (JSON) vào cột M của task."""
    try:
        sheet = _lay_sheet_fresh()
        o_tim = sheet.find(str(task_id), in_column=1)
        if o_tim:
            sheet.update_cell(o_tim.row, 13, json.dumps(checklist, ensure_ascii=False))
            lay_danh_sach_cong_viec.clear()
    except Exception:
        pass


# Ánh xạ tên field → số cột trong sheet Tasks
_COL_MAP_TASK = {
    "Công Ty":      2,   # B
    "Công Số":      3,   # C
    "Năm":          4,   # D
    "Tên Công Việc":5,   # E
    "Mô Tả":        6,   # F
    "Nhân Viên":    7,   # G
    "Trạng Thái":   8,   # H
    "Hạn Hoàn Thành":10, # J
    "Người Phê Duyệt":12,# L
    "Công Đoạn":    15,  # O
    "Loại Máy":     16,  # P
    "Tình Trạng":   17,  # Q
    "Công Suất":    18,  # R
    "Số Cực":       19,  # S
    "Mã Số":        20,  # T
    "Số PO Nội Bộ": 21,  # U
    "Số PO KH/HĐ":  22,  # V
    "Số Báo Giá":   23,  # W
    "Ngày Xóa":     26,  # Z
}


def cap_nhat_nhieu_truong_task(task_id: int, gia_tri: dict):
    """Cập nhật nhiều trường của task trong 1 lần kết nối.
    gia_tri: {tên_cột: giá_trị_mới}
    """
    try:
        sheet = _lay_sheet_fresh()
        o_tim = sheet.find(str(task_id), in_column=1) or sheet.find(f"{task_id}.0", in_column=1)
        if not o_tim:
            return
        # Chuẩn hóa ID về integer nếu đang là float
        if sheet.cell(o_tim.row, 1).value.endswith(".0"):
            sheet.update_cell(o_tim.row, 1, str(task_id))
        r = o_tim.row
        for ten_col, val in gia_tri.items():
            col_num = _COL_MAP_TASK.get(ten_col)
            if col_num:
                sheet.update_cell(r, col_num, str(val))
        lay_danh_sach_cong_viec.clear()
    except Exception:
        pass


# ============================================================
# QUẢN LÝ TÀI KHOẢN NGƯỜI DÙNG (USERS SHEET)
# ============================================================
_TIEUDE_USERS = ["ID", "Username", "Password", "HoTen", "NgaySinh", "VaiTro", "NgayTao"]


def _ma_hoa_mat_khau(mat_khau: str) -> str:
    """Hash mật khẩu bằng SHA-256."""
    return hashlib.sha256(mat_khau.encode("utf-8")).hexdigest()


@st.cache_resource
def lay_sheet_users():
    """Mở (hoặc tự tạo) worksheet 'Users' để lưu tài khoản."""
    try:
        bang_tinh = _lay_bang_tinh()
        try:
            sheet = bang_tinh.worksheet("Users")
        except gspread.exceptions.WorksheetNotFound:
            sheet = bang_tinh.add_worksheet(title="Users", rows=500, cols=len(_TIEUDE_USERS) + 1)
            sheet.append_row(_TIEUDE_USERS)
            # Tạo tài khoản admin mặc định ngay khi sheet được tạo lần đầu
            sheet.append_row([
                1,
                "admin",
                _ma_hoa_mat_khau("admin123"),
                "Quản Trị Viên",
                "",
                "admin",
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            ])
        return sheet
    except Exception as e:
        raise ConnectionError(_xoa_cache_va_thong_bao(e)) from e


@st.cache_data(ttl=30)
def _fetch_users_cached() -> pd.DataFrame:
    """Fetch users từ GSheets — raise on error (không cache kết quả lỗi)."""
    sheet = lay_sheet_users()
    allvals = sheet.get_all_values()
    if len(allvals) <= 1:
        raise RuntimeError("Users sheet trống")
    n = len(_TIEUDE_USERS)
    rows = [r[:n] for r in allvals[1:] if any(r[:n])]
    return pd.DataFrame(rows, columns=_TIEUDE_USERS)


def lay_danh_sach_users() -> pd.DataFrame:
    """Đọc danh sách tài khoản, tự retry nếu connection stale."""
    try:
        return _fetch_users_cached()
    except Exception:
        try:
            lay_sheet_users.clear()
            _fetch_users_cached.clear()
            return _fetch_users_cached()
        except Exception:
            return pd.DataFrame(columns=_TIEUDE_USERS)


lay_danh_sach_users.clear = _fetch_users_cached.clear


def kiem_tra_dang_nhap(username: str, mat_khau: str):
    """
    Kiểm tra thông tin đăng nhập.
    Trả về dict user nếu đúng, None nếu sai.
    """
    df = lay_danh_sach_users()
    if df.empty:
        return None
    pw_hash = _ma_hoa_mat_khau(mat_khau)
    row = df[(df["Username"].str.lower() == username.strip().lower()) &
             (df["Password"] == pw_hash)]
    if row.empty:
        return None
    r = row.iloc[0]
    return {"id": r["ID"], "username": r["Username"],
            "ho_ten": r["HoTen"], "ngay_sinh": r["NgaySinh"],
            "vai_tro": r["VaiTro"]}


def dang_ky_tai_khoan(username: str, mat_khau: str, ho_ten: str,
                      ngay_sinh: str, vai_tro: str = "nhan_vien") -> tuple:
    """
    Đăng ký tài khoản mới.
    Trả về (True, 'OK') nếu thành công, (False, 'lý do') nếu thất bại.
    """
    username = username.strip()
    if not username or not mat_khau or not ho_ten:
        return False, "Vui lòng điền đầy đủ thông tin."
    if len(username) < 3:
        return False, "Username phải có ít nhất 3 ký tự."
    if len(mat_khau) < 6:
        return False, "Mật khẩu phải có ít nhất 6 ký tự."

    # Kiểm tra username đã tồn tại chưa
    df = lay_danh_sach_users()
    if not df.empty and username.lower() in df["Username"].str.lower().tolist():
        return False, "Username đã tồn tại. Vui lòng chọn tên khác."

    sheet = lay_sheet_users()
    id_moi = len(sheet.col_values(1))  # header + data rows
    sheet.append_row([
        id_moi,
        username,
        _ma_hoa_mat_khau(mat_khau),
        ho_ten.strip(),
        ngay_sinh,
        vai_tro,
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    ])
    # Xóa cache để đọc lại danh sách mới nhất
    lay_danh_sach_users.clear()
    lay_danh_sach_nhan_vien.clear()
    return True, "OK"


def doi_mat_khau(username: str, mat_khau_cu: str, mat_khau_moi: str) -> tuple:
    """
    Đổi mật khẩu cho tài khoản.
    Trả về (True, 'OK') nếu thành công, (False, 'lý do') nếu thất bại.
    """
    if not mat_khau_moi:
        return False, "Vui lòng nhập mật khẩu mới."
    if len(mat_khau_moi) < 6:
        return False, "Mật khẩu mới phải có ít nhất 6 ký tự."

    df = lay_danh_sach_users()
    if df.empty:
        return False, "Không tìm thấy tài khoản."

    pw_hash_cu = _ma_hoa_mat_khau(mat_khau_cu)
    row = df[(df["Username"].str.lower() == username.strip().lower()) &
             (df["Password"] == pw_hash_cu)]
    if row.empty:
        return False, "Mật khẩu hiện tại không đúng."

    sheet = lay_sheet_users()
    o_tim = sheet.find(username.strip(), in_column=2)
    if o_tim is None:
        return False, "Không tìm thấy tài khoản trong hệ thống."

    sheet.update_cell(o_tim.row, 3, _ma_hoa_mat_khau(mat_khau_moi))
    lay_danh_sach_users.clear()
    return True, "OK"


@st.cache_data(ttl=60)
def lay_danh_sach_nhan_vien() -> list:
    """Trả về danh sách Họ Tên nhân viên (đọc từ Users sheet, vai trò nhan_vien)."""
    try:
        df = lay_danh_sach_users()
        if df.empty:
            return []
        # Lấy tất cả người dùng (kể cả admin) để có thể giao việc
        nv = df[df["VaiTro"] == "nhan_vien"]["HoTen"].dropna().tolist()
        return [n for n in nv if n.strip()]
    except Exception:
        return []


# ============================================================
# TẠO PDF BIÊN BẢN NGHIỆM THU
# ============================================================
class PDFNghiemThu(FPDF):
    """
    Class PDF tùy chỉnh, kế thừa FPDF.
    Dùng font DejaVuSans để hiển thị đầy đủ tiếng Việt.
    Header/Footer trống để tự quản lý layout theo mẫu.
    """

    def _add_fonts(self):
        """Nạp font Unicode vào PDF (gọi 1 lần sau khi khởi tạo)."""
        self.add_font("DejaVu", "",  FONT_REGULAR)
        self.add_font("DejaVu", "B", FONT_BOLD)

    def header(self):
        """Không dùng header mặc định — layout được quản lý thủ công."""
        pass

    def footer(self):
        """Số trang nhỏ ở chân trang."""
        self.set_y(-12)
        self.set_font("DejaVu", "", 8)
        self.cell(0, 8, f"Trang {self.page_no()}", align="C")


def _tai_anh_tam(url: str):
    """Tải ảnh từ URL về file tạm, trả về đường dẫn hoặc None nếu lỗi."""
    try:
        r = requests.get(url, timeout=20)
        if r.status_code != 200:
            return None
        suffix = ".png" if url.lower().endswith(".png") else ".jpg"
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as f:
            f.write(r.content)
            return f.name
    except Exception:
        return None


def tao_pdf_nghiem_thu(thong_tin_task: dict) -> bytes:
    """
    Tạo file PDF biên bản nghiệm thu theo mẫu Điện Cơ Ngọc Trâm.

    Trang 1 : Header, tiêu đề BBNT, Engine/Customer/Address,
              Thời gian–địa điểm, bảng 12 hạng mục, footer doc.
    Trang 2 : Header doc, 3 bảng điện trở, mỗi bảng có ô ảnh
              (ảnh 1-3 → Stator1, 4-6 → Stator2, 7-9 → Rotor).
    """
    M_LEFT   = 10
    M_TOP    = 10
    USABLE_W = 190

    # ── Parse dữ liệu ──────────────────────────────────────────
    ten_dong_co  = str(thong_tin_task.get("Tên Công Việc", ""))
    khach_hang   = str(thong_tin_task.get("Công Ty", ""))
    cong_so      = str(thong_tin_task.get("Công Số", ""))
    mo_ta        = str(thong_tin_task.get("Mô Tả", ""))
    nhan_vien    = str(thong_tin_task.get("Nhân Viên", ""))
    ngay_tao_str = str(thong_tin_task.get("Ngày Tạo", ""))

    # Ảnh đo lường theo từng nhãn  {"U1–V1": [url,...], ...}
    anh_do_luong = doc_anh_do_luong(str(thong_tin_task.get("Ảnh Đo Lường", "") or ""))

    # Format ngày
    try:
        dt = datetime.strptime(ngay_tao_str[:10], "%Y-%m-%d")
        ngay_en = dt.strftime("%-d %B %Y") if os.name != "nt" else dt.strftime("%d %B %Y").lstrip("0")
        ngay_vi = f"Ngày {dt.day:02d} Tháng {dt.month:02d} Năm {dt.year}"
    except Exception:
        ngay_en = ngay_tao_str
        ngay_vi = ngay_tao_str

    hang_muc = [h.strip() for h in mo_ta.split("\n") if h.strip()]

    # ── Màu sắc ────────────────────────────────────────────────
    BLUE_HDR  = (68,  114, 196)   # tiêu đề bảng (đậm xanh)
    BLUE_CELL = (189, 215, 238)   # ô label pha (xanh nhạt)
    WHITE     = (255, 255, 255)
    BLACK     = (0,   0,   0)
    PURPLE    = (112, 48,  160)   # màu chữ công ty

    # ── Khởi tạo PDF ───────────────────────────────────────────
    pdf = PDFNghiemThu()
    pdf._add_fonts()
    pdf.set_margins(M_LEFT, M_TOP)
    pdf.set_auto_page_break(auto=False)

    # ========================================================= #
    #  TRANG 1                                                  #
    # ========================================================= #
    pdf.add_page()
    W = USABLE_W

    # ── Header công ty ─────────────────────────────────────────
    pdf.set_text_color(*PURPLE)
    pdf.set_font("DejaVu", "B", 13)
    pdf.cell(W, 7, "CÔNG TY TNHH MỘT THÀNH VIÊN ĐIỆN CƠ NGỌC TRÂM",
             align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.set_text_color(*BLACK)
    pdf.set_font("DejaVu", "", 8.5)
    pdf.cell(W, 5, "Địa chỉ : 8/5, hẻm 4, tổ 9, Khu phố Kim Sơn, Phường Long Thành, Thành phố Đồng Nai",
             align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(W, 5, "Website: ngoctrammotor.com   Mail: kd@ngoctrammotor.com",
             align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(W, 5, "MST: 3603238978  DT: 0907 042 043 (Mr.Hiệp) – 0908 062 291 ( Ms.Linh)",
             align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)

    # ── Tiêu đề BBNT ──────────────────────────────────────────
    pdf.set_font("DejaVu", "B", 12)
    pdf.cell(W, 7, "REPAIR ACCEPTANCE CERTIFICATE",
             align="C", new_x="LMARGIN", new_y="NEXT")

    # "BIÊN BẢN NGHIỆM THU" trong khung viền xanh
    y_box = pdf.get_y()
    pdf.set_draw_color(*BLUE_HDR)
    pdf.set_line_width(0.8)
    pdf.set_font("DejaVu", "B", 15)
    pdf.cell(W, 10, "BIÊN BẢN NGHIỆM THU",
             border=1, align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.set_line_width(0.2)
    pdf.set_draw_color(*BLACK)
    pdf.ln(2)

    # ── Bảng Engine / Customer / Address ──────────────────────
    W_LBL = 32   # cột nhãn (EN + VI stacked)
    W_VAL = W - W_LBL
    ROW_H = 11   # chiều cao mỗi hàng

    def _info_row(en: str, vi: str, val: str):
        y0 = pdf.get_y()
        # Vẽ khung 2 ô bằng rect (đảm bảo thẳng hàng)
        pdf.rect(M_LEFT,         y0, W_LBL, ROW_H)
        pdf.rect(M_LEFT + W_LBL, y0, W_VAL, ROW_H)
        # EN text (bold, trên)
        pdf.set_xy(M_LEFT + 1, y0 + 1)
        pdf.set_font("DejaVu", "B", 9)
        pdf.cell(W_LBL - 2, 5, en)
        # VI text (normal, dưới)
        pdf.set_xy(M_LEFT + 1, y0 + 5.5)
        pdf.set_font("DejaVu", "", 8)
        pdf.cell(W_LBL - 2, 5, vi)
        # Value text (căn giữa dọc trong ô)
        pdf.set_xy(M_LEFT + W_LBL + 2, y0 + 1.5)
        pdf.set_font("DejaVu", "", 9)
        pdf.cell(W_VAL - 4, ROW_H - 3, val, align="L")
        # Xuống dòng tiếp theo
        pdf.set_xy(M_LEFT, y0 + ROW_H)

    _info_row("Engine",   "Động cơ",    ten_dong_co)
    _info_row("Customer", "Khách hàng", khach_hang)
    _info_row("Address",  "Địa chỉ",   "")
    pdf.ln(3)

    # ── Section I ─────────────────────────────────────────────
    pdf.set_font("DejaVu", "B", 9.5)
    pdf.cell(W, 6, "I. Time and place of the test / Thời gian và địa điểm kiểm tra",
             new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("DejaVu", "", 9.5)
    pdf.cell(W, 6, f"At 7:30 AM on {ngay_en}, at Ngoc Tram Motor",
             new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("DejaVu", "", 9)
    pdf.cell(W, 6, f"Lúc 7h30 - {ngay_vi} , tại Điện cơ Ngọc Trâm",
             new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)

    # ── Section II: bảng 12 hạng mục ─────────────────────────
    pdf.set_font("DejaVu", "B", 9.5)
    pdf.cell(W, 6, "II. Hạng mục sửa chữa / Hạng mục sửa chữa",
             new_x="LMARGIN", new_y="NEXT")

    # ── Kích thước cố định bảng Section II ──────────────────
    W_STT  = 15
    W_DATE = 35
    W_PASS = 38
    W_HM   = W - W_STT - W_DATE - W_PASS   # 190-15-35-38 = 102

    HDR_H  = 9
    ROW_H2 = 7

    # Tọa độ x tuyệt đối — tính 1 lần, dùng cho cả header lẫn rows
    xSTT  = M_LEFT                          # 10
    xHM   = xSTT  + W_STT                  # 25
    xDATE = xHM   + W_HM                   # 127
    xPASS = xDATE + W_DATE                 # 162

    y_hdr = pdf.get_y()

    # ── Header: vẽ nền trước bằng rect, rồi viền, rồi text ──
    # 1) Fill toàn bộ hàng header một lần
    pdf.set_fill_color(*BLUE_CELL)
    pdf.rect(xSTT, y_hdr, W, HDR_H, style="F")

    # 2) Vẽ viền từng ô bằng rect (không fill, chỉ border)
    pdf.rect(xSTT,  y_hdr, W_STT,  HDR_H)
    pdf.rect(xHM,   y_hdr, W_HM,   HDR_H)
    pdf.rect(xDATE, y_hdr, W_DATE, HDR_H)
    pdf.rect(xPASS, y_hdr, W_PASS, HDR_H)

    # 3) Text header — border=0, set_xy rõ ràng cho từng ô
    pdf.set_font("DejaVu", "B", 8.5)
    pdf.set_text_color(*BLACK)

    pdf.set_xy(xSTT, y_hdr)
    pdf.cell(W_STT, HDR_H, "STT", border=0, align="C")

    pdf.set_xy(xHM, y_hdr)
    pdf.cell(W_HM, HDR_H, "Repair catalog / Hạng mục sửa chữa", border=0, align="C")

    pdf.set_xy(xDATE, y_hdr)
    pdf.cell(W_DATE, HDR_H, "Date / Ngày", border=0, align="C")

    # "Passed / Thông qua" — 2 dòng, không có đường ngăn giữa
    pdf.set_font("DejaVu", "B", 8)
    pdf.set_xy(xPASS, y_hdr + 0.5)
    pdf.cell(W_PASS, HDR_H / 2, "Passed /", border=0, align="C")
    pdf.set_xy(xPASS, y_hdr + HDR_H / 2)
    pdf.cell(W_PASS, HDR_H / 2, "Thông qua", border=0, align="C")

    # ── 12 dòng hạng mục ─────────────────────────────────────
    pdf.set_font("DejaVu", "", 8.5)
    for i in range(1, 11):
        noi_dung = hang_muc[i - 1] if (i - 1) < len(hang_muc) else ""
        y_r = y_hdr + HDR_H + (i - 1) * ROW_H2

        # Viền từng ô
        pdf.rect(xSTT,  y_r, W_STT,  ROW_H2)
        pdf.rect(xHM,   y_r, W_HM,   ROW_H2)
        pdf.rect(xDATE, y_r, W_DATE, ROW_H2)
        pdf.rect(xPASS, y_r, W_PASS, ROW_H2)

        # Text
        pdf.set_xy(xSTT, y_r + 1)
        pdf.cell(W_STT, ROW_H2 - 1, f"{i} -", border=0, align="C")

        pdf.set_xy(xHM, y_r + 1)
        pdf.cell(W_HM, ROW_H2 - 1, noi_dung, border=0, align="L")

    # Đặt cursor xuống dưới bảng
    pdf.set_xy(M_LEFT, y_hdr + HDR_H + 10 * ROW_H2)

    # Định nghĩa W1/W2/W3 dùng chung cho _draw_img_hdr
    BOX_H1 = 8
    BOX_H2 = 8
    W1 = 16   # logo
    W2 = int((W - W1) / 3)
    W3 = W - W1 - W2 - W2

    # ========================================================= #
    #  TRANG ẢNH 1/5 → 5/5                                     #
    # ========================================================= #

    # ── Hàm vẽ header tài liệu cho mỗi trang ảnh ─────────
    def _draw_img_hdr(page_str: str):
        """Vẽ header box 2 hàng bằng nhau, NT span 2 hàng, giống mẫu Excel.
        Dùng rect() cho tất cả viền để đảm bảo thẳng hàng tuyệt đối."""
        RH = BOX_H1   # chiều cao mỗi hàng = 8mm
        y1 = M_TOP
        y2 = M_TOP + RH

        # Tọa độ x cố định cho 4 cột
        xA = M_LEFT           # NT logo
        xB = M_LEFT + W1      # Quotation / Management doc
        xC = M_LEFT + W1 + W2 # Engine number / Edition date
        xD = M_LEFT + W1 + W2 * 2  # Order number / Page

        # ── Vẽ tất cả viền bằng rect() ──────────────────────────
        pdf.rect(xA, y1, W1, RH * 2)   # NT logo span 2 hàng
        pdf.rect(xB, y1, W2, RH)        # Quotation (hàng 1)
        pdf.rect(xC, y1, W2, RH)        # Engine number (hàng 1)
        pdf.rect(xD, y1, W3, RH)        # Order number (hàng 1)
        pdf.rect(xB, y2, W2, RH)        # Management doc (hàng 2)
        pdf.rect(xC, y2, W2, RH)        # Edition date (hàng 2)
        pdf.rect(xD, y2, W3, RH)        # Page (hàng 2)

        # ── NT logo — canh giữa theo chiều dọc 2 hàng ──
        pdf.set_font("DejaVu", "B", 9)
        pdf.set_xy(xA, y1 + RH - 3)
        pdf.cell(W1, 6, "NT", border=0, align="C")

        # ── Hàng 1 — text canh trái ô ──
        pdf.set_font("DejaVu", "", 7)
        # Quotation
        pdf.set_xy(xB, y1)
        pdf.cell(W2, RH, "Quotation / Báo giá :", border=0, align="L")
        # Engine number
        pdf.set_xy(xC, y1)
        pdf.cell(W2, RH, "Engine number / Số máy :", border=0, align="L")
        # Order number (có giá trị)
        pdf.set_xy(xD, y1)
        pdf.cell(W3, RH, f"Order number / Số ĐH: {cong_so}", border=0, align="L")

        # ── Hàng 2 — mỗi ô chia 2 dòng nhỏ bằng multi_cell ──
        sub_h = RH / 2  # 4mm mỗi sub-row

        # Management doc
        pdf.set_font("DejaVu", "", 6.5)
        pdf.set_xy(xB, y2)
        pdf.cell(W2, sub_h, "Management document / Tài liệu quản lý:", border=0, align="L")
        pdf.set_xy(xB, y2 + sub_h)
        pdf.cell(W2, sub_h, "QT-NT-029-1A", border=0, align="L")

        # Edition date
        pdf.set_xy(xC, y2)
        pdf.cell(W2, sub_h, "Edition date / Ngày ban hành:", border=0, align="L")
        pdf.set_xy(xC, y2 + sub_h)
        pdf.cell(W2, sub_h, "24/04/2025", border=0, align="L")

        # Page
        pdf.set_xy(xD, y2)
        pdf.cell(W3, sub_h, "Page / Trang:", border=0, align="L")
        pdf.set_xy(xD, y2 + sub_h)
        pdf.cell(W3, sub_h, page_str, border=0, align="L")

        # Đặt cursor sau header
        pdf.set_xy(M_LEFT, M_TOP + RH * 2 + 4)

    # ── Hàm vẽ 1 bảng có ô ảnh ─────────────────────────────
    def _bang_co_anh(ten_en: str, ten_vi: str, nhan_pha_display: list,
                     danh_sach_anh_bang: list, img_h: int = 55):
        """
        Vẽ bảng n cột:
          - Dòng 1: tiêu đề full width (xanh đậm/trắng) — bỏ qua nếu ten_en rỗng
          - Dòng 2: n nhãn cột (xanh nhạt)
          - Dòng 3: n ô ảnh (hoặc trống)
        img_h : chiều cao vùng ảnh (mm), điều chỉnh theo số bảng/trang.
        """
        HDR_H  = 8
        LBL_H  = 8
        n = len(nhan_pha_display)
        if n == 0:
            return
        base_w = int(W / n)
        col_ws = [base_w] * (n - 1) + [W - base_w * (n - 1)]
        col_xs = [M_LEFT + sum(col_ws[:i]) for i in range(n)]

        y_start = pdf.get_y()

        # -- Dòng tiêu đề (chỉ vẽ nếu có tên) --
        if ten_en:
            pdf.set_xy(M_LEFT, y_start)
            pdf.set_fill_color(*BLUE_HDR)
            pdf.set_text_color(*WHITE)
            pdf.set_font("DejaVu", "B", 9.5)
            hdr_txt = f"{ten_en} / {ten_vi}" if ten_vi else ten_en
            pdf.cell(W, HDR_H, hdr_txt, border=1, align="C", fill=True,
                     new_x="LMARGIN", new_y="NEXT")
            pdf.set_text_color(*BLACK)
            y_lbl = pdf.get_y()
        else:
            y_lbl = y_start

        # -- Dòng nhãn --
        pdf.set_fill_color(*BLUE_CELL)
        pdf.set_font("DejaVu", "B", 9)
        for idx_p, nhan in enumerate(nhan_pha_display):
            pdf.set_xy(col_xs[idx_p], y_lbl)
            pdf.cell(col_ws[idx_p], LBL_H, nhan, border=1, align="C", fill=True)
        pdf.set_xy(M_LEFT, y_lbl + LBL_H)

        # -- Dòng ô ảnh --
        y_img = pdf.get_y()
        for idx_a in range(n):
            x_img = col_xs[idx_a]
            cw    = col_ws[idx_a]
            pdf.rect(x_img, y_img, cw, img_h)
            if idx_a < len(danh_sach_anh_bang) and danh_sach_anh_bang[idx_a]:
                try:
                    pad = 1
                    pdf.image(danh_sach_anh_bang[idx_a],
                              x=x_img + pad, y=y_img + pad,
                              w=cw - 2*pad, h=img_h - 2*pad)
                except Exception:
                    pass
        pdf.set_xy(M_LEFT, y_img + img_h)
        pdf.ln(4)

    # ── Nhóm theo trang ảnh (trang X/5) ────────────────────
    # Cấu trúc mỗi phần tử: (page_str, img_h, [(ten_en, ten_vi, display_labels, storage_keys)])
    # ten_en = "" → không vẽ header xanh đậm (dùng cho nhóm tiếp theo cùng section)
    IMAGE_PAGES = [
        # Trang 1/5 — Stator + Temperature sensor
        ("1/5", 55, [
            ("Stator 1 coil resistance",
             "Điện trở cuộn dây Stator 1",
             ["U1 – U2", "V1 – V2", "W1 –  W2"],
             ["U1–U2",   "V1–V2",   "W1–W2"]),

            ("Resistance of the temperature sensor",
             "điện trở của cảm biến nhiệt độ",
             ["PTC", "PT100", "HEATER"],
             ["PTC", "PT100", "HEATER"]),
        ]),

        # Trang 2/5 — No-load test (3 sub-bảng chung 1 header xanh đầu)
        ("2/5", 50, [
            ("No-load test",
             "Kiểm tra không tải",
             ["Frequency / Tần số", "Voltage /  Voltage", "Current / Dòng điện"],
             ["Tần số",              "Voltage",             "Dòng điện"]),

            ("",   # không vẽ lại header
             "",
             ["Radial ↔ DE / AS", "Radial ↕ DE / AS", "Axial (X) DE / AS"],
             ["Radial ↔ DE / AS", "Radial ↕ DE / AS", "Axial (X) DE / AS"]),

            ("",
             "",
             ["Radial ↔ NDE / AS", "Radial ↕ NDE / AS", "Axial (X) NDE / AS"],
             ["Radial ↔ NDE / AS", "Radial ↕ NDE / AS", "Axial (X) NDE / AS"]),
        ]),

        # Trang 3/5 — Engine overview + Nắp
        ("3/5", 55, [
            ("Engine overview",
             "Tổng quan động cơ",
             ["Engine / Động cơ", "Nameplate / Bảng tên", "Quạt làm mát / Cooling fan"],
             ["Engine / Động cơ", "Nameplate / Bảng tên", "Quạt làm mát / Cooling fan"]),

            ("Nắp động cơ", "",
             ["DE", "NDE"],
             ["Nắp DE", "Nắp NDE"]),
        ]),

        # Trang 4/5 — Trục, Phớt, Bearing (4 bảng → img_h nhỏ hơn)
        ("4/5", 42, [
            ("Trục động cơ", "",
             ["Bạc đạn DE", "Bạc đạn NDE"],
             ["Bạc đạn DE", "Bạc đạn NDE"]),

            ("",  "",
             ["Phớt", "Đầu ren , chốt lavet", "Cánh quạt làm mát"],
             ["Phớt", "Đầu ren, chốt lavet",  "Cánh quạt làm mát"]),

            ("Phớt chặn", "",
             ["Phớt 1", "Phớt 2"],
             ["Phớt chặn 1", "Phớt chặn 2"]),

            ("Bearing / Vòng bi", "",
             ["DE", "NDE"],
             ["Vòng bi DE", "Vòng bi NDE"]),
        ]),

        # Trang 5/5 — Grease pump, Coil, Cân bằng
        ("5/5", 50, [
            ("Grease pump",
             "Bơm mỡ bôi trơn",
             ["Bearing / Vòng bi", "Grease cap / Mặt gam"],
             ["Vòng bi bơm mỡ",    "Mặt gam"]),

            ("Coil",
             "Cuộn dây",
             ["Vào dây", "Đai đấu"],
             ["Vào dây", "Đai đầu"]),

            ("Cân bằng động", "",
             ["Gá lên máy", "Sau khi cân"],
             ["Gá lên máy", "Sau khi cân"]),
        ]),
    ]

    # ── Tải toàn bộ ảnh về temp files ───────────────────────
    # Làm phẳng IMAGE_PAGES thành danh sách bảng + map index
    _all_tables = []
    for _, _, tables in IMAGE_PAGES:
        for t in tables:
            _all_tables.append(t)

    temp_files_all = []
    all_temps = []
    for _, _, _, storage_keys in _all_tables:
        nhom_paths = []
        for lbl in storage_keys:
            urls = anh_do_luong.get(lbl, [])
            p = _tai_anh_tam(urls[0]) if urls else None
            nhom_paths.append(p)
            if p:
                all_temps.append(p)
        temp_files_all.append(nhom_paths)

    # ── Vẽ từng trang ảnh ───────────────────────────────────
    tbl_idx = 0
    for page_str, img_h, tables in IMAGE_PAGES:
        pdf.add_page()
        _draw_img_hdr(page_str)
        for ten_en, ten_vi, display_labels, _ in tables:
            _bang_co_anh(ten_en, ten_vi, display_labels, temp_files_all[tbl_idx], img_h)
            tbl_idx += 1

    # ── Dọn temp files ───────────────────────────────────────
    for p in all_temps:
        if p and os.path.exists(p):
            try:
                os.unlink(p)
            except Exception:
                pass

    # ── Xuất PDF ─────────────────────────────────────────────
    return bytes(pdf.output())


# ============================================================
# TẠO EXCEL BIÊN BẢN NGHIỆM THU
# ============================================================
def tao_excel_nghiem_thu(thong_tin_task: dict) -> bytes:
    """Tạo file Excel biên bản nghiệm thu — toàn bộ trên 1 sheet 'Biên Bản'.
    Phần 1: Header công ty, tiêu đề BBNT, Engine/Customer, Section I-II, footer.
    Phần 2: Tất cả bảng ảnh đo lường (IMAGE_PAGES 1/5→5/5) từ trên xuống.
    Dùng 6 cột đều nhau (A-F) xuyên suốt.
    """
    import io, os as _os
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side)
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker
    from openpyxl.worksheet.pagebreak import Break

    # ── Logo công ty ────────────────────────────────────────────
    _LOGO_PATH = _os.path.join(_os.path.dirname(__file__), "image.png")
    _logo_bio: io.BytesIO | None = None
    if _os.path.exists(_LOGO_PATH):
        with open(_LOGO_PATH, "rb") as _lf:
            _logo_bio = io.BytesIO(_lf.read())

    # ── Parse dữ liệu ──────────────────────────────────────────
    ten_dong_co  = str(thong_tin_task.get("Tên Công Việc", ""))
    khach_hang   = str(thong_tin_task.get("Công Ty", ""))
    cong_so      = str(thong_tin_task.get("Công Số", ""))
    ma_so        = str(thong_tin_task.get("Mã Số", ""))
    so_po_kh     = str(thong_tin_task.get("Số PO KH/HĐ", ""))
    so_bao_gia   = str(thong_tin_task.get("Số Báo Giá", ""))
    mo_ta        = str(thong_tin_task.get("Mô Tả", ""))
    nhan_vien    = str(thong_tin_task.get("Nhân Viên", ""))
    ngay_tao_str = str(thong_tin_task.get("Ngày Tạo", ""))
    anh_do_luong = doc_anh_do_luong(
        str(thong_tin_task.get("Ảnh Đo Lường", "") or ""))

    # Lookup địa chỉ khách hàng từ danh sách công ty
    _dia_chi_kh = ""
    try:
        _df_ct = lay_danh_sach_cong_ty()
        _match = _df_ct[_df_ct["Tên Công Ty"].str.strip() == khach_hang.strip()]
        if not _match.empty:
            _dia_chi_kh = str(_match.iloc[0].get("Địa Chỉ", "") or "")
    except Exception:
        pass

    try:
        dt = datetime.strptime(ngay_tao_str[:10], "%Y-%m-%d")
        ngay_en = dt.strftime("%d %B %Y")
        ngay_vi = f"Ngày {dt.day:02d} Tháng {dt.month:02d} Năm {dt.year}"
    except Exception:
        ngay_en = ngay_tao_str
        ngay_vi = ngay_tao_str

    # Lấy các mục checklist đã tick (done=True) làm hạng mục sửa chữa
    _raw_cl = str(thong_tin_task.get("Checklist", "") or "[]")
    try:
        _cl_items = json.loads(_raw_cl)
    except Exception:
        _cl_items = []
    hang_muc = [
        it["text"].strip() for it in _cl_items
        if isinstance(it, dict) and it.get("done") and it.get("text", "").strip()
    ]
    if not hang_muc:  # fallback: dùng Mô Tả nếu chưa có checklist
        hang_muc = [h.strip() for h in mo_ta.split("\n") if h.strip()]

    # ── Styles ──────────────────────────────────────────────────
    thin  = Side(style="thin",   color="000000")
    thick = Side(style="medium", color="000000")
    brd_all   = Border(left=thin, right=thin, top=thin, bottom=thin)
    brd_thick = Border(left=thick, right=thick, top=thick, bottom=thick)

    BLUE_HDR   = "5B84CC"
    BLUE_CELL  = "BDD7EE"
    BLUE_LABEL = "E2EFF8"   # nhạt hơn BLUE_CELL, dùng cho label row khung ảnh
    PURPLE    = "70309F"
    WHITE     = "FFFFFF"
    YELLOW    = "FFF2CC"

    def _font(bold=False, size=10, color="000000", italic=False):
        return Font(bold=bold, size=size, color=color, italic=italic,
                    name="Times New Roman")

    def _fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def _align(h="center", v="center", wrap=True):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    def _sc(row, col, value="", bold=False, size=10, color="000000",
            h_align="center", fill_color=None, border=None, wrap=True):
        """Set cell shorthand."""
        c = ws.cell(row=row, column=col, value=value)
        c.font      = _font(bold=bold, size=size, color=color)
        c.alignment = _align(h=h_align, wrap=wrap)
        if fill_color:
            c.fill  = _fill(fill_color)
        if border:
            c.border = border
        return c

    def _brd_merge(row, c1, c2, brd, r_end=None):
        """Set outer borders on ALL cells of a merged range (not just top-left).
        This is needed because openpyxl only renders the border from the top-left
        cell of a merge; the right/bottom edges of the merge are missing otherwise.
        """
        _re = r_end if r_end is not None else row
        for r in range(row, _re + 1):
            for c in range(c1, c2 + 1):
                left   = brd.left   if c == c1 else Side(style=None)
                right  = brd.right  if c == c2 else Side(style=None)
                top    = brd.top    if r == row else Side(style=None)
                bottom = brd.bottom if r == _re  else Side(style=None)
                ws.cell(row=r, column=c).border = Border(
                    left=left, right=right, top=top, bottom=bottom)

    # ── Tạo workbook — 1 sheet duy nhất ─────────────────────────
    # Dùng 6 cột đều nhau (A-F) xuyên suốt:
    #   A=6 (STT/label hẹp), B-F=25 (nội dung)
    # Bảng 3 cột ảnh → merge cặp  A:B | C:D | E:F
    # Bảng 2 cột ảnh → merge bộ 3 A:C | D:F
    NCOLS = 6
    COL_W = 25
    # COL_PX không còn dùng (dùng TwoCellAnchor thay thế)

    wb = Workbook()
    ws = wb.active
    ws.title = "Biên Bản"
    for ci in "ABCDEF":
        ws.column_dimensions[ci].width = COL_W

    row = 1

    # ═══════════════════════════════════════════════════════════
    # PHẦN 1 — BIÊN BẢN NGHIỆM THU
    # ═══════════════════════════════════════════════════════════

    # ── Header công ty (A:E = text, F = logo) ──────────────────
    # Header: cột A = logo (span 4 hàng), B:F = nội dung công ty (canh giữa)
    _hdr_start = row
    _hdr_end   = row + 3   # 4 hàng cố định: 1 tên + 3 địa chỉ

    # ── Ô logo bên TRÁI (A, span 4 hàng) ────────────────────────
    ws.merge_cells(f"A{_hdr_start}:A{_hdr_end}")
    _brd_merge(_hdr_start, 1, 1, brd_all, r_end=_hdr_end)
    if _logo_bio is not None:
        _logo_bio.seek(0)
        _xl_logo = XLImage(io.BytesIO(_logo_bio.read()))
        PAD_L = 18288   # ~2pt padding
        _logo_anchor = TwoCellAnchor(editAs="twoCell")
        _logo_anchor._from = AnchorMarker(col=0, colOff=PAD_L,
                                          row=_hdr_start - 1, rowOff=PAD_L)
        _logo_anchor.to   = AnchorMarker(col=1, colOff=-PAD_L,
                                          row=_hdr_end, rowOff=-PAD_L)
        _xl_logo.anchor = _logo_anchor
        ws.add_image(_xl_logo)

    # ── Tên công ty (B:F, canh giữa) ─────────────────────────────
    ws.merge_cells(f"B{row}:F{row}")
    _sc(row, 2, "CÔNG TY TNHH MỘT THÀNH VIÊN ĐIỆN CƠ NGỌC TRÂM",
        bold=True, size=14, color=PURPLE, h_align="center")
    ws.row_dimensions[row].height = 22
    row += 1

    for txt in [
        "Địa chỉ : 8/5, hẻm 4, tổ 9, Khu phố Kim Sơn, Phường Long Thành, Thành phố Đồng Nai",
        "Website: ngoctrammotor.com   Mail: kd@ngoctrammotor.com",
        "MST: 3603238978  ĐT: 0907 042 043 (Mr.Hiệp) – 0908 062 291 (Ms.Linh)",
    ]:
        ws.merge_cells(f"B{row}:F{row}")
        _sc(row, 2, txt, size=13, h_align="center")
        ws.row_dimensions[row].height = 20
        row += 1
    # Outer border cho toàn bộ khối B:F header
    _brd_merge(_hdr_start, 2, 6, brd_all, r_end=_hdr_end)

    row += 1  # khoảng trống
    ws.row_dimensions[row - 1].height = 20

    # ── Tiêu đề BBNT ────────────────────────────────────────────
    ws.merge_cells(f"A{row}:F{row}")
    _sc(row, 1, "REPAIR ACCEPTANCE CERTIFICATE",
        bold=True, size=13, color="1F3864", h_align="center")
    _brd_merge(row, 1, 6, brd_all)
    ws.row_dimensions[row].height = 22
    row += 1

    ws.merge_cells(f"A{row}:F{row}")
    c = ws.cell(row=row, column=1, value="BIÊN BẢN NGHIỆM THU")
    c.font      = _font(bold=True, size=16, color="1F3864")
    c.alignment = _align()
    c.fill      = _fill(BLUE_CELL)
    _brd_merge(row, 1, 6, brd_all)
    ws.row_dimensions[row].height = 28
    row += 1

    row += 1  # khoảng trống
    ws.row_dimensions[row - 1].height = 20

    # ── Bảng Engine / Customer / Address ────────────────────────
    for en_lbl, vi_lbl, val in [
        ("Engine",   "Động cơ",    ten_dong_co),
        ("Customer", "Khách hàng", khach_hang),
        ("Address",  "Địa chỉ",   _dia_chi_kh),
    ]:
        # A:B merged = label, C:F merged = value
        ws.merge_cells(f"A{row}:B{row}")
        _sc(row, 1, f"{en_lbl} / {vi_lbl}",
            bold=True, size=13, fill_color=BLUE_CELL, border=brd_all, h_align="left")
        _brd_merge(row, 1, 2, brd_all)
        ws.merge_cells(f"C{row}:F{row}")
        _sc(row, 3, val, size=13, border=brd_all, h_align="left")
        _brd_merge(row, 3, 6, brd_all)
        ws.row_dimensions[row].height = 30
        row += 1

    row += 1
    ws.row_dimensions[row - 1].height = 20

    # ── Section I ────────────────────────────────────────────────
    ws.merge_cells(f"A{row}:F{row}")
    _sc(row, 1,
        "I. Time and place of the test / Thời gian và địa điểm kiểm tra",
        bold=True, size=13, h_align="left", border=brd_all)
    _brd_merge(row, 1, 6, brd_all)
    ws.row_dimensions[row].height = 26
    row += 1

    ws.merge_cells(f"A{row}:F{row}")
    _sc(row, 1, f"At 7:30 AM on {ngay_en}, at Ngoc Tram Motor",
        size=13, h_align="left", border=brd_all)
    _brd_merge(row, 1, 6, brd_all)
    ws.row_dimensions[row].height = 26
    row += 1

    ws.merge_cells(f"A{row}:F{row}")
    _sc(row, 1, f"Lúc 7h30 - {ngay_vi}, tại Điện cơ Ngọc Trâm",
        size=13, h_align="left", border=brd_all)
    _brd_merge(row, 1, 6, brd_all)
    ws.row_dimensions[row].height = 26
    row += 1
    row += 1
    ws.row_dimensions[row - 1].height = 20

    # ── Confirmation paragraph (bordered merged A:F, ~5 rows) ────
    _confirm_start = row
    _confirm_end   = row + 4
    ws.merge_cells(f"A{_confirm_start}:F{_confirm_end}")
    _cc = ws.cell(row=_confirm_start, column=1)
    _en_text = (
        "We hereby confirm that all electrical tests on this machine have been "
        "performed in accordance with the relevant standards. This electronically "
        "generated report is also valid without a handwritten signature."
    )
    _vi_text = (
        "Chúng tôi xác nhận rằng tất cả các thử nghiệm điện trên máy này đã được "
        "thực hiện theo các tiêu chuẩn liên quan. Báo cáo được tạo bằng điện tử "
        "này cũng có giá trị mà không cần chữ ký viết tay."
    )
    _cc.value     = f"{_en_text}\n{_vi_text}"
    _cc.font      = Font(name="Times New Roman", size=13, italic=True)
    _cc.alignment = Alignment(horizontal="center", vertical="center",
                              wrap_text=True)
    _cc.fill = _fill("F2F2F2")   # nền xám nhạt
    _brd_merge(_confirm_start, 1, 6, brd_all, r_end=_confirm_end)
    for _r in range(_confirm_start, _confirm_end + 1):
        ws.row_dimensions[_r].height = 28
    row = _confirm_end + 1
    row += 1
    ws.row_dimensions[row - 1].height = 20

    # ── CLIENT / CONTRACTOR signature block ──────────────────────
    # Header row — nền xanh nhạt, chữ đen đậm (dễ đọc khi in đen trắng)
    ws.merge_cells(f"A{row}:C{row}")
    _sc(row, 1, "CLIENT / CHỦ ĐẦU TƯ",
        bold=True, size=13, color="000000", fill_color=BLUE_CELL,
        border=brd_all, h_align="center")
    _brd_merge(row, 1, 3, brd_all)
    ws.merge_cells(f"D{row}:F{row}")
    _sc(row, 4, "CONTRACTOR / NHÀ THẦU",
        bold=True, size=13, color="000000", fill_color=BLUE_CELL,
        border=brd_all, h_align="center")
    _brd_merge(row, 4, 6, brd_all)
    ws.row_dimensions[row].height = 28
    row += 1

    # Position row
    ws.merge_cells(f"A{row}:C{row}")
    _sc(row, 1, "Position / Chức vụ :",
        size=13, border=brd_all, h_align="left")
    _brd_merge(row, 1, 3, brd_all)
    ws.merge_cells(f"D{row}:F{row}")
    _sc(row, 4, "Position / Chức vụ : Trưởng bộ phận Kỹ thuật",
        size=13, border=brd_all, h_align="left")
    _brd_merge(row, 4, 6, brd_all)
    ws.row_dimensions[row].height = 28
    row += 1

    # Merge 3 rows thành 1 ô ký tên mỗi bên (khoảng trắng để ký tay)
    _sig_start = row
    _sig_end   = row + 2
    ws.merge_cells(f"A{_sig_start}:C{_sig_end}")
    _sc(_sig_start, 1, "", size=13, border=brd_all)
    _brd_merge(_sig_start, 1, 3, brd_all, r_end=_sig_end)
    ws.merge_cells(f"D{_sig_start}:F{_sig_end}")
    _sc(_sig_start, 4, "", size=13, border=brd_all)
    _brd_merge(_sig_start, 4, 6, brd_all, r_end=_sig_end)
    for _r in range(_sig_start, _sig_end + 1):
        ws.row_dimensions[_r].height = 40
    row = _sig_end + 1

    # Full name row — bỏ tên nhân viên, để trống để điền tay
    ws.merge_cells(f"A{row}:C{row}")
    _sc(row, 1, "Full name / Họ và tên :",
        size=13, border=brd_all, h_align="left")
    _brd_merge(row, 1, 3, brd_all)
    ws.merge_cells(f"D{row}:F{row}")
    _sc(row, 4, "Full name / Họ và tên :",
        size=13, border=brd_all, h_align="left")
    _brd_merge(row, 4, 6, brd_all)
    ws.row_dimensions[row].height = 28
    row += 1

    # Date row
    ws.merge_cells(f"A{row}:C{row}")
    _sc(row, 1, "Date / Ngày :",
        size=13, border=brd_all, h_align="left")
    _brd_merge(row, 1, 3, brd_all)
    ws.merge_cells(f"D{row}:F{row}")
    _sc(row, 4, "Date / Ngày :",
        size=13, border=brd_all, h_align="left")
    _brd_merge(row, 4, 6, brd_all)
    ws.row_dimensions[row].height = 28
    row += 1
    row += 1
    ws.row_dimensions[row - 1].height = 20

    # ── Section II ───────────────────────────────────────────────
    ws.merge_cells(f"A{row}:F{row}")
    _sc(row, 1, "II. Hạng mục sửa chữa / Repair catalog",
        bold=True, size=13, h_align="left")
    ws.row_dimensions[row].height = 26
    row += 1

    # Header bảng — nền xanh nhạt, chữ đen đậm (dễ in đen trắng)
    _sc(row, 1, "STT", bold=True, size=13, color="000000",
        fill_color=BLUE_CELL, border=brd_all)
    ws.merge_cells(f"B{row}:D{row}")
    _sc(row, 2, "Repair catalog / Hạng mục sửa chữa",
        bold=True, size=13, color="000000", fill_color=BLUE_CELL, border=brd_all)
    _brd_merge(row, 2, 4, brd_all)
    _sc(row, 5, "Date / Ngày", bold=True, size=13, color="000000",
        fill_color=BLUE_CELL, border=brd_all)
    _sc(row, 6, "Passed / Thông qua", bold=True, size=13, color="000000",
        fill_color=BLUE_CELL, border=brd_all)
    ws.row_dimensions[row].height = 28
    row += 1

    _hm_rows = hang_muc if hang_muc else [""]
    for i, noi_dung in enumerate(_hm_rows, start=1):
        fill = YELLOW if i % 2 == 0 else WHITE
        _sc(row, 1, f"{i}", size=13, border=brd_all, fill_color=fill)
        ws.merge_cells(f"B{row}:D{row}")
        _sc(row, 2, noi_dung, size=13, border=brd_all,
            h_align="left", fill_color=fill)
        _brd_merge(row, 2, 4, brd_all)
        _sc(row, 5, "", size=13, border=brd_all, fill_color=fill)
        _sc(row, 6, "", size=13, border=brd_all, fill_color=fill)
        ws.row_dimensions[row].height = 26
        row += 1

    row += 1
    ws.row_dimensions[row - 1].height = 14

    # ═══════════════════════════════════════════════════════════
    # PHẦN 2 — ẢNH ĐO LƯỜNG (IMAGE_PAGES 1/5 → 5/5)
    # Tiếp tục trên cùng sheet, cuộn xuống từ phần BBNT
    # ═══════════════════════════════════════════════════════════

    IMAGE_PAGES = [
        ("3/8", 75, [
            ("Resistance / Điện trở", "",
             ["R (U1 – U2)", "R (V1 – V2)", "R (W1 – W2)"],
             ["R_U1U2", "R_V1V2", "R_W1W2"]),
            ("", "",
             ["R (PTC)", "R (PT100)", "R (HEATER)"],
             ["R_PTC", "R_PT100", "R_HEATER"]),
        ]),
        ("4/8", 75, [
            ("Insulation Resistance / Cách điện", "",
             ["IR (U – V)", "IR (U – W)", "IR (V – W)"],
             ["IR_UV", "IR_UW", "IR_VW"]),
            ("", "",
             ["IR (PTC – E)", "IR (PT100 – E)", "IR (HEATER – E)"],
             ["IR_PTC_E", "IR_PT100_E", "IR_HEATER_E"]),
            ("", "",
             ["IR (U – E)", "IR (V – E)", "IR (W – E)"],
             ["IR_U_E", "IR_V_E", "IR_W_E"]),
        ]),
        ("5/8", 72, [
            ("Engine Overview / Tổng Quan Động Cơ", "",
             ["Engine / Động cơ", "Nameplate / Bảng thông số"],
             ["eng_overview", "eng_nameplate"]),
            ("Terminal Box / Hộp Điện", "",
             ["Before", "After"],
             ["tb_before", "tb_after"]),
            ("Terminal Block / Cầu Đấu Điện", "",
             ["Before", "After"],
             ["tbl_before", "tbl_after"]),
        ]),
        ("6/8", 72, [
            ("Cover Fan / Chụp Bảo Vệ Cánh Quạt", "",
             ["Before", "After"],
             ["cf_before", "cf_after"]),
            ("Rotor Fan / Cánh Quạt", "",
             ["Before", "After"],
             ["rf_before", "rf_after"]),
            ("Coil / Cuộn dây", "",
             ["Before", "After"],
             ["coil_before", "coil_after"]),
        ]),
        ("7/8", 72, [
            ("End Cover DE / Nắp Đầu Tải", "",
             ["Before", "After"],
             ["ec_de_before", "ec_de_after"]),
            ("End Cover NDE / Nắp Đầu Không Tải", "",
             ["Before", "After"],
             ["ec_nde_before", "ec_nde_after"]),
            ("Shaft at DE / Trục Đầu Tải", "",
             ["Before", "After"],
             ["shaft_de_before", "shaft_de_after"]),
        ]),
        ("8/8", 72, [
            ("Shaft at NDE / Trục Đầu Không Tải", "",
             ["Before", "After"],
             ["shaft_nde_before", "shaft_nde_after"]),
            ("DE Bearing / Vòng Bi Đầu Tải", "",
             ["Before", "After"],
             ["de_brg_before", "de_brg_after"]),
            ("NDE Bearing / Vòng Bi Đầu Không Tải", "",
             ["Before", "After"],
             ["nde_brg_before", "nde_brg_after"]),
        ]),
    ]

    # ── Tải ảnh vào BytesIO ngay (không giữ file path) ──────────
    _all_skeys: list = []
    for _, _, tbls in IMAGE_PAGES:
        for _, _, _, skeys in tbls:
            _all_skeys.extend(skeys)

    _img_cache: dict = {}   # key → BytesIO | None
    for key in _all_skeys:
        if key not in _img_cache:
            urls = anh_do_luong.get(key, [])
            _img_cache[key] = None
            if urls:
                tmp = _tai_anh_tam(urls[0])
                if tmp and os.path.exists(tmp):
                    try:
                        with open(tmp, "rb") as _f:
                            _img_cache[key] = io.BytesIO(_f.read())
                    except Exception:
                        pass
                    try:
                        os.unlink(tmp)
                    except Exception:
                        pass
    # ── Lọc chỉ giữ trang có ít nhất 1 ảnh + tính lại tổng trang ──
    _active_pages = [
        (_, img_h, tables)
        for _, img_h, tables in IMAGE_PAGES
        if any(
            _img_cache.get(k) is not None
            for _, _, _, skeys in tables
            for k in skeys
        )
    ]
    _total_pages = 2 + len(_active_pages)  # 1=BBNT, 2=terminal, 3..N=ảnh

    def _col_ranges(n: int) -> list:
        if n == 3:
            return [(1, 2), (3, 4), (5, 6)]   # A:B | C:D | E:F
        if n == 2:
            return [(1, 3), (4, 6)]            # A:C | D:F
        step = NCOLS // n
        return [(i * step + 1, (i + 1) * step) for i in range(n)]

    ws.row_breaks.append(Break(id=row - 1))   # ── page break: page 1 → 2
    row += 1  # blank row = đầu trang 2

    # ═══════════════════════════════════════════════════════════
    # TRANG 2 — BẢNG SỐ LIỆU ĐO LƯỜNG (manual entry)
    # ═══════════════════════════════════════════════════════════

    def _apply_outer_thick(r_start, r_end):
        """Bọc khung thick (dày) bên ngoài block rows r_start:r_end, cols A-F."""
        tk = Side(style="thick", color="000000")
        for r in range(r_start, r_end + 1):
            for c in range(1, 7):
                cell = ws.cell(row=r, column=c)
                b = cell.border
                cell.border = Border(
                    top    = tk if r == r_start else b.top,
                    bottom = tk if r == r_end   else b.bottom,
                    left   = tk if c == 1       else b.left,
                    right  = tk if c == 6       else b.right,
                )

    def _nt_mini_header(page_lbl):
        """Render NT mini header (2 rows) at current row, advance row by 3."""
        nonlocal row
        ws.merge_cells(f"A{row}:A{row+1}")
        _nt2 = ws.cell(row=row, column=1,
                       value="" if _logo_bio is not None else "NT")
        _nt2.font      = _font(bold=True, size=11)
        _nt2.alignment = _align()
        _brd_merge(row, 1, 1, brd_all, r_end=row + 1)
        if _logo_bio is not None:
            _logo_bio.seek(0)
            _xl_nt2 = XLImage(io.BytesIO(_logo_bio.read()))
            PAD_NT2 = 9144
            _nt2_anchor = TwoCellAnchor(editAs="twoCell")
            _nt2_anchor._from = AnchorMarker(col=0, colOff=PAD_NT2,
                                              row=row - 1, rowOff=PAD_NT2)
            _nt2_anchor.to   = AnchorMarker(col=1, colOff=-PAD_NT2,
                                              row=row + 1, rowOff=-PAD_NT2)
            _xl_nt2.anchor = _nt2_anchor
            ws.add_image(_xl_nt2)
        ws.merge_cells(f"B{row}:C{row}")
        _sc(row, 2, f"Quotation / Báo giá: {so_bao_gia}", size=12, border=brd_all, h_align="left")
        _brd_merge(row, 2, 3, brd_all)
        ws.merge_cells(f"D{row}:E{row}")
        _sc(row, 4, f"Engine number / Số máy: {ma_so}", size=12, border=brd_all, h_align="left")
        _brd_merge(row, 4, 5, brd_all)
        _sc(row, 6, f"Order number / Số ĐH: {so_po_kh}", size=12, border=brd_all, h_align="left")
        ws.merge_cells(f"B{row+1}:C{row+1}")
        _sc(row + 1, 2,
            "Management document / Tài liệu quản lý: QT-NT-029-1A",
            size=12, border=brd_all, h_align="left")
        _brd_merge(row + 1, 2, 3, brd_all)
        ws.merge_cells(f"D{row+1}:E{row+1}")
        _sc(row + 1, 4, "Edition date / Ngày ban hành: 24/04/2025",
            size=12, border=brd_all, h_align="left")
        _brd_merge(row + 1, 4, 5, brd_all)
        _sc(row + 1, 6, f"Page / Trang: {page_lbl}", size=12, border=brd_all, h_align="left")
        ws.row_dimensions[row].height     = 26
        ws.row_dimensions[row + 1].height = 24
        ws.row_dimensions[row + 2].height = 14   # blank gap sau mini header
        _apply_outer_thick(row, row + 1)          # viền ngoài dày cho mini header
        row += 3

    # NT mini header for page 2
    _nt_mini_header(f"2/{_total_pages}")

    LIGHT_BLUE = "DAEEF3"

    def _val_cell(r, c1, c2):
        """Empty bordered value cell (merged if c1 != c2)."""
        if c1 != c2:
            ws.merge_cells(
                f"{get_column_letter(c1)}{r}:{get_column_letter(c2)}{r}")
        _sc(r, c1, "", size=9, border=brd_all)
        _brd_merge(r, c1, c2, brd_all)

    def _cat_label(r_start, r_end, label):
        """Dark-blue category label merged vertically in col A+B (2 columns)."""
        ws.merge_cells(f"A{r_start}:B{r_end}")
        c = ws.cell(row=r_start, column=1, value=label)
        c.font      = Font(bold=True, size=13, color=WHITE,
                           name="Times New Roman")
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True, text_rotation=0)
        c.fill   = _fill(BLUE_HDR)
        _brd_merge(r_start, 1, 2, brd_all, r_end=r_end)

    # ── Group 1: Terminal / Đầu cực đấu ─────────────────────────
    _terminal_tbl_start = row   # để bọc viền ngoài toàn bảng Terminal
    # Col A+B: blank xanh (cùng màu category), Col C-F: Terminal header
    _sc(row, 1, "", size=13, border=brd_all, fill_color=BLUE_HDR)
    _sc(row, 2, "", size=13, border=brd_all, fill_color=BLUE_HDR)
    ws.merge_cells(f"C{row}:F{row}")
    _tc = ws.cell(row=row, column=3,
                  value="Terminal / Đầu cực đấu")
    _tc.font      = Font(bold=True, size=13, color=WHITE, name="Times New Roman")
    _tc.alignment = Alignment(horizontal="center", vertical="center")
    _tc.fill      = _fill(BLUE_HDR)
    _brd_merge(row, 3, 6, brd_all)
    ws.row_dimensions[row].height = 26
    row += 1

    # Category: Resistance (mΩ) / Điện trở — 6 rows
    _res_rows = [
        ("R(U1-U2)", "R_U1U2"),
        ("R(V1-V2)", "R_V1V2"),
        ("R(W1-W2)", "R_W1W2"),
        ("R(PTC)",   "R_PTC"),
        ("R(PT100)", "R_PT100"),
        ("R(HEATER)","R_HEATER"),
    ]
    _res_start = row
    _res_end   = row + len(_res_rows) - 1
    _cat_label(_res_start, _res_end, "Resistance (Ω) / Điện trở")
    for _lbl, _vkey in _res_rows:
        _val = str(anh_do_luong.get(f"{_vkey}_val", "") or "")
        _sc(row, 3, _lbl, bold=True, size=13, border=brd_all, h_align="left", fill_color=LIGHT_BLUE)
        ws.merge_cells(f"D{row}:F{row}")
        _sc(row, 4, _val, size=13, border=brd_all)
        _brd_merge(row, 4, 6, brd_all)
        ws.row_dimensions[row].height = 28
        row += 1

    # Category: Insulation Resistance (MΩ) / Cách điện — 7 rows
    # IR(U;V;W-E) gộp giá trị từ 3 key riêng: IR_U_E, IR_V_E, IR_W_E
    _ir_rows = [
        ("IR(U-V)",     "IR_UV"),
        ("IR(U-W)",     "IR_UW"),
        ("IR(V-W)",     "IR_VW"),
        ("IR(U;V;W-E)", ["IR_U_E", "IR_V_E", "IR_W_E"]),
        ("IR(PTC-E)",   "IR_PTC_E"),
        ("IR(PT100-E)", "IR_PT100_E"),
        ("IR(HEATER-E)","IR_HEATER_E"),
    ]
    _ir_start = row
    _ir_end   = row + len(_ir_rows) - 1
    _cat_label(_ir_start, _ir_end, "Insulation Resistance (MΩ) / Cách điện")
    for _lbl, _vkey in _ir_rows:
        if isinstance(_vkey, list):
            _parts = [str(anh_do_luong.get(f"{k}_val", "") or "") for k in _vkey]
            _val = " / ".join(p for p in _parts if p)
        else:
            _val = str(anh_do_luong.get(f"{_vkey}_val", "") or "")
        _sc(row, 3, _lbl, bold=True, size=13, border=brd_all, h_align="left", fill_color=LIGHT_BLUE)
        ws.merge_cells(f"D{row}:F{row}")
        _sc(row, 4, _val, size=13, border=brd_all)
        _brd_merge(row, 4, 6, brd_all)
        ws.row_dimensions[row].height = 28
        row += 1

    # Category: No-load test / Kiểm tra không tải — 3 rows (with L1/L2/L3 sub-header)
    _nl_start = row
    _nl_end   = row + 2   # sub-header + Voltage + Current = 3 rows
    _cat_label(_nl_start, _nl_end, "No-load test / Kiểm tra không tải")
    # Sub-header: col3=blank/blue, col4=L1, col5=L2, col6=L3
    _sc(row, 3, "", size=13, border=brd_all, fill_color=BLUE_HDR)
    for _c, _lbl in zip([4, 5, 6], ["L1", "L2", "L3"]):
        _sc(row, _c, _lbl, bold=True, size=13, color=WHITE,
            border=brd_all, fill_color=BLUE_HDR)
    ws.row_dimensions[row].height = 28
    row += 1
    for _lbl in ["Voltage (V)", "Current (A)"]:
        _sc(row, 3, _lbl, bold=True, size=13, border=brd_all, h_align="left", fill_color=LIGHT_BLUE)
        if _lbl == "Voltage (V)":
            for _c in range(4, 7):
                _sc(row, _c, "380V", size=13, border=brd_all)
        else:
            for _c, _ck in zip([4, 5, 6], ["cur_L1", "cur_L2", "cur_L3"]):
                _sc(row, _c, anh_do_luong.get(f"{_ck}_val", ""), size=13, border=brd_all, h_align="center")
        ws.row_dimensions[row].height = 28
        row += 1

    # Vibration row — xanh đậm chữ trắng như mẫu
    _vib_labels = [
        "Radial ↔ DE/AS", "Radial ↕ DE/AS", "Axial (X) DE/AS",
        "Radial ↔ NDE/AS", "Radial ↕ NDE/AS", "Axial (X) NDE/AS",
    ]
    for _c_idx, _lbl in enumerate(_vib_labels, start=1):
        _sc(row, _c_idx, _lbl, bold=True, size=13, color=WHITE,
            border=brd_all, fill_color=BLUE_HDR, h_align="center")
    ws.row_dimensions[row].height = 28
    row += 1
    _vib_keys = ["vib_rad_h_de", "vib_rad_v_de", "vib_axial_de",
                 "vib_rad_h_nde", "vib_rad_v_nde", "vib_axial_nde"]
    for _c_idx, _vk in enumerate(_vib_keys, start=1):
        _vib_val = anh_do_luong.get(f"{_vk}_val", "")
        _sc(row, _c_idx, _vib_val, size=13, border=brd_all, h_align="center")
    ws.row_dimensions[row].height = 30
    row += 1          # past vib values
    _apply_outer_thick(_terminal_tbl_start, row - 1)  # viền ngoài dày toàn bảng Terminal
    # blank filler rows để lấp đầy trang 2
    for _fi in range(8):
        ws.row_dimensions[row].height = 40
        row += 1
    ws.row_breaks.append(Break(id=row - 1))   # ── page break: page 2 → 3

    for _pi, (_, img_h_mm, tables) in enumerate(_active_pages):
        page_str = f"{_pi + 3}/{_total_pages}"
        # ── Header NT mini (2 hàng) ──────────────────────────────
        _nt_mini_header(page_str)

        for ten_en, ten_vi, display_labels, storage_keys in tables:
            _tbl_start = row          # theo dõi dòng đầu của bảng này
            n      = len(display_labels)
            ranges = _col_ranges(n)
            # 2-col cells rộng hơn 3-col → tăng chiều cao tương ứng để hình cân đối
            _h_scale = 1.4 if n == 2 else 1.0
            img_row_h_pt = img_h_mm * 2.835 * _h_scale

            # Kiểm tra trước: có ảnh nào trong bảng này không?
            _has_any_img = any(_img_cache.get(k) is not None for k in storage_keys)
            _no_brd    = Border()
            _white_fill = PatternFill("solid", fgColor="FFFFFF")

            def _blank_row(r, c_from, c_to):
                """Xoá viền + trắng hoàn toàn cho một hàng."""
                # Unmerge bất kỳ merged range nào nằm trong hàng này trước
                for _mr in list(ws.merged_cells.ranges):
                    if _mr.min_row <= r <= _mr.max_row and _mr.min_col >= c_from and _mr.max_col <= c_to:
                        ws.unmerge_cells(str(_mr))
                for _c in range(c_from, c_to + 1):
                    _cell = ws.cell(row=r, column=_c)
                    _cell.border = _no_brd
                    _cell.fill   = _white_fill
                    try:
                        _cell.value = None
                    except AttributeError:
                        pass

            # -- Tiêu đề bảng (xanh đậm) --
            _title_row = None
            if ten_en:
                _title_row = row
                ws.merge_cells(f"A{row}:F{row}")
                title_txt = f"{ten_en} / {ten_vi}" if ten_vi else ten_en
                tc = ws.cell(row=row, column=1, value=title_txt)
                tc.font      = Font(bold=True, size=16, color="000000",
                                    name="Times New Roman")
                tc.alignment = Alignment(horizontal="center", vertical="center")
                tc.fill      = PatternFill("solid", fgColor=BLUE_CELL)
                tc.border    = brd_thick
                _brd_merge(row, 1, 6, brd_thick)
                ws.row_dimensions[row].height = 26
                row += 1

            # -- Nhãn cột (xanh nhạt) --
            _label_row = row
            for lbl, (c1, c2) in zip(display_labels, ranges):
                if c1 != c2:
                    ws.merge_cells(
                        f"{get_column_letter(c1)}{row}:"
                        f"{get_column_letter(c2)}{row}")
                lc = ws.cell(row=row, column=c1, value=lbl)
                lc.font      = Font(bold=True, size=14, name="Times New Roman")
                lc.alignment = Alignment(horizontal="center", vertical="center",
                                         wrap_text=True)
                lc.fill      = PatternFill("solid", fgColor=BLUE_LABEL)
                lc.border    = brd_thick
                _brd_merge(row, c1, c2, brd_thick)
            ws.row_dimensions[row].height = 20
            row += 1

            # -- Hàng ảnh --
            _img_row = row
            ws.row_dimensions[row].height = img_row_h_pt
            for key, (c1, c2) in zip(storage_keys, ranges):
                if c1 != c2:
                    ws.merge_cells(
                        f"{get_column_letter(c1)}{row}:"
                        f"{get_column_letter(c2)}{row}")

                img_bio = _img_cache.get(key)
                if img_bio is not None:
                    # Có ảnh → vẽ viền bình thường
                    ws.cell(row=row, column=c1).border = brd_thick
                    _brd_merge(row, c1, c2, brd_thick)
                else:
                    # Không có ảnh → ẩn ô (bỏ viền + fill trắng)
                    for _c in range(c1, c2 + 1):
                        _cell = ws.cell(row=row, column=_c)
                        _cell.border = _no_brd
                        _cell.fill   = _white_fill

                if img_bio is not None:
                    try:
                        img_bio.seek(0)
                        xl_img = XLImage(io.BytesIO(img_bio.read()))
                        # Dùng TwoCellAnchor để ảnh căng đúng khung ô
                        PAD = 9144   # ~1pt padding (đơn vị EMU)
                        _anchor = TwoCellAnchor(editAs="twoCell")
                        _anchor._from = AnchorMarker(
                            col=c1 - 1, colOff=PAD,
                            row=row - 1, rowOff=PAD)
                        _anchor.to = AnchorMarker(
                            col=c2, colOff=-PAD,
                            row=row, rowOff=-PAD)
                        xl_img.anchor = _anchor
                        ws.add_image(xl_img)
                    except Exception:
                        pass

            # Nếu toàn bộ bảng không có ảnh → ẩn luôn tiêu đề + nhãn
            if not _has_any_img:
                if _title_row:
                    _blank_row(_title_row, 1, 6)
                    ws.row_dimensions[_title_row].height = 0
                _blank_row(_label_row, 1, 6)
                ws.row_dimensions[_label_row].height = 0
                ws.row_dimensions[_img_row].height = 0
            else:
                _apply_outer_thick(_tbl_start, row)   # khung đậm bên ngoài toàn bảng
            row += 1
            ws.row_dimensions[row].height = 5   # khoảng trắng nhỏ giữa các bảng
            row += 1  # trống giữa các bảng

        ws.row_breaks.append(Break(id=row - 1))  # ── page break: mỗi trang ảnh

    # ── Page setup: A4 portrait, fit 1 trang ngang ──────────────
    from openpyxl.worksheet.page import PageMargins
    ws.page_setup.paperSize        = 9          # A4
    ws.page_setup.orientation      = "portrait"
    ws.page_setup.fitToPage        = True
    ws.page_setup.fitToWidth       = 1          # scale vừa 1 trang ngang
    ws.page_setup.fitToHeight      = 0          # chiều cao tự nhiên (nhiều trang)
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    ws.page_margins = PageMargins(
        left=0.39, right=0.39,
        top=0.59,  bottom=0.59,
        header=0.2, footer=0.2,
    )

    # ── Xuất bytes ──────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ============================================================
# FRAGMENT: BADGE + SELECTBOX TRẠNG THÁI (dùng trong dialog, rerun riêng)
# ============================================================
@st.fragment
def _fragment_trang_thai_dialog(hang: dict, ds_trang_thai: list):
    """Badge màu + selectbox đổi trạng thái — đặt ở đầu _task_dialog.
    Vì là @st.fragment riêng nên badge cập nhật ngay khi selectbox thay đổi.
    """
    try:
        task_id = int(float(hang.get("ID", 0)))
    except (ValueError, TypeError):
        task_id = 0
    trang_thai = hang.get("Trạng Thái", "Chờ Làm")
    _MAU = {
        "Đang Kiểm Tra":           ("#1d4ed8", "#dbeafe"),
        "Đã Phê Duyệt":            ("#15803d", "#dcfce7"),
        "Đã Báo Giá":              ("#b45309", "#fef3c7"),
        "Có Đơn":                  ("#7c3aed", "#ede9fe"),
        "Chờ Giao":                ("#a16207", "#fef9c3"),
        "Đã Hoàn Thành - Giao Máy":("#166534", "#bbf7d0"),
        "Đã Xuất Hóa Đơn":        ("#374151", "#f3f4f6"),
        "Bảo Hành - Trả Lại":     ("#b91c1c", "#fee2e2"),
        "Chờ Làm":                 ("#dc2626", "#fee2e2"),
        "Đang Làm":                ("#d97706", "#fef3c7"),
        "Hoàn Thành":              ("#16a34a", "#dcfce7"),
    }
    # Track trạng thái đã lưu để tránh gửi background thread nhiều lần
    _saved_key = f"_dlg_saved_tt_{task_id}"
    if _saved_key not in st.session_state:
        st.session_state[_saved_key] = trang_thai

    tt_cur = st.session_state.get(f"tt_select_dlg_{task_id}", st.session_state[_saved_key])
    tt_color, tt_bg = _MAU.get(tt_cur, ("#6b7280", "#f3f4f6"))
    st.markdown(
        f'<div style="margin:0 0 6px 0">'
        f'<span style="display:inline-block;padding:5px 14px;border-radius:20px;'
        f'font-size:0.82rem;font-weight:700;letter-spacing:0.3px;'
        f'border:1.5px solid {tt_color}40;color:{tt_color};background:{tt_bg};">'
        f'{tt_cur}</span></div>',
        unsafe_allow_html=True,
    )
    # Chỉ lấy index khi trạng thái hiện tại có trong danh sách;
    # nếu không có → giữ nguyên _saved_key, KHÔNG để index=0 tự chọn sai
    _saved_val = st.session_state[_saved_key]
    if _saved_val in ds_trang_thai:
        _idx = ds_trang_thai.index(_saved_val)
    else:
        # Trạng thái task không có trong danh sách → thêm tạm vào đầu để không mất dữ liệu
        ds_trang_thai = [_saved_val] + ds_trang_thai
        _idx = 0

    def _on_tt_change():
        _new = st.session_state.get(f"tt_select_dlg_{task_id}")
        _old = st.session_state.get(_saved_key)
        if _new and _new != _old:
            _ten_task  = hang.get("Tên Công Việc", "")
            _nguoi_doi = st.session_state.get("ho_ten", "")
            _ct        = hang.get("Công Ty", "")
            _tb_tt = (
                f"🔄 **{_nguoi_doi}** đã chuyển công việc **{_ten_task}**"
                + (f" ({_ct})" if _ct else "")
                + f" từ **{_old}** → **{_new}**"
            )
            st.session_state[_saved_key] = _new
            if "_tt_overrides" not in st.session_state:
                st.session_state["_tt_overrides"] = {}
            st.session_state["_tt_overrides"][task_id] = _new
            st.session_state[f"_tt_toast_{task_id}"] = _new
            def _bg_doi_tt(_tid, _tt, _msg, _tid2):
                cap_nhat_trang_thai(_tid, _tt)
                lay_danh_sach_cong_viec.clear()
                them_thong_bao_tat_ca(_msg, task_id=_tid2, loai="trang_thai", tru_nguoi="")
            threading.Thread(target=_bg_doi_tt, args=(task_id, _new, _tb_tt, task_id), daemon=True).start()

    st.selectbox(
        "Chọn trạng thái",
        options=ds_trang_thai,
        index=_idx,
        key=f"tt_select_dlg_{task_id}",
        label_visibility="collapsed",
        on_change=_on_tt_change,
    )
    _toast_val = st.session_state.pop(f"_tt_toast_{task_id}", None)
    if _toast_val:
        st.toast(f"✅ Đã chuyển → **{_toast_val}**")


# ============================================================
# FRAGMENT: CHI TIẾT TASK NHÂN VIÊN (rerun cục bộ khi tick checkbox / đổi trạng thái)
# ============================================================
@st.fragment
def _fragment_chi_tiet_task(hang: dict, ds_trang_thai: list, show_status: bool = True):
    """
    Hiển thị toàn bộ thông tin một task cho nhân viên:
    - Thông tin đầy đủ (công ty, công số, năm, mô tả, hạn, người phê duyệt)
    - Dropdown chọn trạng thái (lưu ngay khi đổi) với CSS badge màu
    - Checklist tương tác (tick → lưu lên Google Sheets)
    - Công việc con (card đẹp)
    - Upload ảnh + xuất PDF
    Dùng @st.fragment nên chỉ rerun vùng này, không reload toàn trang.
    """
    # ── CSS toàn cục cho fragment này ────────────────────────
    st.markdown("""
    <style>
    /* Badge trạng thái */
    .badge-trang-thai {
        display: inline-block;
        padding: 5px 14px;
        border-radius: 20px;
        font-size: 0.82rem;
        font-weight: 700;
        letter-spacing: 0.3px;
        border: 1.5px solid;
        white-space: nowrap;
    }
    /* Card công việc con */
    .cv-con-card {
        background: #f8f7ff;
        border: 1px solid #e0dcff;
        border-radius: 12px;
        padding: 8px 12px;
        margin-bottom: 4px;
    }
    .cv-con-card.done {
        background: #f0fdf4;
        border-color: #bbf7d0;
        opacity: 0.75;
    }
    .cv-con-meta {
        font-size: 0.8rem;
        color: #6b7280;
        display: flex;
        gap: 12px;
        flex-wrap: wrap;
        margin-top: 2px;
        margin-left: 28px;
        padding-bottom: 4px;
    }
    .cv-con-meta span { display:inline-flex; align-items:center; gap:3px; }
    /* Checklist done: giữ nguyên chữ, chỉ tick màu đỏ */
    /* Ghost delete button */
    button[title="Xóa"], button[title="Xóa công việc này"] {
        background: transparent !important;
        border: none !important;
        color: #d1d5db !important;
        box-shadow: none !important;
    }
    /* Fix input bị cắt đáy trong column */
    [data-testid="stColumn"] > [data-testid="stVerticalBlock"] {
        overflow: visible !important;
        padding-bottom: 4px !important;
    }
    </style>
    """, unsafe_allow_html=True)

    # ── Bảng màu trạng thái ───────────────────────────────────
    _MAU_TRANG_THAI = {
        "Đang Kiểm Tra":           ("#1d4ed8", "#dbeafe"),   # xanh dương
        "Đã Phê Duyệt":            ("#15803d", "#dcfce7"),   # xanh lá
        "Đã Báo Giá":              ("#b45309", "#fef3c7"),   # cam
        "Có Đơn":                  ("#7c3aed", "#ede9fe"),   # tím
        "Chờ Giao":                ("#a16207", "#fef9c3"),   # vàng
        "Đã Hoàn Thành - Giao Máy":("#166534", "#bbf7d0"),   # xanh đậm
        "Đã Xuất Hóa Đơn":        ("#374151", "#f3f4f6"),   # xám
        "Bảo Hành - Trả Lại":     ("#b91c1c", "#fee2e2"),   # đỏ
        "Chờ Làm":                 ("#dc2626", "#fee2e2"),
        "Đang Làm":                ("#d97706", "#fef3c7"),
        "Hoàn Thành":              ("#16a34a", "#dcfce7"),
    }

    try:
        task_id = int(float(hang.get("ID", 0)))
    except (ValueError, TypeError):
        task_id = 0
    trang_thai = hang.get("Trạng Thái", "Chờ Làm")

    # Khởi đầu danh sách ảnh từ hang (nếu chưa có trong session_state)
    _anh_key = f"anh_editable_{task_id}"
    if _anh_key not in st.session_state:
        st.session_state[_anh_key] = doc_danh_sach_anh(str(hang.get("Link Ảnh", "")))

    if show_status:
        # ── Badge + selectbox trạng thái (khi hiển thị standalone, không qua dialog) ──
        tt_hien_thi = st.session_state.get(f"tt_select_{task_id}", trang_thai)
        tt_color, tt_bg = _MAU_TRANG_THAI.get(tt_hien_thi, ("#6b7280", "#f3f4f6"))
        st.markdown(
            f'<div style="margin:0 0 6px 0">'
            f'<span class="badge-trang-thai" style="color:{tt_color};background:{tt_bg};border-color:{tt_color}40;">'
            f'{tt_hien_thi}</span></div>',
            unsafe_allow_html=True,
        )
        if trang_thai in ds_trang_thai:
            idx_hien_tai   = ds_trang_thai.index(trang_thai)
            _ds_tt_display = ds_trang_thai
        else:
            _ds_tt_display = [trang_thai] + ds_trang_thai
            idx_hien_tai   = 0

        def _on_tt_nv_change():
            _new = st.session_state.get(f"tt_select_{task_id}")
            if _new and _new != trang_thai:
                _ten_task  = hang.get("Tên Công Việc", "")
                _nguoi_doi = st.session_state.get("ho_ten", "")
                _ct        = hang.get("Công Ty", "")
                _tb_tt = (
                    f"🔄 **{_nguoi_doi}** đã chuyển công việc "
                    f"**{_ten_task}**"
                    + (f" ({_ct})" if _ct else "")
                    + f" từ **{trang_thai}** → **{_new}**"
                )
                st.session_state[f"_tt_toast_nv_{task_id}"] = _new
                def _bg_doi_tt(_tid, _tt, _msg, _tid2):
                    cap_nhat_trang_thai(_tid, _tt)
                    lay_danh_sach_cong_viec.clear()
                    them_thong_bao_tat_ca(_msg, task_id=_tid2, loai="trang_thai", tru_nguoi="")
                threading.Thread(target=_bg_doi_tt, args=(task_id, _new, _tb_tt, task_id), daemon=True).start()

        st.selectbox(
            "Chọn trạng thái",
            options=_ds_tt_display,
            index=idx_hien_tai,
            key=f"tt_select_{task_id}",
            label_visibility="collapsed",
            on_change=_on_tt_nv_change,
        )
        _toast_nv = st.session_state.pop(f"_tt_toast_nv_{task_id}", None)
        if _toast_nv:
            st.toast(f"✅ Đã chuyển → **{_toast_nv}**")

    st.markdown("---")
    # Công ty, Công số, Năm, Hạn, Ngày tạo
    st.markdown(f"🏢 **Công ty:** `{hang.get('Công Ty', '')}`")
    _cong_so_hien = hang.get('Công Số', '')
    _nam_hien     = hang.get('Năm', '')
    _dong_cs = []
    if _cong_so_hien:
        _dong_cs.append(f"📄 **Công số:** `{_cong_so_hien}`")
    if _nam_hien:
        _dong_cs.append(f"📅 **Năm:** `{_nam_hien}`")
    if _dong_cs:
        st.markdown(" &nbsp;·&nbsp; ".join(_dong_cs))
    # ── Hạn Hoàn Thành (nhân viên có thể chỉnh) ─────────────────────────────
    _hht_hien_tai = hang.get("Hạn Hoàn Thành", "") or ""
    try:
        from datetime import date as _date
        _hht_default = _date.fromisoformat(_hht_hien_tai.strip()) if _hht_hien_tai.strip() else None
    except Exception:
        _hht_default = None

    _hht_val = st.date_input(
        "📅 Hạn hoàn thành",
        value=_hht_default,
        format="YYYY-MM-DD",
        key=f"hht_{task_id}",
    )
    _hht_str = _hht_val.strftime("%Y-%m-%d") if _hht_val else ""
    if _hht_str != _hht_hien_tai.strip():
        with st.spinner("Đang lưu..."):
            cap_nhat_han_hoan_thanh(task_id, _hht_str)

    st.markdown(f"🕐 **Ngày tạo:** `{hang.get('Ngày Tạo', '')}`")

    # ── Ngày Kết Thúc (nhân viên tự điền) ───────────────────────────────────
    _nkt_hien_tai = hang.get("Ngày Kết Thúc", "") or ""
    try:
        from datetime import date as _date
        _nkt_default = _date.fromisoformat(_nkt_hien_tai.strip()) if _nkt_hien_tai.strip() else None
    except Exception:
        _nkt_default = None

    _nkt_val = st.date_input(
        "📅 Ngày kết thúc",
        value=_nkt_default,
        format="YYYY-MM-DD",
        key=f"nkt_{task_id}",
    )
    _nkt_str = _nkt_val.strftime("%Y-%m-%d") if _nkt_val else ""
    if _nkt_str != _nkt_hien_tai.strip():
        with st.spinner("Đang lưu..."):
            cap_nhat_ngay_ket_thuc(task_id, _nkt_str)

    # ── Thông số kỹ thuật / thương mại (có thể chỉnh sửa) ───────────────────
    st.markdown("---")
    st.markdown("**🔧 Thông số kỹ thuật**")

    _ds_loai_may   = [""] + lay_ten_cac_loai_may()
    _ds_tinh_trang = [""] + lay_ten_cac_tinh_trang()

    _lm_cur  = hang.get("Loại Máy", "") or ""
    _tt_cur  = hang.get("Tình Trạng", "") or ""
    _cs_cur  = hang.get("Công Suất", "") or ""
    _sc_cur  = hang.get("Số Cực", "") or ""
    _ms_cur  = hang.get("Mã Số", "") or ""
    _po_cur  = hang.get("Số PO Nội Bộ", "") or ""
    _pkh_cur = hang.get("Số PO KH/HĐ", "") or ""
    _bg_cur  = hang.get("Số Báo Giá", "") or ""

    _col1, _col2 = st.columns(2)
    with _col1:
        _lm_idx = _ds_loai_may.index(_lm_cur) if _lm_cur in _ds_loai_may else 0
        _lm_new = st.selectbox("🔧 Loại Máy", _ds_loai_may, index=_lm_idx, key=f"edit_lm_{task_id}")
    with _col2:
        _tt_idx = _ds_tinh_trang.index(_tt_cur) if _tt_cur in _ds_tinh_trang else 0
        _tt_new = st.selectbox("🛠️ Tình Trạng", _ds_tinh_trang, index=_tt_idx, key=f"edit_tt_{task_id}")

    _col3, _col4 = st.columns(2)
    with _col3:
        _cs_new = st.text_input("⚡ Công Suất", value=_cs_cur, key=f"edit_cs_{task_id}")
    with _col4:
        _sc_new = st.text_input("🔩 Số Cực", value=_sc_cur, key=f"edit_sc_{task_id}")

    _col5, _col6 = st.columns(2)
    with _col5:
        _ms_new = st.text_input("🏷️ Mã Số", value=_ms_cur, key=f"edit_ms_{task_id}")
    with _col6:
        _po_new = st.text_input("📄 Số PO Nội Bộ", value=_po_cur, key=f"edit_po_{task_id}")

    _col7, _col8 = st.columns(2)
    with _col7:
        _pkh_new = st.text_input("📋 Số PO KH/HĐ", value=_pkh_cur, key=f"edit_pkh_{task_id}")
    with _col8:
        _bg_new = st.text_input("💰 Số Báo Giá", value=_bg_cur, key=f"edit_bg_{task_id}")

    _thay_doi = {
        k: v for k, v in {
            "Loại Máy":    _lm_new,
            "Tình Trạng":  _tt_new,
            "Công Suất":   _cs_new,
            "Số Cực":      _sc_new,
            "Mã Số":       _ms_new,
            "Số PO Nội Bộ":_po_new,
            "Số PO KH/HĐ": _pkh_new,
            "Số Báo Giá":  _bg_new,
        }.items()
        if v != hang.get(k, "") and not (v == "" and not hang.get(k, ""))
    }
    if _thay_doi:
        if st.button("💾 Lưu thông số", key=f"save_ts_{task_id}", type="primary", use_container_width=True):
            with st.spinner("Đang lưu..."):
                cap_nhat_nhieu_truong_task(task_id, _thay_doi)
            st.success("✅ Đã lưu!")

    st.divider()

    # ── Người Phê Duyệt ───────────────────────────────────────
    _pd_cur = hang.get("Người Phê Duyệt", "") or ""
    _ds_pd  = ["-- Không chọn --"] + lay_danh_sach_nhan_vien()
    _pd_idx = _ds_pd.index(_pd_cur) if _pd_cur in _ds_pd else 0

    def _luu_nguoi_pd():
        _val = st.session_state.get(f"edit_pd_{task_id}", "-- Không chọn --")
        _luu = _val if _val != "-- Không chọn --" else ""
        cap_nhat_nhieu_truong_task(task_id, {"Người Phê Duyệt": _luu})
        lay_danh_sach_cong_viec.clear()

    _pd_new = st.selectbox(
        "👤 Người Phê Duyệt", _ds_pd, index=_pd_idx,
        key=f"edit_pd_{task_id}",
        on_change=_luu_nguoi_pd,
    )
    _pd_chon = _pd_new if _pd_new != "-- Không chọn --" else ""

    _dang_kiem_tra = (trang_thai == "Đang Kiểm Tra")
    _co_nguoi_pd   = bool(_pd_chon)

    def _on_gui_pd_change():
        _checked = st.session_state.get(f"gui_pd_{task_id}", False)
        if _checked:
            cap_nhat_nhieu_truong_task(task_id, {"Người Phê Duyệt": _pd_chon})
            cap_nhat_trang_thai(task_id, "Đang Kiểm Tra")
            lay_danh_sach_cong_viec.clear()

    st.checkbox(
        "Gửi cho Người Phê Duyệt",
        value=_dang_kiem_tra,
        disabled=not _co_nguoi_pd,
        key=f"gui_pd_{task_id}",
        on_change=_on_gui_pd_change,
    )

    st.divider()

    # ── Checklist ─────────────────────────────────────────────
    raw_cl = hang.get("Checklist", "") or "[]"
    try:
        _cl_parsed = json.loads(raw_cl) if raw_cl else []
    except Exception:
        _cl_parsed = []

    _cl_key = f"cl_editable_{task_id}"
    if _cl_key not in st.session_state:
        st.session_state[_cl_key] = [
            ({"text": x, "done": False} if isinstance(x, str) else x)
            for x in _cl_parsed
        ]
    _checklist = st.session_state[_cl_key]

    so_xong = sum(1 for it in _checklist if isinstance(it, dict) and it.get("done"))

    _cl_open_key = f"cl_open_{task_id}"
    if _cl_open_key not in st.session_state:
        st.session_state[_cl_open_key] = False  # mặc định đóng

    _cl_chevron = "▼" if st.session_state[_cl_open_key] else "▶"
    def _toggle_cl():
        st.session_state[_cl_open_key] = not st.session_state[_cl_open_key]
    st.button(
        f"{_cl_chevron}  ☑️ Checklist  {so_xong}/{len(_checklist)} mục",
        key=f"dlg_cl_tog_{task_id}",
        use_container_width=True,
        on_click=_toggle_cl,
    )

    if st.session_state[_cl_open_key]:
        _mk_cl = f"dlgcl{task_id}"
        st.markdown(f"""<style>
        [data-testid="stHorizontalBlock"]:has(.{_mk_cl}) {{
            flex-wrap: nowrap !important; align-items: center !important; gap: 6px !important;
        }}
        [data-testid="stHorizontalBlock"]:has(.{_mk_cl}) > [data-testid="stColumn"]:first-child {{
            flex: 0 0 36px !important; min-width: 36px !important; max-width: 36px !important;
        }}
        [data-testid="stHorizontalBlock"]:has(.{_mk_cl}) > [data-testid="stColumn"]:last-child {{
            flex: 1 1 0% !important; min-width: 0 !important;
        }}
        [data-testid="stHorizontalBlock"]:has(.{_mk_cl}) [data-testid="stCheckbox"] {{ margin:0 !important; }}
        [data-testid="stHorizontalBlock"]:has(.{_mk_cl}) p {{
            margin: 0 !important; font-size: 0.93rem !important; color: #1e1b4b;
            padding: 4px 2px;
        }}
        </style>""", unsafe_allow_html=True)

        def _cb_cl_done(idx):
            val = st.session_state.get(f"dlg_ck_{task_id}_{idx}", False)
            st.session_state[_cl_key][idx]["done"] = val
            cap_nhat_checklist(task_id, st.session_state[_cl_key])

        def _cb_cl_text(idx):
            val = st.session_state.get(f"dlg_cl_txt_{task_id}_{idx}", "")
            st.session_state[_cl_key][idx]["text"] = val
            cap_nhat_checklist(task_id, st.session_state[_cl_key])

        for _ci, _cit in enumerate(_checklist):
            if isinstance(_cit, str):
                _cit = {"text": _cit, "done": False}
                _checklist[_ci] = _cit
            _done_v = bool(_cit.get("done", False))
            _txt0   = _cit.get("text", "") or f"Mục {_ci+1}"
            col_ck, col_txt = st.columns([0.5, 9], gap="small")
            with col_ck:
                st.markdown(f"<span class='{_mk_cl}' style='display:none'></span>", unsafe_allow_html=True)
                st.checkbox("", value=_done_v, key=f"dlg_ck_{task_id}_{_ci}",
                            label_visibility="collapsed",
                            on_change=_cb_cl_done, args=(_ci,))
            with col_txt:
                st.text_input(
                    "", value=_txt0,
                    key=f"dlg_cl_txt_{task_id}_{_ci}",
                    label_visibility="collapsed",
                    on_change=_cb_cl_text, args=(_ci,),
                )

        # ── Thêm mục checklist mới ────────────────────────────────
        _cl_new_key = f"dlg_cl_new_{task_id}"
        _cl_new_v   = f"dlg_cl_new_v_{task_id}"
        if _cl_new_v not in st.session_state:
            st.session_state[_cl_new_v] = 0

        col_inp, col_add = st.columns([8, 1], gap="small")
        with col_inp:
            st.text_input(
                "", placeholder="➕ Thêm mục checklist...",
                key=f"{_cl_new_key}_{st.session_state[_cl_new_v]}",
                label_visibility="collapsed",
            )
        with col_add:
            def _cb_cl_add():
                _txt = st.session_state.get(f"{_cl_new_key}_{st.session_state[_cl_new_v]}", "").strip()
                if _txt:
                    st.session_state[_cl_key].append({"text": _txt, "done": False})
                    cap_nhat_checklist(task_id, st.session_state[_cl_key])
                    st.session_state[_cl_new_v] += 1  # reset input
            st.button("➕", key=f"dlg_cl_add_btn_{task_id}", on_click=_cb_cl_add, use_container_width=True)

    st.divider()

    # ── Khởi tạo _do_key (ảnh đo lường) — cần trước vòng lặp CVC ──
    _do_key = f"do_luong_{task_id}"
    if _do_key not in st.session_state:
        st.session_state[_do_key] = doc_anh_do_luong(str(hang.get("Ảnh Đo Lường", "") or ""))

    # ── Công việc con ─────────────────────────────────────────
    raw_cv = hang.get("Công Việc Con", "") or "[]"
    try:
        _cv_parsed = json.loads(raw_cv) if raw_cv else []
    except Exception:
        _cv_parsed = []

    _cv_key = f"cv_editable_{task_id}"
    if _cv_key not in st.session_state:
        if _cv_parsed:
            st.session_state[_cv_key] = [
                {
                    "ten":              cv.get("ten", cv.get("Tên", "")),
                    "nhan_vien":        cv.get("nhan_vien", cv.get("Nhân Viên", cv.get("nguoi", ""))),
                    "done":             bool(cv.get("done", False)),
                    "anh":              cv.get("anh", []),
                    "ngay_hoan_thanh":  cv.get("ngay_hoan_thanh", ""),
                }
                for cv in _cv_parsed if isinstance(cv, dict)
            ]
        else:
            # CVC rỗng trong sheet — restore từ template mặc định và ghi lại sheet
            _ds_cd = lay_ten_cac_cong_doan()
            st.session_state[_cv_key] = [
                {"ten": cd, "nhan_vien": "", "done": False, "anh": [], "ngay_hoan_thanh": ""}
                for cd in _ds_cd
            ]
            try:
                _save_cv_to_sheet(task_id, _cv_key)
            except Exception:
                pass
    ds_cv_con = st.session_state[_cv_key]
    _ds_nv_cv = ["-- Không chọn --"] + lay_danh_sach_nhan_vien()



    _done_cv = sum(1 for cv in ds_cv_con if cv.get("done"))
    st.markdown(
        f"**📋 Công Việc Con** &nbsp;<span style='color:#6b7280;font-size:0.82rem;'>{_done_cv}/{len(ds_cv_con)} hoàn thành</span>",
        unsafe_allow_html=True,
    )

    _mk_cv = f"dlgcv{task_id}"
    st.markdown(f"""<style>
    [data-testid="stHorizontalBlock"]:has(.{_mk_cv}) {{
        flex-wrap: nowrap !important; align-items: center !important; gap: 6px !important;
        margin-bottom: 2px !important;
    }}
    [data-testid="stHorizontalBlock"]:has(.{_mk_cv}) > [data-testid="stColumn"]:nth-child(1) {{
        flex: 0 0 32px !important; min-width: 32px !important; max-width: 32px !important;
    }}
    [data-testid="stHorizontalBlock"]:has(.{_mk_cv}) > [data-testid="stColumn"]:nth-child(2) {{
        flex: 1 1 0% !important; min-width: 0 !important;
    }}
    [data-testid="stHorizontalBlock"]:has(.{_mk_cv}) > [data-testid="stColumn"]:nth-child(3) {{
        flex: 0 0 36px !important; min-width: 36px !important; max-width: 36px !important;
    }}
    [data-testid="stHorizontalBlock"]:has(.{_mk_cv}) [data-testid="stCheckbox"] {{ margin: 0 !important; }}
    [data-testid="stHorizontalBlock"]:has(.{_mk_cv}) p {{
        margin: 0 !important; font-size: 0.93rem !important; font-weight: 600;
        color: #4c1d95; padding: 2px 0;
    }}
    [data-testid="stHorizontalBlock"]:has(.{_mk_cv}) [data-testid="stColumn"]:nth-child(3) button {{
        background: transparent !important; border: none !important;
        box-shadow: none !important; color: #ef4444 !important;
        font-size: 1.1rem !important; padding: 2px 4px !important; min-height: 32px !important;
    }}
    [data-testid="stHorizontalBlock"]:has(.{_mk_cv}) [data-testid="stColumn"]:nth-child(3) button:hover {{
        background: #fee2e2 !important; border-radius: 6px !important;
    }}
    .cv-nv-row-{task_id} [data-testid="stSelectbox"] > div > div {{
        font-size: 0.9rem !important; border-color: #e0d7ff !important;
        background: #f5f3ff !important; color: #4c1d95 !important;
        min-height: 34px !important;
    }}
    .cv-nv-row-{task_id} [data-testid="stSelectbox"] [class*="singleValue"],
    .cv-nv-row-{task_id} [data-testid="stSelectbox"] [class*="SingleValue"] {{
        direction: ltr !important; font-size: 0.9rem !important;
        overflow: visible !important; white-space: nowrap !important;
        text-overflow: clip !important; max-width: none !important;
    }}
    .cv-nv-row-{task_id} label {{ font-size: 0.78rem !important; color: #6b7280 !important; }}
    .cv-item-wrap-{task_id} {{
        border-bottom: 2px solid #c7d2fe;
        padding: 4px 0 8px 0; margin-bottom: 6px;
    }}
    .cv-item-wrap-{task_id}.done {{ opacity: 0.6; }}
    [data-testid="stHorizontalBlock"]:has(.{_mk_cv}) input[type="text"] {{
        color: #4c1d95 !important; font-weight: 700 !important;
        font-size: 0.95rem !important; border-color: #ddd6fe !important;
    }}
    </style>""", unsafe_allow_html=True)

    def _cb_cv_done(i):
        val = st.session_state.get(f"dlg_cv_ck_{task_id}_{i}", False)
        st.session_state[_cv_key][i]["done"] = val
        if val:
            st.session_state[_cv_key][i]["ngay_hoan_thanh"] = datetime.now().strftime("%Y-%m-%d")
        else:
            st.session_state[_cv_key][i].pop("ngay_hoan_thanh", None)
        _save_cv_to_sheet(task_id, _cv_key)

    def _cb_cv_nv(i):
        val    = st.session_state.get(f"dlg_cv_nv_{task_id}_{i}", "-- Không chọn --")
        nv_cu  = st.session_state[_cv_key][i].get("nhan_vien", "")
        nv_moi = "" if val == "-- Không chọn --" else val
        st.session_state[_cv_key][i]["nhan_vien"] = nv_moi
        # Gửi thông báo nếu nhân viên thay đổi
        if nv_moi and nv_moi != nv_cu:
            ten_cv      = st.session_state[_cv_key][i].get("ten", "")
            nguoi_giao  = st.session_state.get("ho_ten", "")
            ten_task_tb = hang.get("Tên Công Việc", f"Task #{task_id}")
            cong_ty_tb  = hang.get("Công Ty", "")
            _suffix     = f"**{ten_task_tb}**" + (f" ({cong_ty_tb})" if cong_ty_tb else "")
            them_thong_bao(nv_moi,
                f"🔧 Bạn được giao việc: **{ten_cv}** • {_suffix}",
                task_id=task_id, loai="cong_viec_con")
            them_thong_bao_tat_ca(
                f"🔧 **{nguoi_giao}** đã giao **{ten_cv}** cho **{nv_moi}** • {_suffix}",
                task_id=task_id, loai="cong_viec_con", tru_nguoi=nv_moi)
        _save_cv_to_sheet(task_id, _cv_key)

    def _cb_cv_del(i):
        if 0 <= i < len(st.session_state[_cv_key]):
            st.session_state[_cv_key].pop(i)
            for k in list(st.session_state.keys()):
                if k.startswith(f"dlg_cv_nv_{task_id}_"):
                    del st.session_state[k]
            _save_cv_to_sheet(task_id, _cv_key)

    def _cb_del_cv_media(cvi, url_m):
        _cvl = st.session_state.get(_cv_key, [])
        if 0 <= cvi < len(_cvl):
            _cvl[cvi]["anh"] = [u for u in _cvl[cvi].get("anh", []) if u != url_m]
            _save_cv_to_sheet(task_id, _cv_key)

    for _cvi, cv in enumerate(ds_cv_con):
        if not isinstance(cv, dict): continue
        _tcv  = cv.get("ten") or f"Việc {_cvi+1}"
        _nvcv = cv.get("nhan_vien", "") or ""
        _dcv  = bool(cv.get("done", False))
        _nv_idx = _ds_nv_cv.index(_nvcv) if _nvcv in _ds_nv_cv else 0
        _wrap_cls = f"cv-item-wrap-{task_id}" + (" done" if _dcv else "")

        st.markdown(f"<div class='{_wrap_cls}'>", unsafe_allow_html=True)
        # Hàng 1: checkbox + tên công việc + nút xóa
        col_ck, col_txt, col_del = st.columns([0.4, 5, 0.5], gap="small")
        with col_ck:
            st.markdown(f"<span class='{_mk_cv}' style='display:none'></span>", unsafe_allow_html=True)
            st.checkbox("", value=_dcv, key=f"dlg_cv_ck_{task_id}_{_cvi}",
                        label_visibility="collapsed",
                        on_change=_cb_cv_done, args=(_cvi,))
        with col_txt:
            _ten_key = f"dlg_cv_ten_{task_id}_{_cvi}"
            _ten_input = st.text_input(
                "Tên công đoạn", value=_tcv, key=_ten_key,
                label_visibility="collapsed",
            )
            if _ten_input.strip() and _ten_input.strip() != _tcv:
                def _cb_luu_ten(_cvi=_cvi, _ten_key=_ten_key):
                    _new = st.session_state.get(_ten_key, "").strip()
                    if _new:
                        st.session_state[_cv_key][_cvi]["ten"] = _new
                        _save_cv_to_sheet(task_id, _cv_key)
                st.button("💾 Lưu tên", key=f"dlg_cv_luu_ten_{task_id}_{_cvi}",
                          on_click=_cb_luu_ten, use_container_width=True)
        with col_del:
            st.button("🗑️", key=f"dlg_cv_del_{task_id}_{_cvi}",
                      use_container_width=True,
                      on_click=_cb_cv_del, args=(_cvi,))
        # Hàng 2: selectbox nhân viên full width
        st.markdown(f"<div class='cv-nv-row-{task_id}'>", unsafe_allow_html=True)
        st.selectbox("👤 Nhân viên thực hiện", options=_ds_nv_cv, index=_nv_idx,
                     key=f"dlg_cv_nv_{task_id}_{_cvi}",
                     on_change=_cb_cv_nv, args=(_cvi,))
        st.markdown("</div>", unsafe_allow_html=True)
        st.markdown("</div>", unsafe_allow_html=True)

        # ── Hình / Video cho công việc con ────────────────────
        _cv_media = cv.get("anh", [])
        _exp_lbl = f"📎 Hình/Video ({len(_cv_media)})" if _cv_media else "📎 Thêm hình/video"
        _exp_open_key = f"exp_cv_m_open_{task_id}_{_cvi}"
        if _exp_open_key not in st.session_state:
            st.session_state[_exp_open_key] = False
        with st.expander(_exp_lbl, expanded=st.session_state[_exp_open_key]):
            if _cv_media:
                _cols_m = st.columns(min(len(_cv_media), 3))
                for _mi, _url_m in enumerate(_cv_media):
                    with _cols_m[_mi % 3]:
                        if "/view" in _url_m:
                            st.markdown(f"🎬 [Video {_mi + 1}]({_url_m})", unsafe_allow_html=False)
                        else:
                            _hien_thi_anh_drive(_url_m, use_container_width=True)
                        st.button("🗑️ Xoá", key=f"del_cv_m_{task_id}_{_cvi}_{_mi}",
                                  use_container_width=True,
                                  on_click=_cb_del_cv_media, args=(_cvi, _url_m))
                # Nút tải ZIP (chỉ ảnh, bỏ qua video)
                _only_anh = [u for u in _cv_media if "/view" not in u]
                if _only_anh:
                    if st.button("⬇️ Tải tất cả ảnh (ZIP)", key=f"btn_zip_cvm_{task_id}_{_cvi}", use_container_width=True):
                        with st.spinner(f"Đang tải {len(_only_anh)} ảnh..."):
                            _zip_data = _tao_zip_anh(_only_anh, ten_file=f"task_{task_id}_cv{_cvi+1}")
                        st.download_button(
                            "💾 Tải xuống ZIP",
                            data=_zip_data,
                            file_name=f"task_{task_id}_cv{_cvi+1}.zip",
                            mime="application/zip",
                            key=f"dl_zip_cvm_{task_id}_{_cvi}",
                            use_container_width=True,
                        )
            _up_key_m = f"up_cv_m_{task_id}_{_cvi}_{st.session_state.get(f'up_cv_m_v_{task_id}_{_cvi}', 0)}"
            st.file_uploader(
                "Chọn hình hoặc video",
                type=["jpg", "jpeg", "png", "mp4", "mov", "avi"],
                accept_multiple_files=True,
                key=_up_key_m,
                label_visibility="collapsed",
            )
            def _cb_up_cv_m(_tid=task_id, _ci=_cvi, _upk=_up_key_m, _cvk=_cv_key):
                _files_m = st.session_state.get(_upk) or []
                if not _files_m:
                    return
                _cvl = st.session_state.get(_cvk, [])
                _ms = _lay_ma_so_tu_task_id(_tid)
                for _f in _files_m:
                    try:
                        _u = _tai_media_len_drive(_f, ma_so=_ms, task_id=str(_tid))
                        _cvl[_ci].setdefault("anh", []).append(_u)
                    except Exception:
                        pass
                _save_cv_to_sheet(_tid, _cvk)  # gọi trực tiếp để đọc session_state đúng context
                _vk = f"up_cv_m_v_{_tid}_{_ci}"
                st.session_state[_vk] = st.session_state.get(_vk, 0) + 1
                st.session_state[f"exp_cv_m_open_{_tid}_{_ci}"] = True  # giữ expander mở sau upload
            st.button("📤 Tải ảnh", key=f"btn_up_cv_m_{task_id}_{_cvi}", use_container_width=True,
                      disabled=not bool(st.session_state.get(_up_key_m)),
                      on_click=_cb_up_cv_m)

        # ── Ảnh đo lường thuộc công đoạn này ────────────────
        import unicodedata as _ucd
        _cv_name_upper = _ucd.normalize("NFC", _tcv.strip()).upper()
        _cv_do_slots   = _STAGE_DO_LUONG.get(_cv_name_upper)
        if _cv_do_slots:
            _do_open_key = f"cv_do_open_{task_id}_{_cvi}"
            if _do_open_key not in st.session_state:
                st.session_state[_do_open_key] = False
            _n_do = sum(
                len(st.session_state.get(_do_key, {}).get(lk, []))
                for entry in _cv_do_slots for _, lk in entry[1]
                if lk not in _DO_LUONG_NO_IMG_KEYS
            )
            _total_do = sum(
                len([lk for _, lk in entry[1] if lk not in _DO_LUONG_NO_IMG_KEYS])
                for entry in _cv_do_slots
            )
            _do_chevron = "▼" if st.session_state[_do_open_key] else "▶"
            _do_btn_lbl = (
                f"{_do_chevron}  📐 Ảnh Đo Lường  ({_n_do}/{_total_do} ảnh)"
            )
            def _toggle_cv_do(_k=_do_open_key):
                st.session_state[_k] = not st.session_state[_k]
            st.button(_do_btn_lbl, key=f"cv_do_tog_{task_id}_{_cvi}",
                      use_container_width=True, on_click=_toggle_cv_do)
            if st.session_state[_do_open_key]:
                _render_do_luong_inline(task_id, _do_key, _cv_do_slots, cvi=_cvi)

    # ── Thêm công việc con — nhập tay tên công đoạn ──────────
    _cv_add_v  = f"dlg_cv_add_v_{task_id}"
    if _cv_add_v not in st.session_state: st.session_state[_cv_add_v] = 0
    _v = st.session_state[_cv_add_v]

    col_cd, col_nv_add, col_btn_add = st.columns([3, 3, 1], vertical_alignment="bottom")
    with col_cd:
        st.text_input("Công đoạn", placeholder="Nhập tên công đoạn...",
                      key=f"dlg_cv_new_cd_{task_id}_{_v}")
    with col_nv_add:
        st.selectbox("Nhân viên", options=_ds_nv_cv,
                     key=f"dlg_cv_new_nv_{task_id}_{_v}",
                     label_visibility="visible")
    with col_btn_add:
        def _cb_cv_add():
            cd_v  = st.session_state.get(f"dlg_cv_new_cd_{task_id}_{_v}", "").strip()
            nv_v  = st.session_state.get(f"dlg_cv_new_nv_{task_id}_{_v}", "-- Không chọn --")
            if cd_v:
                nv_moi = "" if nv_v == "-- Không chọn --" else nv_v
                st.session_state[_cv_key].append({
                    "ten":       cd_v,
                    "nhan_vien": nv_moi,
                    "done":      False,
                })
                # Gửi thông báo cho nhân viên được giao + broadcast cho tất cả
                if nv_moi:
                    nguoi_giao  = st.session_state.get("ho_ten", "")
                    ten_task_tb = hang.get("Tên Công Việc", f"Task #{task_id}")
                    cong_ty_tb  = hang.get("Công Ty", "")
                    _suffix     = f"**{ten_task_tb}**" + (f" ({cong_ty_tb})" if cong_ty_tb else "")
                    them_thong_bao(nv_moi,
                        f"🔧 Bạn được giao việc: **{cd_v}** • {_suffix}",
                        task_id=task_id, loai="cong_viec_con")
                    them_thong_bao_tat_ca(
                        f"🔧 **{nguoi_giao}** đã giao **{cd_v}** cho **{nv_moi}** • {_suffix}",
                        task_id=task_id, loai="cong_viec_con", tru_nguoi=nv_moi)
                st.session_state[_cv_add_v] += 1
                _save_cv_to_sheet(task_id, _cv_key)
        st.button("＋", key=f"dlg_cv_add_{task_id}",
                  use_container_width=True, on_click=_cb_cv_add)

    st.divider()

    # ── Ảnh đo lường — fallback nếu không có công đoạn tiêu chuẩn ───────────
    _has_stage_cvc = any(
        (cv.get("ten") or "").strip().upper() in _STAGE_DO_LUONG
        for cv in ds_cv_con if isinstance(cv, dict)
    )
    if not _has_stage_cvc:
        st.markdown("**📐 Ảnh Đo Lường**")
        _fragment_upload_do_luong(task_id, _do_key)

    # PDF
    tt_pdf = st.session_state.get(f"tt_select_{task_id}", trang_thai)
    if tt_pdf == "Đã Hoàn Thành - Giao Máy" or "Hoàn Thành" in tt_pdf or tt_pdf == "Chờ Giao":
        st.divider()
        if st.button("📊 Tạo Biên Bản Excel", key=f"xl_{task_id}", use_container_width=True):
            with st.spinner("Đang tạo Excel..."):
                df_moi = lay_danh_sach_cong_viec()
                rows = df_moi[df_moi["ID"].astype(str) == str(task_id)]
                if not rows.empty:
                    du_lieu_excel = tao_excel_nghiem_thu(rows.iloc[0].to_dict())
                    st.download_button(
                        "💾 Tải Xuống Excel",
                        data=du_lieu_excel,
                        file_name=f"BBNT_task_{task_id}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"dl_xl_{task_id}",
                        use_container_width=True,
                    )


# ============================================================
# FRAGMENT: CHECKLIST & CÔNG VIỆC CON (rerun cục bộ, không reload toàn trang)
# ============================================================

_FRAGMENT_CSS = """
<style>
.cl-card {
    display: flex; align-items: center; gap: 10px;
    border-radius: 12px; padding: 11px 14px;
    margin-bottom: 4px; border: 1.5px solid #e5e7eb;
    background: #fff; transition: background 0.2s;
}
.cl-card.cl-done {
    background: #f0fdf4; border-color: #86efac;
}
.cl-num {
    flex-shrink: 0; width: 26px; height: 26px; border-radius: 50%;
    display: inline-flex; align-items: center; justify-content: center;
    font-size: 0.68rem; font-weight: 800; color: #fff;
    background: #7c3aed;
}
.cl-card.cl-done .cl-num { background: #16a34a; }
.cl-txt {
    flex: 1; font-size: 0.9rem; color: #1e293b; word-break: break-word;
}
.cl-card.cl-done .cl-txt {
    color: #1e293b;
}
.cl-badge {
    flex-shrink: 0; font-size: 0.68rem; font-weight: 700;
    color: #16a34a; background: #dcfce7;
    padding: 2px 8px; border-radius: 20px;
}
/* Fix input bị cắt đáy — cascade overflow visible qua mọi wrapper */
[data-testid="stTextInput"],
[data-testid="stTextInput"] > div,
[data-testid="stTextArea"],
[data-testid="stTextArea"] > div,
.element-container {
    overflow: visible !important;
}
/* Ép 2 nút nằm cùng hàng trên mọi màn hình */
div[data-testid="stHorizontalBlock"]:has(> div[data-testid="stColumn"]) {
    flex-wrap: nowrap !important;
    gap: 6px !important;
}
div[data-testid="stHorizontalBlock"] > div[data-testid="stColumn"] {
    min-width: 0 !important;
    flex: 1 1 0 !important;
    width: 50% !important;
    overflow: visible !important;
    padding-bottom: 4px !important;
}
div[data-testid="stHorizontalBlock"] > div[data-testid="stColumn"] > div,
div[data-testid="stHorizontalBlock"] > div[data-testid="stColumn"] > div > div,
div[data-testid="stHorizontalBlock"] > div[data-testid="stColumn"] > div > div > div,
div[data-testid="stHorizontalBlock"] > div[data-testid="stColumn"] > div > div > div > div {
    overflow: visible !important;
}
.cl-btn-row { margin: 0 0 10px 0; gap: 6px; display: flex; }
/* Nút icon nhỏ trong checklist (row có checkbox) */
[data-testid="stHorizontalBlock"]:has([data-testid="stCheckbox"]) .stButton > button {
    background: transparent !important;
    border: 1.5px solid #e5e7eb !important;
    box-shadow: none !important;
    color: #6b7280 !important;
    font-size: 1rem !important;
    padding: 3px 6px !important;
    min-height: 32px !important;
    border-radius: 8px !important;
}
[data-testid="stHorizontalBlock"]:has([data-testid="stCheckbox"]) .stButton > button:hover {
    background: #f3f4f6 !important;
    border-color: #9ca3af !important;
    transform: none !important;
    box-shadow: none !important;
}
/* CVC card */
.cvc-card {
    border: 1.5px solid #e5e7eb; border-radius: 12px;
    padding: 10px 14px 8px; background: #fafafa;
    margin-bottom: 4px;
}
.cvc-card.cvc-done { background: #f0fdf4; border-color: #86efac; }
.cvc-title { font-size: 0.9rem; font-weight: 700; color: #1e293b; }
.cvc-card.cvc-done .cvc-title { color: #1e293b; }
.cvc-meta { font-size: 0.78rem; color: #64748b; margin-top: 3px; }
.cvc-badge {
    display: inline-block; font-size: 0.68rem; font-weight: 700;
    color: #16a34a; background: #dcfce7;
    padding: 2px 8px; border-radius: 20px; margin-top: 4px;
}
</style>
"""

# 12 mục checklist mặc định cho motor repair
_MAC_DINH_CHECKLIST = [
    {"text": "Quấn",                      "done": False},
    {"text": "Thay bạc đạn trước",         "done": False},
    {"text": "Thay bạc đạn sau",           "done": False},
    {"text": "Thay phốt",                  "done": False},
    {"text": "Đóng sơ mi trước",           "done": False},
    {"text": "Đóng sơ mi sau",             "done": False},
    {"text": "Thay tụ",                    "done": False},
    {"text": "Thay cánh quạt",             "done": False},
    {"text": "Thay dây nguồn/dây điện",    "done": False},
    {"text": "Thay nắp chụp",              "done": False},
    {"text": "Đắp cốt",                    "done": False},
    {"text": "Gia công mới cốt",           "done": False},
]

# Màu nền theo trạng thái (cho board)
_STATUS_BG = {
    "Đang Kiểm Tra":              "#00b4d8",
    "Đã Phê Duyệt":               "#f44336",
    "Đã Báo Giá":                 "#f9c74f",
    "Có Đơn":                     "#4361ee",
    "Chờ Giao":                   "#f8961e",
    "Đã Hoàn Thành - Giao Máy":   "#4caf50",
    "Đã Xuất Hóa Đơn":            "#78909c",
    "Bảo Hành - Trả Lại":         "#9b5de5",
    "Chờ Làm":                    "#ef476f",
    "Đang Làm":                   "#ffd166",
    "Hoàn Thành":                 "#06d6a0",
}

# Màu chữ + nền nhạt cho header kanban board (text_color, bg_color)
_STATUS_HEADER_COLOR = {
    "Đang Kiểm Tra":              ("#0077a8", "#e0f7fc"),
    "Đã Phê Duyệt":               ("#b71c1c", "#fce4e4"),
    "Đã Báo Giá":                 ("#7a5c00", "#fff8dc"),
    "Có Đơn":                     ("#2b3eb5", "#e8ecff"),
    "Chờ Giao":                   ("#b35c00", "#fff3e0"),
    "Đã Hoàn Thành - Giao Máy":   ("#1b5e20", "#e6f4ea"),
    "Đã Xuất Hóa Đơn":            ("#37474f", "#eceff1"),
    "Bảo Hành - Trả Lại":         ("#5e1891", "#f3e5ff"),
    "Chờ Làm":                    ("#b71c1c", "#fce4e4"),
    "Đang Làm":                   ("#7a5c00", "#fff8dc"),
    "Hoàn Thành":                 ("#004d40", "#e0f2f1"),
}


@st.fragment
def _fragment_checklist(key_prefix: str, show_done: bool = True, default_items=None):
    cl_key   = f"{key_prefix}_checklist"
    cl_inp_v = f"{key_prefix}_cl_inp_v"
    if cl_key   not in st.session_state:
        st.session_state[cl_key] = [dict(x) for x in default_items] if default_items else []
    if cl_inp_v not in st.session_state: st.session_state[cl_inp_v] = 0

    st.markdown(_FRAGMENT_CSS, unsafe_allow_html=True)
    st.markdown("**☑️ Checklist**")

    items = st.session_state[cl_key]

    _xoa = None

    # CSS dùng :has() với marker bên trong cột — đây là cách duy nhất hoạt động
    # vì st.markdown div KHÔNG wrap st.columns trong DOM thực tế
    mk = f"clm{key_prefix}"  # class marker ngắn
    st.markdown(f"""<style>
    /* Row chứa marker → luôn nằm ngang, không wrap */
    [data-testid="stHorizontalBlock"]:has(.{mk}) {{
        flex-wrap: nowrap !important;
        align-items: center !important;
        gap: 4px !important;
    }}
    /* Cột checkbox: cố định 36px */
    [data-testid="stHorizontalBlock"]:has(.{mk}) > [data-testid="stColumn"]:first-child {{
        flex: 0 0 36px !important; min-width: 36px !important; max-width: 36px !important;
        overflow: visible !important;
    }}
    /* Cột text: chiếm phần còn lại */
    [data-testid="stHorizontalBlock"]:has(.{mk}) > [data-testid="stColumn"]:nth-child(2) {{
        flex: 1 1 0% !important; min-width: 0 !important; overflow: hidden !important;
    }}
    /* Cột xóa: cố định 36px */
    [data-testid="stHorizontalBlock"]:has(.{mk}) > [data-testid="stColumn"]:last-child {{
        flex: 0 0 36px !important; min-width: 36px !important; max-width: 36px !important;
    }}
    /* Text input trông như plain text */
    [data-testid="stHorizontalBlock"]:has(.{mk}) [data-testid="stTextInput"] input {{
        border: 1.5px solid transparent !important;
        background: transparent !important; box-shadow: none !important;
        padding: 3px 6px !important; font-size: 0.93rem !important;
        color: #1e1b4b !important; min-height: 32px !important;
    }}
    [data-testid="stHorizontalBlock"]:has(.{mk}) [data-testid="stTextInput"] input:focus {{
        border-color: #7c3aed !important; background: white !important;
        border-radius: 6px !important;
        box-shadow: 0 0 0 2px rgba(124,58,237,0.12) !important;
    }}
    [data-testid="stHorizontalBlock"]:has(.{mk}) [data-testid="stTextInput"] > div {{
        border: none !important; box-shadow: none !important; background: transparent !important;
    }}
    /* Nút xóa: transparent, đỏ */
    [data-testid="stHorizontalBlock"]:has(.{mk}) [data-testid="stColumn"]:last-child button {{
        background: transparent !important; border: none !important;
        box-shadow: none !important; color: #ef4444 !important;
        font-size: 1.1rem !important; padding: 2px 4px !important;
        min-height: 32px !important; transform: none !important;
    }}
    [data-testid="stHorizontalBlock"]:has(.{mk}) [data-testid="stColumn"]:last-child button:hover {{
        background: #fee2e2 !important; border-radius: 6px !important;
        box-shadow: none !important; transform: none !important;
    }}
    /* Bỏ margin thừa của checkbox */
    [data-testid="stHorizontalBlock"]:has(.{mk}) [data-testid="stCheckbox"] {{
        margin: 0 !important;
    }}
    </style>""", unsafe_allow_html=True)

    def _save_cl(i):
        val = st.session_state.get(f"{key_prefix}_cl_txt_{i}", "").strip()
        if val:
            st.session_state[cl_key][i]["text"] = val

    def _cb_ck_done(i):
        st.session_state[cl_key][i]["done"] = st.session_state.get(f"{key_prefix}_ck_{i}", False)

    def _cb_cl_del(i):
        if i < len(st.session_state[cl_key]):
            st.session_state[cl_key].pop(i)
        for k in list(st.session_state.keys()):
            if k.startswith(f"{key_prefix}_cl_txt_"):
                del st.session_state[k]

    def _cb_cl_add():
        val = st.session_state.get(f"{key_prefix}_cl_inp_{st.session_state[cl_inp_v]}", "").strip()
        if val:
            st.session_state[cl_key].append({"text": val, "done": False})
            st.session_state[cl_inp_v] += 1

    for i, item in enumerate(items):
        txt      = item.get("text", "") or f"Mục {i+1}"
        done_val = bool(item.get("done", False))

        col_ck, col_txt, col_del = st.columns([0.6, 8.5, 0.7], gap="small")
        with col_ck:
            # Marker ở đây để CSS :has() tìm đúng HorizontalBlock này
            st.markdown(f"<span class='{mk}' style='display:none'></span>", unsafe_allow_html=True)
            st.checkbox(
                "", value=done_val,
                key=f"{key_prefix}_ck_{i}",
                label_visibility="collapsed",
                on_change=_cb_ck_done, args=(i,),
            )
        with col_txt:
            st.text_input(
                "", value=txt,
                key=f"{key_prefix}_cl_txt_{i}",
                label_visibility="collapsed",
                on_change=_save_cl, args=(i,),
            )
        with col_del:
            st.button("🗑️", key=f"{key_prefix}_cl_del_{i}", use_container_width=True,
                      on_click=_cb_cl_del, args=(i,))

    st.markdown("<div style='margin-top:8px'></div>", unsafe_allow_html=True)
    col_i, col_b = st.columns([5, 2])
    with col_i:
        st.text_input(
            "", placeholder="Nhập tên checklist...",
            key=f"{key_prefix}_cl_inp_{st.session_state[cl_inp_v]}",
            label_visibility="collapsed",
        )
    with col_b:
        st.button("＋ Thêm", key=f"{key_prefix}_cl_add",
                  use_container_width=True,
                  on_click=_cb_cl_add)



@st.fragment
def _fragment_cong_viec_con(key_prefix: str, ds_nhan_vien: list, show_done: bool = True):
    cv_key      = f"{key_prefix}_cong_viec_con"
    cv_inp_v    = f"{key_prefix}_cv_inp_v"
    cv_seeded   = f"{key_prefix}_cv_seeded"

    # Lần đầu: seed tất cả công đoạn sẵn vào danh sách
    if cv_seeded not in st.session_state:
        ds_cd = lay_ten_cac_cong_doan()
        st.session_state[cv_key]    = [{"ten": cd, "nhan_vien": "", "done": False} for cd in ds_cd]
        st.session_state[cv_inp_v]  = 0
        st.session_state[cv_seeded] = True
    else:
        if cv_key   not in st.session_state: st.session_state[cv_key]   = []
        if cv_inp_v not in st.session_state: st.session_state[cv_inp_v] = 0

    st.markdown(_FRAGMENT_CSS, unsafe_allow_html=True)
    st.markdown("**📋 Công Việc Con**")

    # CSS dùng :has() với marker bên trong cột — không bị override bởi Streamlit mobile wrap
    mk_cv = f"cvcm{key_prefix}"
    nv_opts = ["-- Không chọn --"] + ds_nhan_vien
    st.markdown(f"""<style>
    /* Hàng 1: checkbox + tên + xóa */
    [data-testid="stHorizontalBlock"]:has(.{mk_cv}) {{
        flex-wrap: nowrap !important;
        align-items: center !important;
        gap: 4px !important;
    }}
    [data-testid="stHorizontalBlock"]:has(.{mk_cv}) > [data-testid="stColumn"]:first-child {{
        flex: 0 0 36px !important; min-width: 36px !important; max-width: 36px !important;
    }}
    [data-testid="stHorizontalBlock"]:has(.{mk_cv}) > [data-testid="stColumn"]:nth-child(2) {{
        flex: 1 1 0% !important; min-width: 0 !important;
    }}
    [data-testid="stHorizontalBlock"]:has(.{mk_cv}) > [data-testid="stColumn"]:last-child {{
        flex: 0 0 36px !important; min-width: 36px !important; max-width: 36px !important;
    }}
    /* Tên công việc — nổi bật, dễ nhìn */
    [data-testid="stHorizontalBlock"]:has(.{mk_cv}) [data-testid="stTextInput"] input {{
        border: none !important; border-bottom: 2px solid #c4b5fd !important;
        background: transparent !important; box-shadow: none !important;
        padding: 3px 6px !important; color: #4c1d95 !important;
        min-height: 36px !important; font-weight: 600 !important;
    }}
    [data-testid="stHorizontalBlock"]:has(.{mk_cv}) [data-testid="stTextInput"] input:focus {{
        border-bottom-color: #7c3aed !important; background: #f5f3ff !important;
        border-radius: 6px 6px 0 0 !important;
        box-shadow: none !important;
    }}
    [data-testid="stHorizontalBlock"]:has(.{mk_cv}) [data-testid="stTextInput"] > div {{
        border: none !important; box-shadow: none !important; background: transparent !important;
    }}
    /* Căn checkbox ngang hàng text input */
    [data-testid="stHorizontalBlock"]:has(.{mk_cv}) > [data-testid="stColumn"]:first-child {{
        display: flex !important;
        align-items: center !important;
        padding-top: 0 !important;
        padding-bottom: 0 !important;
    }}
    [data-testid="stHorizontalBlock"]:has(.{mk_cv}) > [data-testid="stColumn"]:first-child > div {{
        display: flex !important;
        align-items: center !important;
        width: 100% !important;
        height: 100% !important;
    }}
    [data-testid="stHorizontalBlock"]:has(.{mk_cv}) [data-testid="stCheckbox"] {{
        margin: 0 !important;
        padding: 0 !important;
    }}
    [data-testid="stHorizontalBlock"]:has(.{mk_cv}) [data-testid="stCheckbox"] label {{
        margin: 0 !important; padding: 0 !important;
    }}
    /* Nút xóa: transparent đỏ */
    [data-testid="stHorizontalBlock"]:has(.{mk_cv}) [data-testid="stColumn"]:last-child button {{
        background: transparent !important; border: none !important;
        box-shadow: none !important; color: #ef4444 !important;
        font-size: 1.1rem !important; padding: 2px 4px !important;
        min-height: 32px !important; transform: none !important;
    }}
    [data-testid="stHorizontalBlock"]:has(.{mk_cv}) [data-testid="stColumn"]:last-child button:hover {{
        background: #fee2e2 !important; border-radius: 6px !important;
        box-shadow: none !important; transform: none !important;
    }}
    /* Hàng 2: selectbox nhân viên full width */
    .cv-nv-frag-{key_prefix} [data-testid="stSelectbox"] > div > div {{
        font-size: 0.9rem !important; border-color: #e0d7ff !important;
        background: #f5f3ff !important; color: #4c1d95 !important;
        min-height: 34px !important;
    }}
    .cv-nv-frag-{key_prefix} [data-testid="stSelectbox"] [class*="singleValue"],
    .cv-nv-frag-{key_prefix} [data-testid="stSelectbox"] [class*="SingleValue"] {{
        direction: ltr !important; font-size: 0.9rem !important;
        overflow: visible !important; white-space: nowrap !important;
        text-overflow: clip !important; max-width: none !important;
    }}
    .cv-nv-frag-{key_prefix} label {{ font-size: 0.78rem !important; color: #6b7280 !important; }}
    </style>""", unsafe_allow_html=True)

    items_cv = st.session_state[cv_key]

    def _save_cv_ten(i):
        val = st.session_state.get(f"{key_prefix}_cv_txt_{i}", "").strip()
        if val:
            st.session_state[cv_key][i]["ten"] = val

    def _save_cv_nv(i):
        val = st.session_state.get(f"{key_prefix}_cv_nv_sel_{i}", "-- Không chọn --")
        st.session_state[cv_key][i]["nhan_vien"] = "" if val == "-- Không chọn --" else val

    def _save_cv_done(i):
        val = st.session_state.get(f"{key_prefix}_cv_ck_{i}", False)
        st.session_state[cv_key][i]["done"] = val
        if val:
            st.session_state[cv_key][i]["ngay_hoan_thanh"] = datetime.now().strftime("%Y-%m-%d")
        else:
            st.session_state[cv_key][i].pop("ngay_hoan_thanh", None)

    def _cb_cv_del(i):
        if i < len(st.session_state[cv_key]):
            st.session_state[cv_key].pop(i)
        for k in list(st.session_state.keys()):
            if k.startswith(f"{key_prefix}_cv_txt_") or k.startswith(f"{key_prefix}_cv_nv_sel_"):
                del st.session_state[k]

    for i, cv in enumerate(items_cv):
        done_val = bool(cv.get("done", False))
        ten   = cv.get("ten", "") or f"Việc {i+1}"
        nguoi = cv.get("nhan_vien", cv.get("nguoi", "")) or ""
        nv_idx = nv_opts.index(nguoi) if nguoi in nv_opts else 0

        # Hàng 1: checkbox + tên công việc + nút xóa
        col_ck, col_txt, col_del = st.columns([0.5, 5.5, 0.5], gap="small")
        with col_ck:
            st.markdown(f"<span class='{mk_cv}' style='display:none'></span>", unsafe_allow_html=True)
            st.checkbox(
                "", value=done_val,
                key=f"{key_prefix}_cv_ck_{i}",
                label_visibility="collapsed",
                on_change=_save_cv_done, args=(i,),
            )
        with col_txt:
            st.text_input(
                "", value=ten,
                key=f"{key_prefix}_cv_txt_{i}",
                label_visibility="collapsed",
                on_change=_save_cv_ten, args=(i,),
            )
        with col_del:
            st.button("🗑️", key=f"{key_prefix}_cv_del_{i}", use_container_width=True,
                      on_click=_cb_cv_del, args=(i,))
        # Hàng 2: selectbox nhân viên full width
        st.markdown(f"<div class='cv-nv-frag-{key_prefix}'>", unsafe_allow_html=True)
        st.selectbox(
            "👤 Nhân viên thực hiện", options=nv_opts,
            index=nv_idx,
            key=f"{key_prefix}_cv_nv_sel_{i}",
            on_change=_save_cv_nv, args=(i,),
        )
        st.markdown("</div>", unsafe_allow_html=True)

        # Hàng 3: upload ảnh cho từng công việc con
        _cv_anh = st.session_state[cv_key][i].get("anh", [])
        _exp_anh_lbl = f"📎 Hình/Video ({len(_cv_anh)})" if _cv_anh else "📎 Thêm hình/video"
        _exp_open_key = f"{key_prefix}_exp_anh_open_{i}"
        if not isinstance(st.session_state.get(_exp_open_key), bool):
            st.session_state[_exp_open_key] = False
        with st.expander(_exp_anh_lbl, expanded=st.session_state[_exp_open_key]):
            if _cv_anh:
                _cols_a = st.columns(min(len(_cv_anh), 3))
                for _ai, _url_a in enumerate(_cv_anh):
                    with _cols_a[_ai % 3]:
                        if "/view" in _url_a or _url_a.endswith((".mp4", ".mov", ".avi")):
                            st.markdown(f"🎬 [Video {_ai+1}]({_url_a})")
                        else:
                            _hien_thi_anh_drive(_url_a, use_container_width=True)
                        def _del_cv_anh_frag(idx=i, url=_url_a):
                            _cvl = st.session_state.get(cv_key, [])
                            if 0 <= idx < len(_cvl):
                                _cvl[idx]["anh"] = [u for u in _cvl[idx].get("anh", []) if u != url]
                        st.button("🗑️ Xoá", key=f"{key_prefix}_del_cvanh_{i}_{_ai}",
                                  use_container_width=True,
                                  on_click=_del_cv_anh_frag)
            _up_ver_key_frag = f"{key_prefix}_up_cv_v_{i}"
            _up_key_frag = f"{key_prefix}_up_cv_{i}_{st.session_state.get(_up_ver_key_frag, 0)}"
            st.file_uploader(
                "Chọn hình hoặc video",
                type=["jpg", "jpeg", "png", "mp4", "mov", "avi"],
                accept_multiple_files=True,
                key=_up_key_frag,
                label_visibility="collapsed",
            )
            def _cb_up_cv_frag(_cvk=cv_key, _idx=i, _upk=_up_key_frag, _eok=_exp_open_key, _vk=_up_ver_key_frag):
                _files = st.session_state.get(_upk) or []
                if not _files:
                    return
                for _f in _files:
                    try:
                        _u = _tai_media_len_drive(_f)
                        st.session_state[_cvk][_idx].setdefault("anh", []).append(_u)
                    except Exception:
                        pass
                st.session_state[_eok] = True
                st.session_state[_vk] = st.session_state.get(_vk, 0) + 1
            st.button("📤 Tải ảnh", key=f"{key_prefix}_btn_up_cv_{i}",
                      use_container_width=True,
                      disabled=not bool(st.session_state.get(_up_key_frag)),
                      on_click=_cb_up_cv_frag)

        # Ảnh Đo Lường — chỉ hiện cho subtask "Nhận máy"
        if ten.strip().upper() == "NHẬN MÁY":
            _nm_do_key   = f"{key_prefix}_do_luong_creation"
            if _nm_do_key not in st.session_state:
                st.session_state[_nm_do_key] = {}
            _nm_nhom_list = _STAGE_DO_LUONG.get("NHẬN MÁY", [])
            _n_do    = sum(len(v) for v in st.session_state[_nm_do_key].values() if isinstance(v, list))
            _total_do = sum(len(lbs) for e in _nm_nhom_list for lbs in [e[1]])
            _do_open_key = f"{key_prefix}_nm_do_open_{i}"
            if _do_open_key not in st.session_state:
                st.session_state[_do_open_key] = False
            _do_chev = "▼" if st.session_state[_do_open_key] else "▶"

            def _toggle_nm_do(_k=_do_open_key):
                st.session_state[_k] = not st.session_state[_k]

            st.button(
                f"{_do_chev}  📐 Ảnh Đo Lường  ({_n_do}/{_total_do} ảnh)",
                key=f"{key_prefix}_nm_do_tog_{i}",
                use_container_width=True,
                on_click=_toggle_nm_do,
            )
            if st.session_state[_do_open_key]:
                _nm_dict = st.session_state[_nm_do_key]
                for _nhom_entry in _nm_nhom_list:
                    _nhom_title, _nhom_lbs = _nhom_entry[0], _nhom_entry[1]
                    st.markdown(
                        f"<div style='background:#dbeafe;border-radius:6px;padding:5px 10px;"
                        f"font-weight:700;font-size:0.82rem;color:#1e3a8a;margin-top:6px;'>"
                        f"{_nhom_title}</div>",
                        unsafe_allow_html=True,
                    )
                    for _lbl_disp, _lbl_key in _nhom_lbs:
                        _urls_lbl = _nm_dict.get(_lbl_key, [])
                        _exp_lbl  = f"📷 Ảnh {_lbl_disp} ✅" if _urls_lbl else f"📷 Ảnh {_lbl_disp}"
                        with st.expander(_exp_lbl, expanded=False):
                            if _urls_lbl:
                                _c_do = st.columns(min(len(_urls_lbl), 3))
                                for _ai, _u in enumerate(_urls_lbl):
                                    with _c_do[_ai % 3]:
                                        _hien_thi_anh_drive(_u, use_container_width=True)
                                        def _del_nm(dk=_nm_do_key, lk=_lbl_key, u=_u):
                                            _d = st.session_state.get(dk, {})
                                            _d[lk] = [x for x in _d.get(lk, []) if x != u]
                                        st.button("🗑️ Xoá", key=f"{key_prefix}_del_nm_{i}_{_lbl_key}_{_ai}",
                                                  use_container_width=True, on_click=_del_nm)
                            _up_k  = f"{key_prefix}_up_nm_{i}_{_lbl_key}"
                            _dn_k  = f"{key_prefix}_dn_nm_{i}_{_lbl_key}"
                            st.file_uploader("Chọn ảnh", type=["jpg", "jpeg", "png"],
                                             key=_up_k, label_visibility="collapsed")
                            def _cb_up_nm(dk=_nm_do_key, lk=_lbl_key, uk=_up_k, dnk=_dn_k):
                                f = st.session_state.get(uk)
                                if not f:
                                    return
                                fid = f"{f.name}_{f.size}"
                                if st.session_state.get(dnk) == fid:
                                    return
                                st.session_state[dnk] = fid
                                url = tai_anh_len_cloudinary(f)
                                st.session_state.setdefault(dk, {}).setdefault(lk, []).append(url)
                            st.button("📤 Tải ảnh", key=f"{key_prefix}_btn_nm_{i}_{_lbl_key}",
                                      use_container_width=True, on_click=_cb_up_nm,
                                      disabled=not bool(st.session_state.get(_up_k)))

    # Thêm thủ công một mục mới
    st.markdown("<div style='margin-top:6px'></div>", unsafe_allow_html=True)
    st.text_input(
        "", placeholder="Tên công việc con...",
        key=f"{key_prefix}_cv_ten_{st.session_state[cv_inp_v]}",
        label_visibility="collapsed",
    )
    st.markdown(f"<div class='cv-nv-frag-{key_prefix}'>", unsafe_allow_html=True)
    st.selectbox(
        "👤 Nhân viên thực hiện", options=nv_opts,
        key=f"{key_prefix}_cv_nv_{st.session_state[cv_inp_v]}",
    )
    st.markdown("</div>", unsafe_allow_html=True)

    def _cb_cv_add():
        ten_val = st.session_state.get(
            f"{key_prefix}_cv_ten_{st.session_state[cv_inp_v]}", "").strip()
        cv_nv   = st.session_state.get(
            f"{key_prefix}_cv_nv_{st.session_state[cv_inp_v]}", "-- Không chọn --")
        if ten_val:
            st.session_state[cv_key].append({
                "ten":       ten_val,
                "nhan_vien": "" if cv_nv == "-- Không chọn --" else cv_nv,
                "done":      False,
            })
            st.session_state[cv_inp_v] += 1

    st.button("＋ Thêm công việc con", key=f"{key_prefix}_cv_add",
              use_container_width=True, on_click=_cb_cv_add)






# ============================================================
# FRAGMENT: UPLOAD ẢNH (dùng trong dialog để tránh đóng dialog khi rerun)
# ============================================================

def _cb_xoa_anh_nt(task_id, anh_key, url_a):
    """Callback xóa ảnh — chạy trước khi re-render, không cần st.rerun()"""
    xoa_url_anh(task_id, url_a)
    st.session_state[anh_key] = [u for u in st.session_state.get(anh_key, []) if u != url_a]


def _cb_upload_anh_nt(task_id, anh_key, up_key):
    """Callback upload ảnh — chạy trước khi re-render, không cần st.rerun()"""
    files = st.session_state.get(up_key) or []
    if not files:
        return
    new_urls = []
    _ms = _lay_ma_so_tu_task_id(task_id)
    for f in files:
        url = tai_anh_len_cloudinary(f, ma_so=_ms)
        cap_nhat_url_anh(task_id, url)
        new_urls.append(url)
    st.session_state[anh_key] = st.session_state.get(anh_key, []) + new_urls
    st.session_state[f"_nt_msg_{task_id}"] = f"✅ Đã upload {len(new_urls)} ảnh!"
    st.session_state[f"up_anh_nt_v_{task_id}"] = st.session_state.get(f"up_anh_nt_v_{task_id}", 0) + 1


def _tao_zip_anh(ds_url: list, ten_file: str = "anh") -> bytes:
    """Fetch tất cả ảnh từ Drive rồi đóng gói thành ZIP, trả bytes."""
    import zipfile, io as _io, re as _re, concurrent.futures as _cf
    def _fetch(idx_url):
        idx, url = idx_url
        m = _re.search(r"[?&]id=([^&]+)", url)
        if m:
            try:
                return idx, _lay_bytes_anh_drive(m.group(1))
            except Exception:
                pass
        return idx, None
    with _cf.ThreadPoolExecutor(max_workers=6) as ex:
        results = list(ex.map(_fetch, enumerate(ds_url)))
    buf = _io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for idx, data in results:
            if data:
                zf.writestr(f"{ten_file}_{idx+1:03d}.jpg", data)
    return buf.getvalue()


def _fragment_upload_anh_nghiem_thu(task_id, anh_key: str):
    """Upload ảnh nghiệm thu — explicit button + spinner để hiện ảnh ngay."""
    up_key = f"up_anh_nt_{task_id}_{st.session_state.get(f'up_anh_nt_v_{task_id}', 0)}"
    ds_anh = st.session_state.get(anh_key, [])
    st.markdown("**📸 Ảnh Nghiệm Thu**")
    if ds_anh:
        st.caption(f"{len(ds_anh)} ảnh đã upload")
        cols_anh = st.columns(min(len(ds_anh), 3))
        for idx_a, url_a in enumerate(ds_anh):
            with cols_anh[idx_a % 3]:
                _hien_thi_anh_drive(url_a, caption=f"Ảnh {idx_a + 1}", use_container_width=True)
                st.button(
                    "🗑️ Xoá", key=f"xoa_anh_{task_id}_{idx_a}",
                    use_container_width=True,
                    on_click=_cb_xoa_anh_nt, args=(task_id, anh_key, url_a),
                )
        # Nút tải ZIP
        if st.button("⬇️ Tải tất cả ảnh (ZIP)", key=f"btn_zip_nt_{task_id}", use_container_width=True):
            with st.spinner(f"Đang tải {len(ds_anh)} ảnh..."):
                _zip_bytes = _tao_zip_anh(ds_anh, ten_file=f"task_{task_id}_anh")
            st.download_button(
                "💾 Tải xuống ZIP",
                data=_zip_bytes,
                file_name=f"task_{task_id}_anh.zip",
                mime="application/zip",
                key=f"dl_zip_nt_{task_id}",
                use_container_width=True,
            )
    else:
        st.caption("Chưa có ảnh nghiệm thu.")

    st.file_uploader(
        "Thêm ảnh (JPG, PNG)",
        type=["jpg", "jpeg", "png"],
        accept_multiple_files=True,
        label_visibility="collapsed",
        key=up_key,
    )
    st.button(
        "📤 Upload ảnh", key=f"btn_up_nt_{task_id}",
        use_container_width=True,
        disabled=not bool(st.session_state.get(up_key)),
        on_click=_cb_upload_anh_nt,
        args=(task_id, anh_key, up_key),
    )


# Cấu trúc: (tiêu đề nhóm, [(display_label, storage_key), ...])
_NHOM_DO = [
    # Trang 3/8 — Resistance
    ("📐 Resistance / Điện trở",
     [("R (U1 – U2)", "R_U1U2"),
      ("R (V1 – V2)", "R_V1V2"),
      ("R (W1 – W2)", "R_W1W2"),
      ("R (PTC)",     "R_PTC"),
      ("R (PT100)",   "R_PT100"),
      ("R (HEATER)",  "R_HEATER")]),

    # Trang 4/8 — Insulation Resistance
    ("🔒 Insulation Resistance / Cách điện",
     [("IR (U – V)",       "IR_UV"),
      ("IR (U – W)",       "IR_UW"),
      ("IR (V – W)",       "IR_VW"),
      ("IR (PTC – E)",     "IR_PTC_E"),
      ("IR (PT100 – E)",   "IR_PT100_E"),
      ("IR (HEATER – E)",  "IR_HEATER_E"),
      ("IR (U – E)",       "IR_U_E"),
      ("IR (V – E)",       "IR_V_E"),
      ("IR (W – E)",       "IR_W_E")]),

    # Rung động — expander đóng mở, chỉ nhập số
    ("📳 Vibration / Rung Động",
     [("Radial ↔ DE/AS",  "vib_rad_h_de"),
      ("Radial ↑ DE/AS",  "vib_rad_v_de"),
      ("Axial (X) DE/AS", "vib_axial_de"),
      ("Radial ↔ NDE/AS", "vib_rad_h_nde"),
      ("Radial ↑ NDE/AS", "vib_rad_v_nde"),
      ("Axial (X) NDE/AS","vib_axial_nde")],
     True),  # True = wrap trong st.expander, mặc định đóng

    # Trang 5/8 — Engine Overview, Terminal Box, Terminal Block
    ("🔧 Engine Overview / Tổng Quan Động Cơ",
     [("Engine / Động cơ",            "eng_overview"),
      ("Nameplate / Bảng thông số",   "eng_nameplate")]),
    ("🔌 Terminal Box / Hộp Điện",
     [("Before / Trước", "tb_before"),
      ("After / Sau",    "tb_after")]),
    ("🔌 Terminal Block / Cầu Đấu Điện",
     [("Before / Trước", "tbl_before"),
      ("After / Sau",    "tbl_after")]),

    # Trang 6/8 — Cover Fan, Rotor Fan, Coil
    ("💨 Cover Fan / Chụp Bảo Vệ Cánh Quạt",
     [("Before / Trước", "cf_before"),
      ("After / Sau",    "cf_after")]),
    ("💨 Rotor Fan / Cánh Quạt",
     [("Before / Trước", "rf_before"),
      ("After / Sau",    "rf_after")]),
    ("🌀 Coil / Cuộn dây",
     [("Before / Trước", "coil_before"),
      ("After / Sau",    "coil_after")]),

    # Trang 7/8 — End Covers, Shaft DE
    ("🔩 End Cover DE / Nắp Đầu Tải",
     [("Before / Trước", "ec_de_before"),
      ("After / Sau",    "ec_de_after")]),
    ("🔩 End Cover NDE / Nắp Đầu Không Tải",
     [("Before / Trước", "ec_nde_before"),
      ("After / Sau",    "ec_nde_after")]),
    ("⚙️ Shaft at DE / Trục Đầu Tải",
     [("Before / Trước", "shaft_de_before"),
      ("After / Sau",    "shaft_de_after")]),

    # Trang 8/8 — Shaft NDE, Bearings
    ("⚙️ Shaft at NDE / Trục Đầu Không Tải",
     [("Before / Trước", "shaft_nde_before"),
      ("After / Sau",    "shaft_nde_after")]),
    ("🔵 DE Bearing / Vòng Bi Đầu Tải",
     [("Before / Trước", "de_brg_before"),
      ("After / Sau",    "de_brg_after")]),
    ("🔵 NDE Bearing / Vòng Bi Đầu Không Tải",
     [("Before / Trước", "nde_brg_before"),
      ("After / Sau",    "nde_brg_after")]),
]

# Keys có field nhập giá trị đo (chỉ Resistance + Insulation Resistance)
_DO_LUONG_VALUE_KEYS = {
    "R_U1U2", "R_V1V2", "R_W1W2", "R_PTC",
    "R_PT100", "R_HEATER",
    "IR_UV", "IR_UW", "IR_VW",
    "IR_PTC_E", "IR_PT100_E", "IR_HEATER_E",
    "IR_U_E", "IR_V_E", "IR_W_E",
    "cur_L1", "cur_L2", "cur_L3",
    "vib_rad_h_de", "vib_rad_v_de", "vib_axial_de",
    "vib_rad_h_nde", "vib_rad_v_nde", "vib_axial_nde",
}
# Keys chỉ chấp nhận số
_DO_LUONG_NUMERIC_KEYS = {
    "R_U1U2", "R_V1V2", "R_W1W2",
    "cur_L1", "cur_L2", "cur_L3",
    "vib_rad_h_de", "vib_rad_v_de", "vib_axial_de",
    "vib_rad_h_nde", "vib_rad_v_nde", "vib_axial_nde",
}
# Keys không cần upload ảnh (chỉ nhập số)
_DO_LUONG_NO_IMG_KEYS = {
    "cur_L1", "cur_L2", "cur_L3",
    "vib_rad_h_de", "vib_rad_v_de", "vib_axial_de",
    "vib_rad_h_nde", "vib_rad_v_nde", "vib_axial_nde",
}
# V1V2, W1W2 không được lệch quá 2% so với U1U2
_DO_LUONG_REF_MAP = {"R_V1V2": "R_U1U2", "R_W1W2": "R_U1U2"}

# Mapping: TÊN CÔNG ĐOẠn (uppercase) → nhóm ảnh đo lường thuộc công đoạn đó
_STAGE_DO_LUONG = {
    "ĐAI ĐẦU": [
        ("📐 Resistance / Điện trở", [
            ("R (U1 – U2)", "R_U1U2"),
            ("R (V1 – V2)", "R_V1V2"),
            ("R (W1 – W2)", "R_W1W2"),
            ("R (PTC)",     "R_PTC"),
        ]),
    ],
    # alias: người dùng có thể nhập 'đấu' (sắc) thay vì 'đầu' (huyền)
    "ĐAI ĐẤU": [
        ("📐 Resistance / Điện trở", [
            ("R (U1 – U2)", "R_U1U2"),
            ("R (V1 – V2)", "R_V1V2"),
            ("R (W1 – W2)", "R_W1W2"),
            ("R (PTC)",     "R_PTC"),
        ]),
    ],
    "NHẬN MÁY": [
        ("🔧 Engine Overview / Tổng Quan Động Cơ", [
            ("Engine / Động cơ",          "eng_overview"),
            ("Nameplate / Bảng thông số", "eng_nameplate"),
        ]),
    ],
    "THÁO MÁY": [
        ("🔌 Terminal Box / Hộp Điện",          [("Before / Trước", "tb_before")]),
        ("🔌 Terminal Block / Cầu Đấu Điện",    [("Before / Trước", "tbl_before")]),
        ("💨 Cover Fan / Chụp Bảo Vệ Cánh Quạt",[("Before / Trước", "cf_before")]),
        ("💨 Rotor Fan / Cánh Quạt",            [("Before / Trước", "rf_before")]),
        ("🔩 End Cover DE / Nắp Đầu Tải",       [("Before / Trước", "ec_de_before")]),
        ("🔩 End Cover NDE / Nắp Đầu Không Tải",[("Before / Trước", "ec_nde_before")]),
        ("⚙️ Shaft at DE / Trục Đầu Tải",       [("Before / Trước", "shaft_de_before")]),
        ("⚙️ Shaft at NDE / Trục Đầu Không Tải",[("Before / Trước", "shaft_nde_before")]),
        ("🔵 DE Bearing / Vòng Bi Đầu Tải",     [("Before / Trước", "de_brg_before")]),
        ("🔵 NDE Bearing / Vòng Bi Đầu Không Tải",[("Before / Trước", "nde_brg_before")]),
    ],
    "ĐỤC MÁY": [
        ("🌀 Coil / Cuộn dây", [("Before / Trước", "coil_before")]),
    ],
    "ĐỤC DÂY": [
        ("📐 Resistance / Điện trở", [
            ("R (U1 – U2)", "R_U1U2"),
            ("R (V1 – V2)", "R_V1V2"),
            ("R (W1 – W2)", "R_W1W2"),
            ("R (PTC)",     "R_PTC"),
        ]),
        ("🌀 Coil / Cuộn dây", [("Before / Trước", "coil_before")]),
    ],
    "VÔ DÂY": [
        ("🌀 Coil / Cuộn dây", [("After / Sau", "coil_after")]),
    ],
    "LẮP MÁY": [
        ("📐 Resistance / Điện trở", [
            ("R (PT100)",  "R_PT100"),
            ("R (HEATER)", "R_HEATER"),
        ]),
        ("🔒 Insulation Resistance / Cách điện", [
            ("IR (U – V)",       "IR_UV"),
            ("IR (U – W)",       "IR_UW"),
            ("IR (V – W)",       "IR_VW"),
            ("IR (PTC – E)",     "IR_PTC_E"),
            ("IR (PT100 – E)",   "IR_PT100_E"),
            ("IR (HEATER – E)",  "IR_HEATER_E"),
            ("IR (U – E)",       "IR_U_E"),
            ("IR (V – E)",       "IR_V_E"),
            ("IR (W – E)",       "IR_W_E"),
        ]),
        ("⚡ Current / Dòng Không Tải (A)",
         [("L1", "cur_L1"),
          ("L2", "cur_L2"),
          ("L3", "cur_L3")]),
        ("📳 Vibration / Rung Động",
         [("Radial ↔ DE/AS",  "vib_rad_h_de"),
          ("Radial ↑ DE/AS",  "vib_rad_v_de"),
          ("Axial (X) DE/AS", "vib_axial_de"),
          ("Radial ↔ NDE/AS", "vib_rad_h_nde"),
          ("Radial ↑ NDE/AS", "vib_rad_v_nde"),
          ("Axial (X) NDE/AS","vib_axial_nde")],
         True),
        ("🔌 Terminal Box / Hộp Điện",          [("After / Sau", "tb_after")]),
        ("🔌 Terminal Block / Cầu Đấu Điện",    [("After / Sau", "tbl_after")]),
        ("💨 Cover Fan / Chụp Bảo Vệ Cánh Quạt",[("After / Sau", "cf_after")]),
        ("💨 Rotor Fan / Cánh Quạt",            [("After / Sau", "rf_after")]),
        ("🔩 End Cover DE / Nắp Đầu Tải",       [("After / Sau", "ec_de_after")]),
        ("🔩 End Cover NDE / Nắp Đầu Không Tải",[("After / Sau", "ec_nde_after")]),
        ("⚙️ Shaft at DE / Trục Đầu Tải",       [("After / Sau", "shaft_de_after")]),
        ("⚙️ Shaft at NDE / Trục Đầu Không Tải",[("After / Sau", "shaft_nde_after")]),
        ("🔵 DE Bearing / Vòng Bi Đầu Tải",     [("After / Sau", "de_brg_after")]),
        ("🔵 NDE Bearing / Vòng Bi Đầu Không Tải",[("After / Sau", "nde_brg_after")]),
    ],
}


def _cb_xoa_do(task_id, do_key, label, url_d):
    """Callback xóa ảnh đo lường — không cần st.rerun()"""
    cur = st.session_state[do_key].get(label, [])
    if url_d in cur:
        cur.remove(url_d)
    st.session_state[do_key][label] = cur
    cap_nhat_anh_do_luong(task_id, st.session_state[do_key])


def _cb_save_val(task_id, do_key, val_key, inp_key):
    """Callback lưu giá trị thông số đo lường — tự động lưu GSheets khi rời field."""
    import re as _re
    val = (st.session_state.get(inp_key) or "").strip()
    _lbl_key = val_key.replace("_val", "")
    if _lbl_key in _DO_LUONG_NUMERIC_KEYS:
        val = _re.sub(r"[^\d.]", "", val)
        parts = val.split(".")
        if len(parts) > 2:
            val = parts[0] + "." + "".join(parts[1:])
    st.session_state[do_key][val_key] = val
    cap_nhat_anh_do_luong(task_id, st.session_state[do_key])


def _cb_upload_do(task_id, do_key, label, up_key, done_key=None):
    """Kept for compatibility — logic đã chuyển vào inline button."""
    pass


@st.fragment
def _render_do_luong_inline(task_id, do_key, nhom_list, cvi=0):
    """Render ảnh đo lường inline bên trong thẻ công việc con."""
    if do_key not in st.session_state:
        st.session_state[do_key] = {}
    for nhom_entry in nhom_list:
        use_expander = len(nhom_entry) >= 3 and nhom_entry[2]
        nhom_title, labels = nhom_entry[0], nhom_entry[1]

        if use_expander:
            _nhom_ctx = st.expander(nhom_title, expanded=False)
        else:
            st.markdown(
                f"<div style='background:#dbeafe;border-radius:6px;padding:5px 10px;"
                f"font-weight:700;font-size:0.82rem;color:#1e3a8a;margin-top:6px;'>"
                f"{nhom_title}</div>",
                unsafe_allow_html=True,
            )
            _nhom_ctx = st.container()

        with _nhom_ctx:
            for lbl_display, lbl_key in labels:
                if lbl_key in _DO_LUONG_VALUE_KEYS:
                    _val_key = f"{lbl_key}_val"
                    _cur_val = st.session_state.get(do_key, {}).get(_val_key, "")
                    _inp_key = f"inp_val_{task_id}_{cvi}_{lbl_key}"
                    _is_num = lbl_key in _DO_LUONG_NUMERIC_KEYS
                    st.text_input(
                        lbl_display,
                        value=_cur_val,
                        key=_inp_key,
                        placeholder="Nhập số (VD: 5.8)..." if _is_num else "Nhập giá trị đo...",
                        on_change=_cb_save_val,
                        args=(task_id, do_key, _val_key, _inp_key),
                    )
                    _ref_key = _DO_LUONG_REF_MAP.get(lbl_key)
                    if _ref_key and _cur_val:
                        _ref_val_str = st.session_state.get(do_key, {}).get(f"{_ref_key}_val", "")
                        try:
                            _ref_f = float(_ref_val_str)
                            _this_f = float(_cur_val)
                            if _ref_f > 0:
                                _dev_pct = abs(_this_f - _ref_f) / _ref_f * 100
                                if _dev_pct > 2:
                                    st.warning(f"⚠️ Lệch {_dev_pct:.1f}% so với R(U1–U2) = {_ref_val_str} (cho phép ≤2%)")
                        except (ValueError, TypeError):
                            pass
                    elif _is_num and _cur_val:
                        try:
                            float(_cur_val)
                        except ValueError:
                            st.error("❌ Chỉ được nhập số")
                if lbl_key not in _DO_LUONG_NO_IMG_KEYS:
                    urls_label = st.session_state.get(do_key, {}).get(lbl_key, [])
                    exp_lbl = f"📷 Ảnh {lbl_display} ✅" if urls_label else f"📷 Ảnh {lbl_display}"
                    with st.expander(exp_lbl, expanded=False):
                        if urls_label:
                            url_d = urls_label[0]
                            col_img, col_del = st.columns([3, 1], gap="small")
                            with col_img:
                                _hien_thi_anh_drive(url_d, use_container_width=True)
                            with col_del:
                                st.button(
                                    "🗑️", key=f"xoa_do_{task_id}_{cvi}_{lbl_key}_0",
                                    use_container_width=True,
                                    on_click=_cb_xoa_do,
                                    args=(task_id, do_key, lbl_key, url_d),
                                )
                        else:
                            up_key = f"up_do_{task_id}_{cvi}_{lbl_key}"
                            st.file_uploader(
                                f"Ảnh {lbl_display}",
                                type=["jpg", "jpeg", "png"],
                                key=up_key,
                                label_visibility="collapsed",
                                accept_multiple_files=False,
                            )
                            if st.button("📤 Tải ảnh", key=f"btn_do_{task_id}_{cvi}_{lbl_key}",
                                         use_container_width=True,
                                         disabled=not bool(st.session_state.get(up_key))):
                                f_do = st.session_state.get(up_key)
                                if f_do:
                                    try:
                                        with st.spinner("Đang tải ảnh lên..."):
                                            # Compress mạnh hơn (800px/70) → file nhỏ hơn → upload nhanh hơn
                                            url_new = tai_anh_len_cloudinary(f_do, ma_so=_lay_ma_so_tu_task_id(task_id))
                                        st.session_state[do_key].setdefault(lbl_key, []).append(url_new)
                                        cap_nhat_anh_do_luong(task_id, st.session_state[do_key])
                                        st.rerun(scope="fragment")
                                    except Exception:
                                        st.error("❌ File ảnh không hợp lệ. Vui lòng chọn lại ảnh và thử lại.")

    st.components.v1.html(
        """<script>
        (function() {
            var NUM_PHS = ['Nhập số (VD: 5.8)...', 'Nhập giá trị đo...'];
            function fixEl(el) {
                if (NUM_PHS.indexOf(el.placeholder) !== -1) {
                    el.setAttribute('inputmode', 'decimal');
                    if (!el._commaFixed) {
                        el._commaFixed = true;
                        el.addEventListener('input', function() {
                            var pos = el.selectionStart;
                            var nv = el.value.replace(/,/g, '.');
                            if (nv !== el.value) {
                                el.value = nv;
                                el.selectionStart = pos;
                                el.selectionEnd = pos;
                                el.dispatchEvent(new Event('input', {bubbles: true}));
                            }
                        });
                    }
                } else {
                    el.removeAttribute('inputmode');
                }
            }
            function applyAll() {
                window.parent.document.querySelectorAll('input').forEach(fixEl);
            }
            applyAll();
            setTimeout(applyAll, 300);
            var obs = new MutationObserver(applyAll);
            obs.observe(window.parent.document.body, {childList: true, subtree: true});
        })();
        </script>""",
        height=0,
    )


@st.fragment
def _fragment_upload_do_luong(task_id, do_key: str):
    """Upload ảnh đo lường — @st.fragment để rerun cục bộ, giữ nguyên scroll của outer fragment."""
    if do_key not in st.session_state:
        st.session_state[do_key] = {}
    for nhom_entry in _NHOM_DO:
        use_exp = len(nhom_entry) >= 3 and nhom_entry[2]
        nhom_title, labels = nhom_entry[0], nhom_entry[1]
        if use_exp:
            _ctx = st.expander(nhom_title, expanded=False)
        else:
            st.markdown(
                f"<div style='background:#dbeafe;border-radius:8px 8px 0 0;padding:8px 14px;"
                f"font-weight:700;font-size:0.88rem;color:#1e3a8a;margin-top:10px;'>"
                f"{nhom_title}</div>",
                unsafe_allow_html=True,
            )
            _ctx = None
        with (_ctx if _ctx else st.container()):
          for item in labels:
            # item là (display_label, storage_key) hoặc chuỗi cũ
            if isinstance(item, tuple):
                lbl_display, lbl_key = item
            else:
                lbl_display = lbl_key = item
            # ── Giá trị thông số đo (chỉ Resistance + IR) ──
            if lbl_key in _DO_LUONG_VALUE_KEYS:
                _val_key = f"{lbl_key}_val"
                _cur_val = st.session_state.get(do_key, {}).get(_val_key, "")
                _inp_key = f"inp_val_{task_id}_{lbl_key}"
                _is_num = lbl_key in _DO_LUONG_NUMERIC_KEYS
                st.text_input(
                    lbl_display,
                    value=_cur_val,
                    key=_inp_key,
                    placeholder="Nhập số (VD: 5.8)..." if _is_num else "Nhập giá trị đo...",
                    on_change=_cb_save_val,
                    args=(task_id, do_key, _val_key, _inp_key),
                )
                # Kiểm tra lệch 2% so với U1U2
                _ref_key = _DO_LUONG_REF_MAP.get(lbl_key)
                if _ref_key and _cur_val:
                    _ref_val_str = st.session_state.get(do_key, {}).get(f"{_ref_key}_val", "")
                    try:
                        _ref_f = float(_ref_val_str)
                        _this_f = float(_cur_val)
                        if _ref_f > 0:
                            _dev_pct = abs(_this_f - _ref_f) / _ref_f * 100
                            if _dev_pct > 2:
                                st.warning(f"⚠️ Lệch {_dev_pct:.1f}% so với R(U1–U2) = {_ref_val_str} (cho phép ≤2%)")
                    except (ValueError, TypeError):
                        pass
                elif _is_num and _cur_val:
                    try:
                        float(_cur_val)
                    except ValueError:
                        st.error("❌ Chỉ được nhập số")
            # ── Ảnh minh chứng (bỏ qua key không cần ảnh) ──────
            if lbl_key in _DO_LUONG_NO_IMG_KEYS:
                continue
            urls_label = st.session_state.get(do_key, {}).get(lbl_key, [])
            _exp_lbl = f"📷 Ảnh {lbl_display} ✅" if urls_label else f"📷 Ảnh {lbl_display}"
            with st.expander(_exp_lbl, expanded=False):
                if urls_label:
                    # Đã có ảnh — hiển thị + nút xoá
                    url_d = urls_label[0]
                    col_img, col_del = st.columns([3, 1], gap="small")
                    with col_img:
                        _hien_thi_anh_drive(url_d, use_container_width=True)
                    with col_del:
                        st.button(
                            "🗑️ Xoá", key=f"xoa_do_{task_id}_{lbl_key}_0",
                            use_container_width=True,
                            on_click=_cb_xoa_do,
                            args=(task_id, do_key, lbl_key, url_d),
                        )
                else:
                    # Chưa có ảnh — hiển thị uploader
                    up_key = f"up_do_{task_id}_{lbl_key}"
                    st.file_uploader(
                        f"Ảnh {lbl_display}",
                        type=["jpg", "jpeg", "png"],
                        key=up_key,
                        label_visibility="collapsed",
                        accept_multiple_files=False,
                    )
                    if st.button("📤 Tải ảnh", key=f"btn_do_{task_id}_{lbl_key}",
                                 use_container_width=True,
                                 disabled=not bool(st.session_state.get(up_key))):
                        f_do = st.session_state.get(up_key)
                        if f_do:
                            try:
                                with st.spinner("Đang tải ảnh lên..."):
                                    url_new = tai_anh_len_cloudinary(f_do, ma_so=_lay_ma_so_tu_task_id(task_id))
                                st.session_state[do_key].setdefault(lbl_key, []).append(url_new)
                                cap_nhat_anh_do_luong(task_id, st.session_state[do_key])
                                st.rerun(scope="fragment")
                            except Exception:
                                st.error("❌ File ảnh không hợp lệ. Vui lòng chọn lại ảnh và thử lại.")


# ─── helper: lưu công việc con về sheet ───────────────────────────────────────
def _parse_date_display(val: str):
    """Chuyển chuỗi DD/MM/YYYY sang date object, trả None nếu lỗi."""
    try:
        return datetime.strptime(str(val)[:10], "%d/%m/%Y").date()
    except Exception:
        return None


def _save_cv_to_sheet(task_id, cv_key):
    data = list(st.session_state.get(cv_key, []))
    task_id = int(task_id)
    def _merge_with_existing(new_data, sh, row):
        """Merge new_data với data hiện có trên sheet.
        - Nếu new_data rỗng hoàn toàn → dùng data cũ (tránh ghi đè mất dữ liệu).
        - Nếu có data → preserve nhan_vien/anh từ sheet cho công đoạn còn tồn tại.
        """
        try:
            raw = sh.cell(row, 14).value or "[]"
            existing = json.loads(raw) if raw.strip() else []
        except Exception:
            existing = []
        if not existing:
            return new_data
        # new_data hoàn toàn rỗng → dùng existing (session chưa load được)
        if not new_data:
            return existing
        ex_by_name = {i.get("ten", ""): i for i in existing if isinstance(i, dict)}
        new_tens = {i.get("ten", "") for i in new_data if isinstance(i, dict)}
        merged = []
        for item in new_data:
            ten = item.get("ten", "")
            ex  = ex_by_name.get(ten, {})
            mi  = dict(item)
            # Giữ nhan_vien từ sheet nếu session không có
            if not mi.get("nhan_vien", "").strip() and ex.get("nhan_vien", "").strip():
                mi["nhan_vien"] = ex["nhan_vien"]
            # Giữ anh từ sheet nếu session không có
            if not mi.get("anh") and ex.get("anh"):
                mi["anh"] = ex["anh"]
            merged.append(mi)
        # Giữ lại các item trên sheet mà session không có (tránh mất khi session load thiếu)
        for ex_item in existing:
            if isinstance(ex_item, dict) and ex_item.get("ten", "") not in new_tens:
                merged.append(ex_item)
        return merged

    def _write():
        try:
            sh = lay_sheet()
            row = _ROW_CACHE.get(task_id)
            if not row:
                o = sh.find(str(task_id), in_column=1)
                if not o: return
                row = o.row
                _ROW_CACHE[task_id] = row
            final = _merge_with_existing(data, sh, row)
            sh.update_cell(row, 14, json.dumps(final, ensure_ascii=False))
            lay_danh_sach_cong_viec.clear()
        except Exception:
            try:
                sh = _lay_sheet_fresh()
                _ROW_CACHE.pop(task_id, None)
                o = sh.find(str(task_id), in_column=1)
                if o:
                    _ROW_CACHE[task_id] = o.row
                    final = _merge_with_existing(data, sh, o.row)
                    sh.update_cell(o.row, 14, json.dumps(final, ensure_ascii=False))
                    lay_danh_sach_cong_viec.clear()
            except Exception:
                pass
    threading.Thread(target=_write, daemon=False).start()


# ============================================================
# DIALOG: TẠO TASK THÀNH CÔNG
# ============================================================
@st.dialog("✅ Tạo Task Thành Công!")
def _dialog_tao_task_thanh_cong(message: str, goto_key: str):
    st.success(message)
    st.balloons()
    if st.button("📋 OK — Xem Bảng Quản Lý", use_container_width=True, type="primary"):
        st.session_state[goto_key] = True
        st.rerun()


@st.dialog("✅ Tạo Task Thành Công!")
def _dialog_nv_tao_task_xong(message: str):
    st.success(message)
    st.balloons()
    if st.button("✅ OK", use_container_width=True, type="primary"):
        st.session_state["_nv_scroll_top"] = True
        st.rerun()


# ============================================================
# KANBAN BOARD HELPER
# ============================================================
@st.dialog("📋 Chi tiết & Chỉnh sửa công việc", width="large")
def _task_dialog(hang_dict, ds_tt):
    """Dialog to hiển thị chi tiết và chỉnh sửa task."""
    # Đánh dấu dialog đang mở (để phát hiện đóng bằng X vs reconnect)
    st.session_state["_dlg_active_flag"] = True
    _tid_qp = str(hang_dict.get("ID", ""))
    if _tid_qp:
        st.query_params["dlg"] = _tid_qp
    try:
        tid = int(float(hang_dict.get("ID", 0)))
    except (ValueError, TypeError):
        tid = 0
    ten  = hang_dict.get("Tên Công Việc", "")
    cty  = hang_dict.get("Công Ty", "")
    tt   = hang_dict.get("Trạng Thái", "")
    # Badge + selectbox trạng thái — fragment riêng để cập nhật ngay khi đổi
    _fragment_trang_thai_dialog(hang_dict, ds_tt)
    # ── Tên công việc có thể chỉnh sửa ──────────────────────────
    def _cb_luu_ten():
        val = st.session_state.get(f"dlg_ten_{tid}", "").strip()
        if val:
            threading.Thread(
                target=cap_nhat_nhieu_truong_task,
                args=(int(tid), {"Tên Công Việc": val}),
                daemon=False,
            ).start()
    _ten_new = st.text_input(
        "📌 Tên công việc",
        value=ten,
        key=f"dlg_ten_{tid}",
        label_visibility="collapsed",
        on_change=_cb_luu_ten,
    )
    if _ten_new and _ten_new != ten:
        if st.button("💾 Lưu tên công việc", key=f"dlg_btn_luu_ten_{tid}", use_container_width=True):
            with st.spinner("Đang lưu..."):
                cap_nhat_nhieu_truong_task(int(tid), {"Tên Công Việc": _ten_new.strip()})
            st.session_state["_board_dirty"] = True
            st.toast("✅ Đã lưu tên!")
    # ── Công ty có thể chỉnh sửa ─────────────────────────────────
    def _cb_luu_cty():
        val = st.session_state.get(f"dlg_cty_{tid}", "")
        threading.Thread(
            target=cap_nhat_nhieu_truong_task,
            args=(int(tid), {"Công Ty": val}),
            daemon=False,
        ).start()
    _ds_cty = [""] + lay_ten_cac_cong_ty()
    _cty_idx = _ds_cty.index(cty) if cty in _ds_cty else 0
    st.selectbox(
        "🏢 Công ty",
        _ds_cty,
        index=_cty_idx,
        key=f"dlg_cty_{tid}",
        label_visibility="collapsed",
        on_change=_cb_luu_cty,
    )
    # ── Mô Tả (ngay bên dưới Công ty) ────────────────────────────
    _mo_ta_cur = hang_dict.get("Mô Tả", "") or ""
    _mo_ta_new = st.text_area(
        "📝 Mô tả công việc",
        value=_mo_ta_cur,
        placeholder="Nhập mô tả công việc...",
        key=f"dlg_mo_ta_{tid}",
        height=100,
    )
    if _mo_ta_new != _mo_ta_cur:
        if st.button("💾 Lưu mô tả", key=f"dlg_save_mo_ta_{tid}", use_container_width=True):
            with st.spinner("Đang lưu..."):
                cap_nhat_nhieu_truong_task(int(tid), {"Mô Tả": _mo_ta_new})
            st.success("✅ Đã lưu!")
    st.divider()
    _fragment_chi_tiet_task(hang_dict, ds_tt, show_status=False)


@st.fragment
def _render_kanban_board(df, ds_tt, board_key="kb", force_open=False, show_delete=True):
    """Kanban board: header màu + toggle thu/mở, cards 4 cột/hàng.
    Bấm 📂 → @st.dialog chỉnh sửa.
    """
    import html as _hl
    import re as _re
    _EXCLUDE   = {"Đã Xuất Hóa Đơn", "Bảo Hành - Trả Lại"}
    _COMPLETED = {"Đã Hoàn Thành - Giao Máy", "Hoàn Thành"}
    ds_show = list(dict.fromkeys(t for t in ds_tt if t not in _EXCLUDE))
    today = datetime.now().date()
    COLS  = 4  # card per row

    # Loại bỏ task trùng ID (giữ dòng đầu tiên) để tránh duplicate key
    if not df.empty and "ID" in df.columns:
        df = df.drop_duplicates(subset=["ID"], keep="first")

    # Áp dụng override trạng thái từ session_state (để board cập nhật ngay sau khi đổi)
    _overrides = st.session_state.get("_tt_overrides", {})
    if _overrides and not df.empty:
        df = df.copy()
        for _oid, _ott in _overrides.items():
            df.loc[df["ID"].astype(str) == str(_oid), "Trạng Thái"] = _ott

    # ── Inject CSS màu cho từng nút toggle ──────────────────
    # Dùng selector theo element-container (wrapper thật của Streamlit)
    css_rules = []
    for tt2 in ds_show:
        txt2, bg2 = _STATUS_HEADER_COLOR.get(tt2, ("#37474f", "#eceff1"))
        safe2     = _re.sub(r"[^a-zA-Z0-9]", "_", tt2)
        mk        = f"kbtog_{board_key}_{safe2}"
        # .element-container:has(.mk) + .element-container button
        sel = (
            f".element-container:has(.{mk}) + .element-container button,"
            f".element-container:has(.{mk}) + .element-container button:hover"
        )
        css_rules.append(
            f"{sel} {{ background:{bg2} !important; color:{txt2} !important;"
            f" border:2px solid {txt2}55 !important; font-weight:800 !important;"
            f" font-size:0.95rem !important; letter-spacing:0.5px !important;"
            f" box-shadow: none !important;"
            f" border-radius:8px !important;"
            f" text-align:left !important; padding:10px 18px !important; }}"
        )
    st.markdown(f"<style>{'  '.join(css_rules)}</style>", unsafe_allow_html=True)

    for tt in ds_show:
        nhom   = df[df["Trạng Thái"] == tt] if not df.empty else df.iloc[0:0]
        so     = len(nhom)
        mau_bg = _STATUS_BG.get(tt, "#607d8b")
        mau_fg = "#1a1a1a" if mau_bg in ("#f9c74f", "#ffd166") else "#ffffff"
        # màu header nhạt có chữ đậm màu
        hdr_txt, hdr_bg = _STATUS_HEADER_COLOR.get(tt, ("#37474f", "#eceff1"))

        # ── Toggle state ──────────────────────────────────────
        _sk = f"{board_key}_open_{tt}"
        if _sk not in st.session_state:
            st.session_state[_sk] = False   # mặc định đóng
        if force_open and not nhom.empty:
            st.session_state[_sk] = True

        is_open = st.session_state[_sk]
        chevron = "▼" if is_open else "▶"
        safe_tt = _re.sub(r"[^a-zA-Z0-9]", "_", tt)
        mk      = f"kbtog_{board_key}_{safe_tt}"

        # ── Header: marker div + full-width button (CSS tô màu) ──
        st.markdown(f"<div class='{mk}' style='display:none'></div>",
                    unsafe_allow_html=True)
        btn_label = f"{chevron}  {tt}  ({so})"
        def _toggle_group(_sk=_sk):
            st.session_state[_sk] = not st.session_state[_sk]
        st.button(btn_label, key=f"{board_key}_tog_{tt}",
                  use_container_width=True, on_click=_toggle_group)

        if not is_open:
            continue

        if nhom.empty:
            st.caption("— Không có công việc nào —")
            continue

        rows_data = [nhom.iloc[i:i+COLS] for i in range(0, len(nhom), COLS)]
        for row_df in rows_data:
            st_cols = st.columns(COLS, gap="small")
            for ci, (_, h) in enumerate(row_df.iterrows()):
                task_id   = h["ID"]
                ten_cv    = str(h.get("Tên Công Việc", "") or "")
                cong_ty   = str(h.get("Công Ty", "") or "")
                nhan_vien = str(h.get("Nhân Viên", "") or "")
                ngay_tao  = str(h.get("Ngày Tạo", ""))[:10]
                mo_ta     = str(h.get("Mô Tả", "") or "").strip()

                is_over = False
                try:
                    dl = str(h.get("Ngày Kết Thúc", "") or "")[:10]
                    if dl and tt not in _COMPLETED:
                        import datetime as _dt
                        is_over = _dt.date.fromisoformat(dl) < today
                except Exception:
                    pass

                anh_lst = doc_danh_sach_anh(str(h.get("Link Ảnh", "") or ""))
                try:
                    cl  = json.loads(str(h.get("Checklist", "") or "[]"))
                    cld = sum(1 for x in cl if isinstance(x, dict) and x.get("done"))
                    clt = len(cl)
                except Exception:
                    cld = clt = 0
                try:
                    cvt = len(json.loads(str(h.get("Công Việc Con", "") or "[]")))
                except Exception:
                    cvt = 0

                with st_cols[ci]:
                    with st.container(border=True):
                        if is_over:
                            st.error("⚠️ Quá hạn", icon="🔴")
                        if anh_lst:
                            _hien_thi_anh_drive(anh_lst[0], use_container_width=True)
                        # Tên + nút xóa cùng hàng
                        if show_delete:
                            _col_ten, _col_del = st.columns([5, 1], gap="small")
                        else:
                            _col_ten = st.container()
                            _col_del = None
                        with _col_ten:
                            short = ten_cv[:40] + ("…" if len(ten_cv) > 40 else "")
                            st.markdown(f"**{short}**")
                        if _col_del:
                            with _col_del:
                                if st.button("✕", key=f"kdel_{board_key}_{task_id}",
                                             help="Xóa công việc này",
                                             use_container_width=True):
                                    _dialog_xac_nhan_xoa(task_id, ten_cv)
                        if mo_ta:
                            short_mo_ta = mo_ta[:120] + ("…" if len(mo_ta) > 120 else "")
                            st.caption(f"📝 {short_mo_ta}")
                        meta = []
                        if cong_ty:   meta.append(f"🏢 {cong_ty}")
                        if nhan_vien: meta.append(f"👤 {nhan_vien}")
                        if ngay_tao:  meta.append(f"📅 {ngay_tao}")
                        if anh_lst:   meta.append(f"📷 {len(anh_lst)}")
                        if clt:       meta.append(f"☑️ {cld}/{clt}")
                        if cvt:       meta.append(f"🔹 {cvt}")
                        if meta:
                            st.caption(" · ".join(meta))
                        if st.button("📂 Xem & chỉnh sửa", key=f"kopen_{board_key}_{task_id}", use_container_width=True):
                            st.session_state.pop("_pending_dlg", None)
                            _task_dialog(h.to_dict(), ds_tt)


def _render_detail_expanders(df, ds_tt):
    pass  # không còn dùng — chi tiết đã chuyển vào _task_dialog


# ============================================================
# FRAGMENT: Công Việc Con AgGrid
# ============================================================
@st.fragment
def _fragment_cvc_content(df_all_cvc, aggrid_css):
    import io

    if df_all_cvc.empty:
        st.info("ℹ️ Chưa có công việc nào.")
        return

    today = datetime.today().date()
    rows_cvc = []
    stt = 0
    for _, task_row in df_all_cvc.iterrows():
        raw = task_row.get("Công Việc Con", "") or "[]"
        try:
            ds_cv = json.loads(raw) if raw else []
        except Exception:
            ds_cv = []
        if not isinstance(ds_cv, list) or not ds_cv:
            continue

        han_str  = str(task_row.get("Hạn Hoàn Thành", "") or "")
        han_date = None
        try:
            if han_str:
                han_date = datetime.strptime(han_str[:10], "%Y-%m-%d").date()
        except Exception:
            pass

        for cv in ds_cv:
            if not isinstance(cv, dict):
                continue
            done = bool(cv.get("done", False))
            if not done:
                continue
            stt += 1
            ten_cv   = cv.get("ten", cv.get("Tên", ""))
            nv       = cv.get("nhan_vien", cv.get("Nhân Viên", cv.get("nguoi", ""))) or ""
            ngay_ht  = cv.get("ngay_hoan_thanh", "") or str(task_row.get("Ngày Tạo", ""))[:10]

            if done and ngay_ht:
                try:
                    ht_date = datetime.strptime(ngay_ht[:10], "%Y-%m-%d").date()
                    if han_date:
                        if ht_date < han_date:
                            tinh_trang_td = "Trước hạn"
                        elif ht_date == han_date:
                            tinh_trang_td = "Đúng hạn"
                        else:
                            tinh_trang_td = "Quá hạn"
                    else:
                        tinh_trang_td = "Hoàn thành"
                except Exception:
                    tinh_trang_td = "Hoàn thành"
            elif done:
                tinh_trang_td = "Hoàn thành"
            elif han_date and today > han_date:
                tinh_trang_td = "Quá hạn"
            elif han_date and today == han_date:
                tinh_trang_td = "Đúng hạn"
            else:
                tinh_trang_td = "Chưa xong"

            ngay_ht_date = None
            ngay_ht_hien = ""
            if done and ngay_ht:
                try:
                    ngay_ht_date = datetime.strptime(ngay_ht[:10], "%Y-%m-%d").date()
                    ngay_ht_hien = ngay_ht_date.strftime("%d/%m/%Y")
                except Exception:
                    ngay_ht_hien = "✅ (chưa rõ ngày)"
            elif done:
                ngay_ht_hien = "✅ (chưa rõ ngày)"

            rows_cvc.append({
                "STT":               stt,
                "Tên Công Ty":       task_row.get("Công Ty", ""),
                "Tên Công Việc":     task_row.get("Tên Công Việc", ""),
                "Công Suất":         task_row.get("Công Suất", ""),
                "Mã Số":             task_row.get("Mã Số", ""),
                "Công Việc Con":     ten_cv,
                "Nhân Viên":         nv,
                "Ngày Hoàn Thành":   ngay_ht_hien,
                "_ngay_ht_raw":      ngay_ht_date,
                "Trạng Thái":        task_row.get("Trạng Thái", ""),
                "Loại Máy":          task_row.get("Loại Máy", ""),
                "Tình Trạng Máy":    str(task_row.get("Tình Trạng", "") or ""),
                "Tình Trạng":        tinh_trang_td,
                "_task_id":          task_row.get("ID", ""),
                "_task_dict":        task_row.to_dict(),
            })

    if not rows_cvc:
        st.info("ℹ️ Chưa có công việc con nào.")
        return

    df_cvc = pd.DataFrame(rows_cvc)

    # ── Bộ lọc theo từng cột ──────────────────────────────────
    _fc1, _fc2, _fc3, _fc4 = st.columns(4)
    _f_ms_cvc  = _fc1.text_input("🔢 Mã số",        key="cvc_f_ms",  placeholder="260619...")
    _f_ct_cvc  = _fc2.text_input("🏢 Công ty",       key="cvc_f_ct",  placeholder="Tên công ty...")
    _f_nv_cvc  = _fc3.text_input("👤 Nhân viên",     key="cvc_f_nv",  placeholder="Tên nhân viên...")
    _f_cv_cvc  = _fc4.text_input("🔧 Công việc con", key="cvc_f_cv",  placeholder="Quấn dây, Lắp máy...")

    _fc5, _fc6, _fc7, _fc8 = st.columns(4)
    _f_tt_cvc  = _fc5.selectbox("📊 Tình trạng", ["Tất cả"] + sorted(df_cvc["Tình Trạng"].dropna().unique().tolist()), key="cvc_f_tt")
    _f_ts_cvc  = _fc6.selectbox("📋 Trạng thái",  ["Tất cả"] + sorted(df_cvc["Trạng Thái"].dropna().unique().tolist()),  key="cvc_f_ts")
    _f_tu_cvc  = _fc7.date_input("📅 Hoàn thành từ", value=None, format="DD/MM/YYYY", key="cvc_f_tu")
    _f_den_cvc = _fc8.date_input("📅 đến ngày",      value=None, format="DD/MM/YYYY", key="cvc_f_den")

    _fc9, _fc10, _fc11, _fc12 = st.columns(4)
    _ttm_opts = ["Tất cả"] + sorted([x for x in df_cvc["Tình Trạng Máy"].dropna().unique().tolist() if x])
    _f_ttm_cvc = _fc9.selectbox("🖥️ Tình Trạng Máy", _ttm_opts, key="cvc_f_ttm")

    df_show_cvc = df_cvc.copy()
    if _f_ms_cvc:
        df_show_cvc = df_show_cvc[df_show_cvc["Mã Số"].astype(str).str.contains(_f_ms_cvc, case=False, na=False)]
    if _f_ct_cvc:
        df_show_cvc = df_show_cvc[df_show_cvc["Tên Công Ty"].str.contains(_f_ct_cvc, case=False, na=False)]
    if _f_nv_cvc:
        df_show_cvc = df_show_cvc[df_show_cvc["Nhân Viên"].str.contains(_f_nv_cvc, case=False, na=False)]
    if _f_cv_cvc:
        df_show_cvc = df_show_cvc[df_show_cvc["Công Việc Con"].str.contains(_f_cv_cvc, case=False, na=False)]
    if _f_tt_cvc != "Tất cả":
        df_show_cvc = df_show_cvc[df_show_cvc["Tình Trạng"] == _f_tt_cvc]
    if _f_ts_cvc != "Tất cả":
        df_show_cvc = df_show_cvc[df_show_cvc["Trạng Thái"] == _f_ts_cvc]
    if _f_tu_cvc:
        df_show_cvc = df_show_cvc[df_show_cvc["_ngay_ht_raw"].apply(lambda d: d is not None and d >= _f_tu_cvc)]
    if _f_den_cvc:
        df_show_cvc = df_show_cvc[df_show_cvc["_ngay_ht_raw"].apply(lambda d: d is not None and d <= _f_den_cvc)]
    if _f_ttm_cvc != "Tất cả":
        df_show_cvc = df_show_cvc[df_show_cvc["Tình Trạng Máy"] == _f_ttm_cvc]

    total = len(df_show_cvc)
    done_count  = len(df_show_cvc[df_show_cvc["Ngày Hoàn Thành"] != ""])
    qua_han     = len(df_show_cvc[df_show_cvc["Tình Trạng"] == "Quá hạn"])
    truoc_han   = len(df_show_cvc[df_show_cvc["Tình Trạng"] == "Trước hạn"])
    kpi_cols = st.columns(4)
    kpi_cols[0].metric("📋 Tổng việc con", total)
    kpi_cols[1].metric("✅ Hoàn thành", done_count)
    kpi_cols[2].metric("🔴 Quá hạn", qua_han)
    kpi_cols[3].metric("🟢 Trước hạn", truoc_han)

    st.divider()

    cols_show = ["STT", "Tên Công Ty", "Tên Công Việc", "Công Suất", "Mã Số",
                 "Công Việc Con", "Nhân Viên", "Ngày Hoàn Thành", "Trạng Thái",
                 "Loại Máy", "Tình Trạng Máy", "Tình Trạng"]
    _df_grid_cvc = df_show_cvc[cols_show].copy().reset_index(drop=True)

    _TT_STYLE = {
        'Trước hạn': 'background-color:#dcfce7;color:#16a34a;font-weight:700',
        'Đúng hạn':  'background-color:#fef9c3;color:#d97706;font-weight:700',
        'Quá hạn':   'background-color:#fee2e2;color:#dc2626;font-weight:700',
        'Hoàn thành':'background-color:#dbeafe;color:#1d4ed8;font-weight:700',
        'Chưa xong': 'background-color:#f3f4f6;color:#6b7280;font-weight:700',
    }
    _TS_STYLE = {
        'Chờ Làm':   'background-color:#fee2e2;color:#dc2626;font-weight:600',
        'Đang Làm':  'background-color:#fef9c3;color:#d97706;font-weight:600',
        'Hoàn Thành':'background-color:#dcfce7;color:#16a34a;font-weight:600',
        'Đang Kiểm Tra':'background-color:#dbeafe;color:#1d4ed8;font-weight:600',
        'Đã Phê Duyệt':  'background-color:#dcfce7;color:#15803d;font-weight:600',
        'Đã Báo Giá':    'background-color:#fef3c7;color:#b45309;font-weight:600',
        'Có Đơn':    'background-color:#ede9fe;color:#7c3aed;font-weight:600',
        'Chờ Giao':  'background-color:#fef9c3;color:#a16207;font-weight:600',
        'Đã Hoàn Thành - Giao Máy':'background-color:#bbf7d0;color:#166534;font-weight:600',
    }
    _styler_cvc = (
        _df_grid_cvc.style
        .map(lambda v: _TT_STYLE.get(str(v), ''), subset=['Tình Trạng'])
        .map(lambda v: _TS_STYLE.get(str(v), ''), subset=['Trạng Thái'])
        .set_table_styles([{
            'selector': 'thead th',
            'props': [('background-color','#f59e0b'),('color','white'),
                      ('font-weight','bold'),('font-size','0.8rem'),
                      ('text-transform','uppercase')],
        }])
    )

    _evt_cvc = st.dataframe(
        _styler_cvc,
        use_container_width=True,
        hide_index=True,
        height=600,
        on_select="rerun",
        selection_mode="single-row",
        key="df_cvc",
        column_config={
            "STT":             st.column_config.NumberColumn("STT",            width="small"),
            "Tên Công Ty":     st.column_config.TextColumn("Tên Công Ty",      width="large"),
            "Tên Công Việc":   st.column_config.TextColumn("Tên Công Việc",    width="large"),
            "Công Suất":       st.column_config.TextColumn("Công Suất",        width="small"),
            "Mã Số":           st.column_config.TextColumn("Mã Số",            width="small"),
            "Công Việc Con":   st.column_config.TextColumn("Công Việc Con",    width="medium"),
            "Nhân Viên":       st.column_config.TextColumn("Nhân Viên",        width="medium"),
            "Ngày Hoàn Thành": st.column_config.TextColumn("Ngày HT",         width="small"),
            "Trạng Thái":      st.column_config.TextColumn("Trạng Thái",       width="medium"),
            "Loại Máy":        st.column_config.TextColumn("Loại Máy",         width="medium"),
            "Tình Trạng Máy":  st.column_config.TextColumn("Tình Trạng Máy",  width="small"),
            "Tình Trạng":      st.column_config.TextColumn("Tình Trạng",       width="small"),
        },
    )
    st.caption(f"Hiển thị {len(df_show_cvc)} / {total} công việc con · 👆 Click vào hàng để xem chi tiết")

    _sel_rows_cvc = _evt_cvc.selection.rows if hasattr(_evt_cvc, "selection") else []
    if _sel_rows_cvc:
        _sel_idx_cvc = _sel_rows_cvc[0]
        _sel_id_cvc  = str(df_show_cvc.iloc[_sel_idx_cvc]["_task_id"])
        if _sel_id_cvc != st.session_state.get("_cvc_prev_sel_id", ""):
            _match_cvc = df_show_cvc[df_show_cvc["_task_id"].astype(str) == _sel_id_cvc]
            if not _match_cvc.empty:
                st.session_state["_pending_dlg"] = (
                    _match_cvc.iloc[0]["_task_dict"],
                    lay_ten_cac_trang_thai() or _DS_TRANG_THAI_MAC_DINH,
                )
                st.session_state["_cvc_prev_sel_id"] = _sel_id_cvc
                st.rerun(scope="app")
        st.session_state["_cvc_prev_sel_id"] = _sel_id_cvc
    else:
        st.session_state["_cvc_prev_sel_id"] = ""

    cols_excel = ["STT", "Tên Công Ty", "Tên Công Việc", "Công Suất", "Mã Số",
                  "Công Việc Con", "Nhân Viên", "Ngày Hoàn Thành", "Trạng Thái",
                  "Loại Máy", "Tình Trạng Máy", "Tình Trạng"]
    df_excel = df_show_cvc[cols_excel].copy()
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df_excel.to_excel(writer, index=False, sheet_name="CongViecCon")
        ws = writer.sheets["CongViecCon"]
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        header_fill = PatternFill("solid", fgColor="F59E0B")
        thin = Side(style="thin", color="DDDDDD")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for cell in ws[1]:
            cell.font      = Font(bold=True, color="FFFFFF", size=11)
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = border
        _mau_excel = {
            "Trước hạn": "DCFCE7", "Đúng hạn": "FEF9C3",
            "Quá hạn":   "FEE2E2", "Hoàn thành": "DBEAFE",
            "Chưa xong": "F3F4F6",
        }
        col_tt_idx = cols_excel.index("Tình Trạng") + 1
        for row_idx, row_cells in enumerate(ws.iter_rows(min_row=2), start=2):
            for cell in row_cells:
                cell.border    = border
                cell.alignment = Alignment(vertical="center")
            tt_val = ws.cell(row=row_idx, column=col_tt_idx).value or ""
            if tt_val in _mau_excel:
                ws.cell(row=row_idx, column=col_tt_idx).fill = PatternFill("solid", fgColor=_mau_excel[tt_val])
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
        ws.row_dimensions[1].height = 30
    buf.seek(0)
    ten_file = f"cong_viec_con_{datetime.now().strftime('%d%m%Y_%H%M')}.xlsx"
    st.download_button(
        label="📥 Xuất Excel",
        data=buf,
        file_name=ten_file,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ============================================================
# FRAGMENT: Theo Dõi Tiến Độ Máy AgGrid
# ============================================================
def _fragment_tdm_content(df_tdm_all, aggrid_css):
    import io

    if df_tdm_all.empty:
        st.info("ℹ️ Chưa có công việc nào.")
        return

    today_tdm = datetime.today().date()
    rows_tdm  = []

    for stt_tdm, (_, r) in enumerate(df_tdm_all.iterrows(), start=1):
        han_str_tdm  = str(r.get("Hạn Hoàn Thành", "") or "")
        ngay_kt_str  = str(r.get("Ngày Kết Thúc", "") or "")
        han_date_tdm = None
        ngay_kt_date = None
        try:
            if han_str_tdm:
                han_date_tdm = datetime.strptime(han_str_tdm[:10], "%Y-%m-%d").date()
        except Exception:
            pass
        try:
            if ngay_kt_str:
                ngay_kt_date = datetime.strptime(ngay_kt_str[:10], "%Y-%m-%d").date()
        except Exception:
            pass

        if ngay_kt_date and han_date_tdm:
            if ngay_kt_date < han_date_tdm:
                tinh_trang_tdm = "Trước hạn"
            elif ngay_kt_date == han_date_tdm:
                tinh_trang_tdm = "Đúng hạn"
            else:
                tinh_trang_tdm = "Quá hạn"
        elif ngay_kt_date:
            tinh_trang_tdm = "Hoàn thành"
        elif han_date_tdm and today_tdm > han_date_tdm:
            tinh_trang_tdm = "Quá hạn"
        elif han_date_tdm and today_tdm == han_date_tdm:
            tinh_trang_tdm = "Đúng hạn"
        else:
            tinh_trang_tdm = "Chưa xong"

        ngay_nhan_hien  = ""
        ngay_giao_hien  = ""
        han_hien        = ""
        try:
            ngay_tao_str = str(r.get("Ngày Tạo", "") or "")
            if ngay_tao_str:
                ngay_nhan_hien = datetime.strptime(ngay_tao_str[:10], "%Y-%m-%d").strftime("%d/%m/%Y")
        except Exception:
            ngay_nhan_hien = str(r.get("Ngày Tạo", "") or "")
        try:
            if han_date_tdm:
                han_hien = han_date_tdm.strftime("%d/%m/%Y")
        except Exception:
            han_hien = han_str_tdm
        try:
            if ngay_kt_date:
                ngay_giao_hien = ngay_kt_date.strftime("%d/%m/%Y")
        except Exception:
            ngay_giao_hien = ngay_kt_str

        rows_tdm.append({
            "STT":              stt_tdm,
            "Tên Công Ty":      r.get("Công Ty", ""),
            "Tên Công Việc":    r.get("Tên Công Việc", ""),
            "Công Suất":        r.get("Công Suất", ""),
            "Mã Số Máy":        r.get("Mã Số", ""),
            "Số PO Nội Bộ":     r.get("Số PO Nội Bộ", ""),
            "Số PO KH/HĐ":      r.get("Số PO KH/HĐ", ""),
            "Số Báo Giá":       r.get("Số Báo Giá", ""),
            "Trạng Thái":       r.get("Trạng Thái", ""),
            "Ngày Nhận Máy":    ngay_nhan_hien,
            "Hạn Hoàn Thành":   han_hien,
            "Ngày Giao Máy":    ngay_giao_hien,
            "_han_raw":         han_date_tdm,
            "_giao_raw":        ngay_kt_date,
            "Loại Máy":         r.get("Loại Máy", ""),
            "Tình Trạng":       tinh_trang_tdm,
            "_task_id":         r.get("ID", ""),
            "_task_dict":       r.to_dict(),
        })

    if not rows_tdm:
        st.info("ℹ️ Chưa có dữ liệu.")
        return

    df_tdm = pd.DataFrame(rows_tdm)

    # ── Bộ lọc theo từng cột ──────────────────────────────────
    _tf1, _tf2, _tf3, _tf4 = st.columns(4)
    _f_ms_tdm  = _tf1.text_input("🔢 Mã số máy",  key="tdm_f_ms",  placeholder="260619...")
    _f_ct_tdm  = _tf2.text_input("🏢 Công ty",     key="tdm_f_ct",  placeholder="Tên công ty...")
    _f_po_tdm  = _tf3.text_input("📄 Số PO / BG",  key="tdm_f_po",  placeholder="DH00619, BG0277...")
    _f_lm_tdm  = _tf4.text_input("🔧 Loại máy",    key="tdm_f_lm",  placeholder="Động cơ AC...")

    _tf5, _tf6, _tf7, _tf8 = st.columns(4)
    _f_tt_tdm  = _tf5.selectbox("📊 Tình trạng",  ["Tất cả"] + sorted(df_tdm["Tình Trạng"].dropna().unique().tolist()), key="tdm_f_tt")
    _f_ts_tdm  = _tf6.selectbox("📋 Trạng thái",  ["Tất cả"] + sorted(df_tdm["Trạng Thái"].dropna().unique().tolist()),  key="tdm_f_ts")
    _f_tu_tdm  = _tf7.date_input("📅 Nhận máy từ", value=None, format="DD/MM/YYYY", key="tdm_f_tu")
    _f_den_tdm = _tf8.date_input("📅 Giao máy đến", value=None, format="DD/MM/YYYY", key="tdm_f_den")

    df_show_tdm = df_tdm.copy()
    if _f_ms_tdm:
        df_show_tdm = df_show_tdm[df_show_tdm["Mã Số Máy"].astype(str).str.contains(_f_ms_tdm, case=False, na=False)]
    if _f_ct_tdm:
        df_show_tdm = df_show_tdm[df_show_tdm["Tên Công Ty"].str.contains(_f_ct_tdm, case=False, na=False)]
    if _f_po_tdm:
        df_show_tdm = df_show_tdm[
            df_show_tdm["Số PO Nội Bộ"].astype(str).str.contains(_f_po_tdm, case=False, na=False) |
            df_show_tdm["Số PO KH/HĐ"].astype(str).str.contains(_f_po_tdm, case=False, na=False) |
            df_show_tdm["Số Báo Giá"].astype(str).str.contains(_f_po_tdm, case=False, na=False)
        ]
    if _f_lm_tdm:
        df_show_tdm = df_show_tdm[df_show_tdm["Loại Máy"].str.contains(_f_lm_tdm, case=False, na=False)]
    if _f_tt_tdm != "Tất cả":
        df_show_tdm = df_show_tdm[df_show_tdm["Tình Trạng"] == _f_tt_tdm]
    if _f_ts_tdm != "Tất cả":
        df_show_tdm = df_show_tdm[df_show_tdm["Trạng Thái"] == _f_ts_tdm]
    if _f_tu_tdm:
        df_show_tdm = df_show_tdm[df_show_tdm["_han_raw"].apply(lambda d: d is not None and d >= _f_tu_tdm)]
    if _f_den_tdm:
        df_show_tdm = df_show_tdm[df_show_tdm["_giao_raw"].apply(lambda d: d is not None and d <= _f_den_tdm)]

    total_tdm     = len(df_show_tdm)
    da_giao       = len(df_show_tdm[df_show_tdm["Ngày Giao Máy"] != ""])
    qua_han_tdm   = len(df_show_tdm[df_show_tdm["Tình Trạng"] == "Quá hạn"])
    truoc_han_tdm = len(df_show_tdm[df_show_tdm["Tình Trạng"] == "Trước hạn"])
    kpi_c = st.columns(4)
    kpi_c[0].metric("🔩 Tổng máy",    total_tdm)
    kpi_c[1].metric("✅ Đã giao",      da_giao)
    kpi_c[2].metric("🔴 Quá hạn",     qua_han_tdm)
    kpi_c[3].metric("🟢 Trước hạn",   truoc_han_tdm)

    st.divider()

    _cols_display_tdm = ["STT", "Tên Công Ty", "Tên Công Việc", "Công Suất", "Mã Số Máy",
                         "Số PO Nội Bộ", "Số PO KH/HĐ", "Số Báo Giá", "Trạng Thái",
                         "Ngày Nhận Máy", "Hạn Hoàn Thành", "Ngày Giao Máy",
                         "Loại Máy", "Tình Trạng"]

    _df_grid_tdm = df_show_tdm[_cols_display_tdm].copy().reset_index(drop=True)

    _TT_STYLE_TDM = {
        'Trước hạn': 'background-color:#dcfce7;color:#16a34a;font-weight:700',
        'Đúng hạn':  'background-color:#fef9c3;color:#d97706;font-weight:700',
        'Quá hạn':   'background-color:#fee2e2;color:#dc2626;font-weight:700',
        'Hoàn thành':'background-color:#dbeafe;color:#1d4ed8;font-weight:700',
        'Chưa xong': 'background-color:#f3f4f6;color:#6b7280;font-weight:700',
    }
    _TS_STYLE_TDM = {
        'Chờ Làm':   'background-color:#fee2e2;color:#dc2626;font-weight:600',
        'Đang Làm':  'background-color:#fef9c3;color:#d97706;font-weight:600',
        'Hoàn Thành':'background-color:#dcfce7;color:#16a34a;font-weight:600',
        'Đang Kiểm Tra':'background-color:#dbeafe;color:#1d4ed8;font-weight:600',
        'Đã Phê Duyệt':  'background-color:#dcfce7;color:#15803d;font-weight:600',
        'Đã Báo Giá':    'background-color:#fef3c7;color:#b45309;font-weight:600',
        'Có Đơn':    'background-color:#ede9fe;color:#7c3aed;font-weight:600',
        'Chờ Giao':  'background-color:#fef9c3;color:#a16207;font-weight:600',
        'Đã Hoàn Thành - Giao Máy':'background-color:#bbf7d0;color:#166534;font-weight:600',
    }
    _styler_tdm = (
        _df_grid_tdm.style
        .map(lambda v: _TT_STYLE_TDM.get(str(v), ''), subset=['Tình Trạng'])
        .map(lambda v: _TS_STYLE_TDM.get(str(v), ''), subset=['Trạng Thái'])
        .set_table_styles([{
            'selector': 'thead th',
            'props': [('background-color','#3b82f6'),('color','white'),
                      ('font-weight','bold'),('font-size','0.8rem'),
                      ('text-transform','uppercase')],
        }])
    )

    _evt_tdm = st.dataframe(
        _styler_tdm,
        use_container_width=True,
        hide_index=True,
        height=600,
        on_select="rerun",
        selection_mode="single-row",
        key="df_tdm",
        column_config={
            "STT":            st.column_config.NumberColumn("STT",           width="small"),
            "Tên Công Ty":    st.column_config.TextColumn("Tên Công Ty",     width="large"),
            "Tên Công Việc":  st.column_config.TextColumn("Tên Công Việc",   width="large"),
            "Công Suất":      st.column_config.TextColumn("Công Suất",       width="small"),
            "Mã Số Máy":      st.column_config.TextColumn("Mã Số",           width="small"),
            "Số PO Nội Bộ":   st.column_config.TextColumn("PO Nội Bộ",       width="medium"),
            "Số PO KH/HĐ":    st.column_config.TextColumn("PO KH/HĐ",        width="medium"),
            "Số Báo Giá":     st.column_config.TextColumn("Báo Giá",          width="medium"),
            "Trạng Thái":     st.column_config.TextColumn("Trạng Thái",       width="medium"),
            "Ngày Nhận Máy":  st.column_config.TextColumn("Ngày Nhận",        width="small"),
            "Hạn Hoàn Thành": st.column_config.TextColumn("Hạn HT",          width="small"),
            "Ngày Giao Máy":  st.column_config.TextColumn("Ngày Giao",        width="small"),
            "Loại Máy":       st.column_config.TextColumn("Loại Máy",         width="medium"),
            "Tình Trạng":     st.column_config.TextColumn("Tình Trạng",       width="small"),
        },
    )
    st.caption(f"Hiển thị {len(df_show_tdm)} / {total_tdm} máy · 👆 Click vào hàng để xem chi tiết")

    _sel_rows_tdm = _evt_tdm.selection.rows if hasattr(_evt_tdm, "selection") else []
    if _sel_rows_tdm:
        _sel_idx_tdm = _sel_rows_tdm[0]
        _sel_id_tdm  = str(df_show_tdm.iloc[_sel_idx_tdm]["_task_id"])
        if _sel_id_tdm != st.session_state.get("_tdm_prev_sel_id", ""):
            _match_tdm = df_show_tdm[df_show_tdm["_task_id"].astype(str) == _sel_id_tdm]
            if not _match_tdm.empty:
                st.session_state["_pending_dlg"] = (
                    _match_tdm.iloc[0]["_task_dict"],
                    lay_ten_cac_trang_thai() or _DS_TRANG_THAI_MAC_DINH,
                )
        st.session_state["_tdm_prev_sel_id"] = _sel_id_tdm
    else:
        st.session_state["_tdm_prev_sel_id"] = ""

    cols_excel_tdm = _cols_display_tdm
    df_excel_tdm = df_show_tdm[cols_excel_tdm].copy()
    buf_tdm = io.BytesIO()
    with pd.ExcelWriter(buf_tdm, engine="openpyxl") as writer_tdm:
        df_excel_tdm.to_excel(writer_tdm, index=False, sheet_name="TheoDoiTienDoMay")
        ws_tdm = writer_tdm.sheets["TheoDoiTienDoMay"]
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        hdr_fill_tdm = PatternFill("solid", fgColor="3B82F6")
        thin2 = Side(style="thin", color="DDDDDD")
        brd2  = Border(left=thin2, right=thin2, top=thin2, bottom=thin2)
        for cell in ws_tdm[1]:
            cell.font      = Font(bold=True, color="FFFFFF", size=11)
            cell.fill      = hdr_fill_tdm
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = brd2
        _mau_excel_tdm = {
            "Trước hạn": "DCFCE7", "Đúng hạn": "FEF9C3",
            "Quá hạn":   "FEE2E2", "Hoàn thành": "DBEAFE",
            "Chưa xong": "F3F4F6",
        }
        col_tt_idx_tdm = cols_excel_tdm.index("Tình Trạng") + 1
        for row_idx, row_cells in enumerate(ws_tdm.iter_rows(min_row=2), start=2):
            for cell in row_cells:
                cell.border    = brd2
                cell.alignment = Alignment(vertical="center")
            tt_val = ws_tdm.cell(row=row_idx, column=col_tt_idx_tdm).value or ""
            if tt_val in _mau_excel_tdm:
                ws_tdm.cell(row=row_idx, column=col_tt_idx_tdm).fill = PatternFill("solid", fgColor=_mau_excel_tdm[tt_val])
        for col in ws_tdm.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=8)
            ws_tdm.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
        ws_tdm.row_dimensions[1].height = 30
    buf_tdm.seek(0)
    ten_file_tdm = f"theo_doi_tien_do_may_{datetime.now().strftime('%d%m%Y_%H%M')}.xlsx"
    st.download_button(
        label="📥 Xuất Excel",
        data=buf_tdm,
        file_name=ten_file_tdm,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="tdm_xuat_excel",
        use_container_width=False,
    )


# ============================================================
# GIAO DIỆN ADMIN
# ============================================================
def giao_dien_admin():
    """Giao diện quản lý dành cho Admin — 4 tab ngang: Cài Đặt / Nhân Viên / Tạo Task / Tổng Quan."""
    st.header("🔧 Bảng Điều Khiển Admin")

    tab_cai_dat, tab_nhan_vien, tab_tao_task, tab_board, tab_cvc, tab_tdtdm, tab_thung_rac = st.tabs(
        ["⚙️  Cài Đặt", "👥  Nhân Viên", "➕  Tạo Công Việc Mới", "🗂️  Bảng Quản Lý", "📋  Công Việc Con", "🔩  Theo Dõi Tiến Độ Máy", "🗑️  Thùng Rác"]
    )

    if st.session_state.pop("_adm_goto_board", False):
        components.html("""
        <script>
        setTimeout(function() {
            var tabs = window.parent.document.querySelectorAll('button[role="tab"]');
            for (var i = 0; i < tabs.length; i++) {
                if (tabs[i].innerText.indexOf('B\u1ea3ng Qu\u1ea3n L\u00fd') !== -1) {
                    tabs[i].click(); break;
                }
            }
        }, 400);
        </script>
        """, height=0)

    # CSS chung cho cả bảng CVC và TDM (AgGrid)
    _aggrid_css = {
        ".ag-header": {"background": "linear-gradient(135deg,#f59e0b 0%,#d97706 100%) !important"},
        ".ag-header-cell-text": {"color": "white !important", "font-weight": "700 !important",
                                 "font-size": "0.8rem !important", "text-transform": "uppercase !important"},
        ".ag-header-icon": {"color": "white !important"},
        ".ag-header-cell-resize::after": {"background-color": "rgba(255,255,255,0.25) !important"},
        ".ag-floating-filter": {"background": "#e07b00 !important", "border-bottom": "1px solid #c56a00 !important"},
        ".ag-floating-filter-input": {"background": "white !important",
                                      "border": "1px solid rgba(255,255,255,0.5) !important",
                                      "border-radius": "4px !important"},
        ".ag-floating-filter .ag-input-field-input": {"color": "#333 !important", "background": "white !important"},
        ".ag-floating-filter .ag-input-field-input::placeholder": {"color": "#aaa !important"},
        ".ag-row-even": {"background-color": "#fffbeb !important"},
        ".ag-row-odd": {"background-color": "#ffffff !important"},
        ".ag-row-hover": {"background-color": "#fef3c7 !important"},
        ".ag-row-selected": {"background-color": "#fde68a !important"},
        ".ag-cell": {"font-size": "0.86rem !important", "color": "#374151 !important"},
    }

    # ── Helper: render section quản lý 1 field đơn ───────────────────────────
    def _section_don_gian(
        ten_section: str,
        key_prefix: str,
        lay_df,
        them_func,
        ten_cot: str,
        placeholder: str = "",
        mo_ta_them: str  = "",
        expanded: bool   = False,
        sua_func=None,
        xoa_func=None,
    ):
        with st.expander(ten_section, expanded=expanded):
            with st.form(f"form_{key_prefix}", clear_on_submit=True):
                gia_tri = st.text_input(mo_ta_them or "Tên *", placeholder=placeholder)
                gui = st.form_submit_button("➕ Thêm", use_container_width=True)
                if gui:
                    if not gia_tri.strip():
                        st.error("⛔ Vui lòng nhập tên!")
                    else:
                        try:
                            with st.spinner("Đang lưu..."):
                                id_moi = them_func(gia_tri.strip())
                            st.success(f"✅ Đã thêm #{id_moi}: **{gia_tri}**!")
                        except (ConnectionError, OSError, Exception) as e:
                            st.error(f"🔌 Lưu thất bại — mất kết nối mạng. Hãy thử lại.\n\n`{e}`")
            st.divider()
            try:
                df = lay_df()
                _loi_mang = False
            except Exception as e:
                df = pd.DataFrame()
                _loi_mang = str(e)
            if _loi_mang:
                st.warning("⚠️ Không thể tải dữ liệu — mất kết nối mạng.")
            elif df.empty:
                st.info("ℹ️ Chưa có dữ liệu. Hãy thêm ở form bên trên!")
            else:
                st.markdown(f"**Tổng cộng: {len(df)} mục**")
                for _i, (_, row) in enumerate(df.iterrows()):
                    rid  = str(row.get("ID", ""))
                    ten  = str(row.get(ten_cot, ""))
                    _bk  = f"{rid}_{_i}"   # key unique dù ID trùng hoặc rỗng
                    c1, c2, c3 = st.columns([7, 1, 1])
                    c1.markdown(f"**{ten}**")
                    if sua_func and c2.button("✏️", key=f"btn_edit_{key_prefix}_{_bk}", help="Chỉnh sửa"):
                        st.session_state[f"editing_{key_prefix}"] = _bk
                        st.session_state.pop(f"deleting_{key_prefix}", None)
                    if xoa_func and c3.button("🗑️", key=f"btn_del_{key_prefix}_{_bk}", help="Xóa"):
                        st.session_state[f"deleting_{key_prefix}"] = _bk
                        st.session_state.pop(f"editing_{key_prefix}", None)
                    # Form chỉnh sửa inline
                    if sua_func and st.session_state.get(f"editing_{key_prefix}") == _bk:
                        with st.form(f"form_edit_{key_prefix}_{_bk}"):
                            ten_moi = st.text_input("Tên mới *", value=ten)
                            c_save, c_cancel = st.columns(2)
                            saved    = c_save.form_submit_button("💾 Lưu", type="primary", use_container_width=True)
                            canceled = c_cancel.form_submit_button("❌ Huỷ", use_container_width=True)
                        if saved:
                            if not ten_moi.strip():
                                st.error("⛔ Tên không được để trống!")
                            else:
                                try:
                                    with st.spinner("Đang lưu..."):
                                        sua_func(rid, ten_moi.strip())
                                    st.session_state.pop(f"editing_{key_prefix}", None)
                                    st.rerun()
                                except Exception as e:
                                    st.error(f"🔌 Lỗi: {e}")
                        if canceled:
                            st.session_state.pop(f"editing_{key_prefix}", None)
                            st.rerun()
                    # Xác nhận xóa
                    if xoa_func and st.session_state.get(f"deleting_{key_prefix}") == _bk:
                        st.warning(f"⚠️ Xác nhận xóa **{ten}**?")
                        c_ok, c_no = st.columns(2)
                        if c_ok.button("✅ Xác nhận xóa", key=f"ok_del_{key_prefix}_{_bk}", use_container_width=True):
                            try:
                                with st.spinner("Đang xóa..."):
                                    xoa_func(rid)
                                st.session_state.pop(f"deleting_{key_prefix}", None)
                                st.rerun()
                            except Exception as e:
                                st.error(f"🔌 Lỗi: {e}")
                        if c_no.button("❌ Huỷ", key=f"no_del_{key_prefix}_{_bk}", use_container_width=True):
                            st.session_state.pop(f"deleting_{key_prefix}", None)
                            st.rerun()

    # ══════════════════════════════════════════════
    # TAB 1 — CÀI ĐẶT
    # ══════════════════════════════════════════════
    with tab_cai_dat:
        with st.expander("🏢 Quản Lý Công Ty", expanded=False):
            with st.form("form_cong_ty", clear_on_submit=True):
                ten_ct_inp = st.text_input("Tên Công Ty *", placeholder="Ví dụ: Công Ty TNHH ABC")
                dc_inp = st.text_input("Địa Chỉ", placeholder="Ví dụ: 123 Nguyễn Văn Linh, Q.7, TP.HCM")
                col_mkh, col_mst = st.columns(2)
                with col_mkh:
                    ma_kh_inp = st.text_input("Mã Khách Hàng", placeholder="Ví dụ: KH001")
                with col_mst:
                    ma_thue_inp = st.text_input("Mã Số Thuế", placeholder="Ví dụ: 0123456789")
                gui_ct = st.form_submit_button("➕ Thêm", use_container_width=True)
                if gui_ct:
                    if not ten_ct_inp.strip():
                        st.error("⛔ Vui lòng nhập Tên Công Ty!")
                    else:
                        try:
                            with st.spinner("Đang lưu..."):
                                id_ct = them_cong_ty(ten_ct_inp.strip(), dc_inp.strip(), ma_kh_inp.strip(), ma_thue_inp.strip())
                            st.success(f"✅ Đã thêm #{id_ct}: **{ten_ct_inp}**!")
                        except (ConnectionError, OSError, Exception) as e:
                            st.error(f"🔌 Lưu thất bại — mất kết nối mạng. Hãy thử lại.\n\n`{e}`")
            st.divider()
            try:
                df_ct = lay_danh_sach_cong_ty()
                _loi_ct = False
            except Exception as e:
                df_ct = pd.DataFrame()
                _loi_ct = str(e)
            if _loi_ct:
                st.warning("⚠️ Không thể tải dữ liệu — mất kết nối mạng.")
            elif df_ct.empty:
                st.info("ℹ️ Chưa có công ty nào. Hãy thêm ở form bên trên!")
            else:
                _ct_filter = st.text_input(
                    "🔍 Tìm công ty...", key="ct_filter_input",
                    placeholder="Nhập tên, địa chỉ, mã KH hoặc mã số thuế...",
                    label_visibility="collapsed",
                )
                if _ct_filter.strip():
                    _f = _ct_filter.strip().lower()
                    df_ct = df_ct[
                        df_ct["Tên Công Ty"].fillna("").str.lower().str.contains(_f, na=False) |
                        df_ct["Địa Chỉ"].fillna("").str.lower().str.contains(_f, na=False) |
                        df_ct["Mã Khách Hàng"].fillna("").str.lower().str.contains(_f, na=False) |
                        df_ct["Mã Số Thuế"].fillna("").str.lower().str.contains(_f, na=False)
                    ]
                st.markdown(f"**Tổng cộng: {len(df_ct)} công ty**")
                for _, row_ct in df_ct.iterrows():
                    rid_ct  = str(row_ct.get("ID", ""))
                    ten_ct  = str(row_ct.get("Tên Công Ty", ""))
                    dc_ct   = str(row_ct.get("Địa Chỉ", ""))
                    mkh_ct  = str(row_ct.get("Mã Khách Hàng", ""))
                    mst_ct  = str(row_ct.get("Mã Số Thuế", ""))
                    c1, c2, c3 = st.columns([7, 1, 1])
                    c1.markdown(f"**{ten_ct}**" + (f"  —  {dc_ct}" if dc_ct else ""))
                    if c2.button("✏️", key=f"btn_edit_ct_{rid_ct}", help="Chỉnh sửa"):
                        st.session_state["editing_ct"] = rid_ct
                        st.session_state.pop("deleting_ct", None)
                    if c3.button("🗑️", key=f"btn_del_ct_{rid_ct}", help="Xóa"):
                        st.session_state["deleting_ct"] = rid_ct
                        st.session_state.pop("editing_ct", None)
                    # Form chỉnh sửa Công Ty
                    if st.session_state.get("editing_ct") == rid_ct:
                        with st.form(f"form_edit_ct_{rid_ct}"):
                            ten_ct_e  = st.text_input("Tên Công Ty *", value=ten_ct)
                            dc_e      = st.text_input("Địa Chỉ", value=dc_ct)
                            col_a, col_b = st.columns(2)
                            mkh_e  = col_a.text_input("Mã Khách Hàng", value=mkh_ct)
                            mst_e  = col_b.text_input("Mã Số Thuế", value=mst_ct)
                            c_s, c_c = st.columns(2)
                            saved_ct    = c_s.form_submit_button("💾 Lưu", type="primary", use_container_width=True)
                            canceled_ct = c_c.form_submit_button("❌ Huỷ", use_container_width=True)
                        if saved_ct:
                            if not ten_ct_e.strip():
                                st.error("⛔ Tên Công Ty không được để trống!")
                            else:
                                try:
                                    with st.spinner("Đang lưu..."):
                                        sua_cong_ty(rid_ct, ten_ct_e.strip(), dc_e.strip(), mkh_e.strip(), mst_e.strip())
                                    st.session_state.pop("editing_ct", None)
                                    st.rerun()
                                except Exception as e:
                                    st.error(f"🔌 Lỗi: {e}")
                        if canceled_ct:
                            st.session_state.pop("editing_ct", None)
                            st.rerun()
                    # Xác nhận xóa Công Ty
                    if st.session_state.get("deleting_ct") == rid_ct:
                        st.warning(f"⚠️ Xác nhận xóa **{ten_ct}**?")
                        c_ok, c_no = st.columns(2)
                        if c_ok.button("✅ Xác nhận xóa", key=f"ok_del_ct_{rid_ct}", use_container_width=True):
                            try:
                                with st.spinner("Đang xóa..."):
                                    xoa_cong_ty(rid_ct)
                                st.session_state.pop("deleting_ct", None)
                                st.rerun()
                            except Exception as e:
                                st.error(f"🔌 Lỗi: {e}")
                        if c_no.button("❌ Huỷ", key=f"no_del_ct_{rid_ct}", use_container_width=True):
                            st.session_state.pop("deleting_ct", None)
                            st.rerun()
        _section_don_gian(
            "🔧 Loại Máy", "loai_may",
            lay_danh_sach_loai_may, them_loai_may, "Tên Loại Máy",
            placeholder="Ví dụ: Động Cơ Điện, Máy Bơm...",
            mo_ta_them="Tên Loại Máy *",
            sua_func=sua_loai_may,
            xoa_func=xoa_loai_may,
        )
        _section_don_gian(
            "🛠️ Tình Trạng", "tinh_trang",
            lay_danh_sach_tinh_trang, them_tinh_trang, "Tên Tình Trạng",
            placeholder="Ví dụ: Bảo hành, Sửa chữa, Trả lại...",
            mo_ta_them="Tên Tình Trạng *",
            sua_func=sua_tinh_trang,
            xoa_func=xoa_tinh_trang,
        )
        _seed_trang_thai_mac_dinh()
        _section_don_gian(
            "📋 Trạng Thái", "trang_thai",
            lay_danh_sach_trang_thai_custom, them_trang_thai_custom, "Tên Trạng Thái",
            placeholder="Ví dụ: Đang Kiểm Tra, Đã Phê Duyệt, Có Đơn...",
            mo_ta_them="Tên Trạng Thái *",
            sua_func=sua_trang_thai_custom,
            xoa_func=xoa_trang_thai_custom,
        )

        _section_don_gian(
            "⚙️ Công Đoạn", "cong_doan",
            lay_danh_sach_cong_doan, them_cong_doan, "Tên Công Đoạn",
            placeholder="Ví dụ: Kiểm Tra Đầu Vào, Tháo Rã, Quấn Dây...",
            mo_ta_them="Tên Công Đoạn *",
            sua_func=sua_cong_doan,
            xoa_func=xoa_cong_doan_item,
        )

        st.divider()
        st.markdown("#### 🔄 Đồng bộ Công Đoạn vào các Task cũ")
        st.caption("Thêm các công đoạn còn thiếu vào **tất cả task** (giữ nguyên dữ liệu đã có, chỉ bổ sung mục mới).")
        if st.button("⚡ Đồng bộ ngay", key="adm_apply_cd_bulk", type="primary"):
            _ds_cd = lay_ten_cac_cong_doan()
            if not _ds_cd:
                st.warning("Chưa có công đoạn nào. Hãy thêm công đoạn trước.")
            else:
                with st.spinner("Đang đồng bộ..."):
                    try:
                        df_all = lay_danh_sach_cong_viec()
                        sheet  = _lay_sheet_fresh()
                        _col_a = sheet.col_values(1)
                        _batch = []   # danh sách range để batch_update 1 lần

                        for _, row in df_all.iterrows():
                            _cv_raw = str(row.get("Công Việc Con", "") or "").strip()
                            try:
                                _hien_co = json.loads(_cv_raw) if _cv_raw else []
                                if not isinstance(_hien_co, list):
                                    _hien_co = []
                            except Exception:
                                _hien_co = []

                            _ten_hien_co = {item.get("ten", "") for item in _hien_co}
                            _them_moi = [
                                {"ten": cd, "nhan_vien": "", "done": False}
                                for cd in _ds_cd if cd not in _ten_hien_co
                            ]
                            if not _them_moi:
                                continue

                            _cv_merged = _hien_co + _them_moi
                            _cv_json   = json.dumps(_cv_merged, ensure_ascii=False)

                            _tid = str(row.get("ID", ""))
                            try:
                                r_idx = _col_a.index(_tid) + 1
                            except ValueError:
                                continue

                            _batch.append({
                                "range": f"N{r_idx}",
                                "values": [[_cv_json]],
                            })

                        if _batch:
                            # 1 API call duy nhất thay vì N calls
                            sheet.batch_update(_batch)
                            lay_danh_sach_cong_viec.clear()
                            st.success(f"✅ Đã đồng bộ {len(_batch)} task (thêm công đoạn còn thiếu).")
                        else:
                            st.info("Tất cả task đã có đầy đủ công đoạn.")
                    except Exception as e:
                        st.error(f"Lỗi: {e}")

        st.divider()
        st.markdown("#### 🔁 Đổi trạng thái hàng loạt")
        st.caption("Đổi tất cả task đang mang trạng thái cũ không hợp lệ sang trạng thái mới.")
        _col_mig1, _col_mig2, _col_mig3 = st.columns([2, 2, 1])
        with _col_mig1:
            _mig_cu = st.text_input("Trạng thái cũ cần đổi", value="Đang Làm", key="adm_mig_cu")
        with _col_mig2:
            _ds_tt_mig = lay_ten_cac_trang_thai() or _DS_TRANG_THAI_MAC_DINH
            _mig_moi_idx = _ds_tt_mig.index("Đang Kiểm Tra") if "Đang Kiểm Tra" in _ds_tt_mig else 0
            _mig_moi = st.selectbox("Đổi sang", _ds_tt_mig, index=_mig_moi_idx, key="adm_mig_moi")
        with _col_mig3:
            st.markdown("<br>", unsafe_allow_html=True)
            _btn_mig = st.button("⚡ Đổi ngay", key="adm_mig_btn", type="primary", use_container_width=True)
        if _btn_mig and _mig_cu.strip():
            with st.spinner(f"Đang đổi '{_mig_cu}' → '{_mig_moi}'..."):
                try:
                    df_all = lay_danh_sach_cong_viec()
                    df_can_doi = df_all[df_all["Trạng Thái"].fillna("") == _mig_cu.strip()]
                    if df_can_doi.empty:
                        st.info(f"Không tìm thấy task nào có trạng thái '{_mig_cu}'.")
                    else:
                        sheet  = _lay_sheet_fresh()
                        _col_a = sheet.col_values(1)
                        _batch_mig = []
                        for _, row in df_can_doi.iterrows():
                            _tid = str(row.get("ID", ""))
                            try:
                                r_idx = _col_a.index(_tid) + 1
                            except ValueError:
                                continue
                            _batch_mig.append({"range": f"H{r_idx}", "values": [[_mig_moi]]})
                        if _batch_mig:
                            sheet.batch_update(_batch_mig)
                            lay_danh_sach_cong_viec.clear()
                            st.success(f"✅ Đã đổi {len(_batch_mig)} task: '{_mig_cu}' → '{_mig_moi}'.")
                except Exception as e:
                    st.error(f"Lỗi: {e}")

    # ══════════════════════════════════════════════
    # TAB 2 — NHÂN VIÊN
    # ══════════════════════════════════════════════
    with tab_nhan_vien:
        st.markdown("#### ➕ Tạo Tài Khoản Nhân Viên Mới")
        with st.expander("📝 Điền thông tin đăng ký", expanded=False):
            with st.form("adm_form_dang_ky_nv", clear_on_submit=True):
                _nv_ho_ten   = st.text_input("👤 Họ và tên *", placeholder="Nguyễn Văn A")
                _nv_username = st.text_input("🔑 Username *", placeholder="nguyenvana")
                _nv_mat_khau = st.text_input("🔒 Mật khẩu *", type="password", placeholder="Ít nhất 6 ký tự")
                _nv_ngay_sinh = st.date_input("🎂 Ngày sinh", value=None,
                                              min_value=date(1970, 1, 1),
                                              max_value=date(2010, 12, 31))
                _nv_submitted = st.form_submit_button("✅ Tạo tài khoản", use_container_width=True)

            if _nv_submitted:
                _ns_str = str(_nv_ngay_sinh) if _nv_ngay_sinh else ""
                _ok, _msg = dang_ky_tai_khoan(
                    username=_nv_username,
                    mat_khau=_nv_mat_khau,
                    ho_ten=_nv_ho_ten,
                    ngay_sinh=_ns_str,
                    vai_tro="nhan_vien",
                )
                if _ok:
                    st.success(f"✅ Đã tạo tài khoản **{_nv_ho_ten}** (@{_nv_username.strip()}) thành công!")
                else:
                    st.error(f"❌ {_msg}")

        st.markdown("---")
        st.markdown("#### 👥 Danh Sách Nhân Viên Đã Đăng Ký")

        if st.button("🔄 Làm mới", key="adm_nv_refresh"):
            lay_danh_sach_users.clear()
            lay_danh_sach_nhan_vien.clear()
            lay_danh_sach_cong_viec.clear()
            st.rerun()

        df_users = lay_danh_sach_users()
        df_nv_users = df_users[df_users["VaiTro"] == "nhan_vien"].reset_index(drop=True) if not df_users.empty else pd.DataFrame()

        if df_nv_users.empty:
            st.info("ℹ️ Chưa có nhân viên nào đăng ký. Hãy chia sẻ link app để họ tự đăng ký.")
        else:
            # Load tasks once for sub-task lookup
            with st.spinner("Đang tải dữ liệu công việc..."):
                df_all_tasks = lay_danh_sach_cong_viec()

            # ── CSS card nhân viên ──────────────────────────────────────
            st.markdown("""
            <style>
            .nv-summary-card {
                background: white;
                border: 1px solid #e5e7eb;
                border-radius: 12px 12px 0 0;
                padding: 14px 16px 12px 16px;
                margin-bottom: 0;
                display: flex;
                flex-direction: column;
                gap: 6px;
            }
            .nv-top-row {
                display: flex;
                align-items: center;
                gap: 12px;
            }
            .nv-avatar {
                width: 40px; height: 40px;
                border-radius: 50%;
                background: linear-gradient(135deg, #7c3aed, #a78bfa);
                display: flex; align-items: center; justify-content: center;
                font-size: 1.1rem; color: white; font-weight: 700;
                flex-shrink: 0;
            }
            .nv-name { font-weight: 700; font-size: 0.97rem; color: #1e1b4b; }
            .nv-meta { font-size: 0.78rem; color: #6b7280; }
            .nv-badge-row {
                background: #f5f3ff;
                border: 1px solid #e5e7eb;
                border-top: none;
                border-radius: 0 0 12px 12px;
                padding: 6px 16px;
                font-size: 0.78rem;
                color: #7c3aed;
                font-weight: 600;
                margin-bottom: 8px;
            }
            /* Nút đóng trang chi tiết */
            button[kind="secondary"][data-testid*="adm_close_nv"] {
                border-color: #ef4444 !important;
                color: #ef4444 !important;
            }
            </style>
            """, unsafe_allow_html=True)

            # ── Build index subtask một lần cho toàn bộ nhân viên ────
            # subtask_index: {ho_ten_lower -> [(hang_dict, ds_cv_full, [(i, cv)])]}
            # task_chinh_count: {ho_ten_lower -> int}
            _subtask_index: dict = {}   # ho_ten_lower → list[(hang_dict, ds_cv, my_sub)]
            _task_ct_count: dict = {}   # ho_ten_lower → int
            if not df_all_tasks.empty:
                # Đếm task chính vectorized
                for _ht in df_all_tasks["Nhân Viên"].fillna("").str.strip().str.lower().unique():
                    _task_ct_count[_ht] = int((df_all_tasks["Nhân Viên"].fillna("").str.strip().str.lower() == _ht).sum())

                # Parse JSON một lần, build subtask index
                for _, _row in df_all_tasks.iterrows():
                    _raw = _row.get("Công Việc Con", "") or "[]"
                    try:
                        _ds_cv = json.loads(_raw)
                    except Exception:
                        _ds_cv = []
                    if not isinstance(_ds_cv, list):
                        continue
                    # Group CV con theo nhân viên được giao
                    _by_nv: dict = {}
                    for _i, _cv in enumerate(_ds_cv):
                        if not isinstance(_cv, dict):
                            continue
                        _nv_key = ((_cv.get("nguoi") or _cv.get("nhan_vien") or _cv.get("Nhân Viên") or "")).strip().lower()
                        if _nv_key:
                            _by_nv.setdefault(_nv_key, []).append((_i, _cv))
                    for _nv_key, _my_sub in _by_nv.items():
                        _subtask_index.setdefault(_nv_key, []).append((_row.to_dict(), _ds_cv, _my_sub))

            def _dem_task_cua(ho_ten: str):
                _k = ho_ten.strip().lower()
                return _task_ct_count.get(_k, 0), sum(len(m) for _, _, m in _subtask_index.get(_k, []))

            st.markdown(f"**Tổng cộng: {len(df_nv_users)} nhân viên**")
            st.markdown("---")

            _ICON_TT_NV = {
                "Đang Kiểm Tra": "🔵", "Đã Phê Duyệt": "🟢",
                "Đã Báo Giá": "🟠", "Có Đơn": "🟣", "Chờ Giao": "🟡",
                "Đã Hoàn Thành - Giao Máy": "✅", "Đã Xuất Hóa Đơn": "⬜",
                "Bảo Hành - Trả Lại": "🔴",
                "Chờ Làm": "🔴", "Đang Làm": "🟡", "Hoàn Thành": "🟢",
            }

            # ── TRANG CHI TIẾT TASK CỦA NHÂN VIÊN ─────────────────────
            adm_xem_nv = st.session_state.get("adm_xem_nv")

            if adm_xem_nv:
                # Nút đóng
                col_close, col_title = st.columns([1, 5])
                with col_close:
                    if st.button("✕ Đóng", key="adm_close_nv", use_container_width=True):
                        st.session_state.pop("adm_xem_nv", None)
                        st.rerun()
                with col_title:
                    tc_d, ts_d = _dem_task_cua(adm_xem_nv)
                    st.markdown(f"### 👤 {adm_xem_nv} &nbsp;·&nbsp; 📋 {tc_d} công việc chính &nbsp;·&nbsp; 🔹 {ts_d} việc con")

                st.divider()

                # Nút làm mới + xuất Excel
                _col_rf, _col_xl = st.columns([1, 1])
                with _col_rf:
                    if st.button("🔄 Làm mới dữ liệu", key="adm_nv_detail_refresh"):
                        lay_danh_sach_cong_viec.clear()
                        st.rerun()

                ds_tt_adm = lay_ten_cac_trang_thai() or _DS_TRANG_THAI_MAC_DINH

                # ── Task chính ──────────────────────────────────────────
                df_task_chinh = df_all_tasks[
                    df_all_tasks["Nhân Viên"].fillna("").str.strip().str.lower() == adm_xem_nv.lower()
                ].copy() if not df_all_tasks.empty else pd.DataFrame()

                # Lọc theo Công Ty
                ds_ct_filter = df_task_chinh["Công Ty"].dropna().unique().tolist() if not df_task_chinh.empty else []
                if ds_ct_filter:
                    loc_ct = st.multiselect(
                        "🏢 Lọc theo Công Ty",
                        options=ds_ct_filter,
                        placeholder="Hiển thị tất cả...",
                        key="adm_nv_loc_ct",
                    )
                    if loc_ct:
                        df_task_chinh = df_task_chinh[df_task_chinh["Công Ty"].isin(loc_ct)]

                # ── Subtask được giao — lấy từ index đã build ─────────
                tasks_co_subtask = _subtask_index.get(adm_xem_nv.strip().lower(), [])

                # ── Tách việc đang làm / đã hoàn thành theo subtask ───
                tasks_dang_lam = []   # task có ít nhất 1 CV con chưa xong
                tasks_da_xong  = []   # task có ít nhất 1 CV con đã xong
                for hang_dict, ds_cv_full, my_subtasks in tasks_co_subtask:
                    undone = [(i, cv) for i, cv in my_subtasks if not cv.get("done", False)]
                    done   = [(i, cv) for i, cv in my_subtasks if cv.get("done", False)]
                    if undone:
                        tasks_dang_lam.append((hang_dict, ds_cv_full, undone))
                    if done:
                        tasks_da_xong.append((hang_dict, ds_cv_full, done))

                n_dang_lam = sum(len(m) for _, _, m in tasks_dang_lam)
                n_da_xong  = sum(len(m) for _, _, m in tasks_da_xong)

                # ── Nút xuất Excel ─────────────────────────────────────
                with _col_xl:
                    def _xuat_excel_nv(ten_nv, dl_lam, dl_xong):
                        import io as _io
                        from openpyxl import Workbook as _WB
                        from openpyxl.styles import (Font as _F, PatternFill as _PF,
                                                      Alignment as _AL, Border as _BD, Side as _SD)
                        _wb = _WB(); _ws = _wb.active; _ws.title = "Công Việc NV"
                        _thin = _SD(style="thin", color="000000")
                        _brd  = _BD(left=_thin, right=_thin, top=_thin, bottom=_thin)
                        _hdr_fill = _PF("solid", fgColor="4472C4")
                        _done_fill = _PF("solid", fgColor="E2EFDA")
                        _todo_fill = _PF("solid", fgColor="FFF2CC")
                        _al_c = _AL(horizontal="center", vertical="center", wrap_text=True)
                        _al_l = _AL(horizontal="left",   vertical="center", wrap_text=True)

                        # Tiêu đề
                        _ws.merge_cells("A1:F1")
                        _tc = _ws.cell(1, 1, f"Báo cáo công việc: {ten_nv}")
                        _tc.font = _F(bold=True, size=14, name="Times New Roman")
                        _tc.alignment = _al_c
                        _ws.row_dimensions[1].height = 26

                        # Header cột
                        headers = ["STT", "Tên Task (Mã số)", "Công Ty", "Công Việc Con",
                                   "Trạng Thái Task", "Ngày Hoàn Thành"]
                        for ci, h in enumerate(headers, 1):
                            _c = _ws.cell(2, ci, h)
                            _c.font = _F(bold=True, size=11, color="FFFFFF", name="Times New Roman")
                            _c.fill = _hdr_fill
                            _c.alignment = _al_c
                            _c.border = _brd
                        _ws.row_dimensions[2].height = 22

                        _row = 3

                        def _write_section(title, tasks_list, fill, is_done):
                            nonlocal _row
                            # Section header
                            _ws.merge_cells(f"A{_row}:F{_row}")
                            _sc2 = _ws.cell(_row, 1, title)
                            _sc2.font = _F(bold=True, size=12, name="Times New Roman")
                            _sc2.fill = fill
                            _sc2.alignment = _al_l
                            _ws.row_dimensions[_row].height = 20
                            _row += 1
                            stt = 1
                            for hd, _, msub in tasks_list:
                                ten_t  = str(hd.get("Tên Công Việc") or "")
                                ma_so  = str(hd.get("Mã Số") or "")
                                cty    = str(hd.get("Công Ty") or "")
                                tt     = str(hd.get("Trạng Thái") or "")
                                lbl    = f"{ten_t} ({ma_so})" if ma_so else ten_t
                                for _, cv in msub:
                                    ten_cv = str(cv.get("ten") or cv.get("Tên") or "")
                                    ngay   = str(cv.get("ngay_hoan_thanh") or "") if is_done else ""
                                    vals   = [stt, lbl, cty, ten_cv, tt, ngay]
                                    for ci, v in enumerate(vals, 1):
                                        _c2 = _ws.cell(_row, ci, v)
                                        _c2.font = _F(size=11, name="Times New Roman")
                                        _c2.alignment = _al_l if ci > 1 else _al_c
                                        _c2.border = _brd
                                    _ws.row_dimensions[_row].height = 18
                                    _row += 1
                                    stt += 1

                        _todo_sec_fill = _PF("solid", fgColor="FFE699")
                        _done_sec_fill = _PF("solid", fgColor="A9D18E")
                        _write_section(f"⏳ VIỆC ĐANG LÀM ({sum(len(m) for _,_,m in dl_lam)})",
                                       dl_lam, _todo_sec_fill, False)
                        _write_section(f"✅ VIỆC ĐÃ HOÀN THÀNH ({sum(len(m) for _,_,m in dl_xong)})",
                                       dl_xong, _done_sec_fill, True)

                        # Column widths
                        for ci, w in enumerate([6, 38, 28, 32, 22, 18], 1):
                            _ws.column_dimensions[chr(64+ci)].width = w

                        _buf = _io.BytesIO(); _wb.save(_buf); _buf.seek(0)
                        return _buf.read()

                    _xl_bytes = _xuat_excel_nv(adm_xem_nv, tasks_dang_lam, tasks_da_xong)
                    st.download_button(
                        "📥 Xuất Excel",
                        data=_xl_bytes,
                        file_name=f"cong_viec_{adm_xem_nv.replace(' ', '_')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )

                # ── Tabs việc đang làm / đã hoàn thành ────────────────
                tab_tc_d, tab_st_d = st.tabs([
                    f"⏳ Việc Đang Làm ({n_dang_lam})",
                    f"✅ Việc Đã Hoàn Thành ({n_da_xong})",
                ])

                def _render_cv_con_tab(tasks_list, show_done: bool):
                    if not tasks_list:
                        msg = "Không có việc nào đã hoàn thành." if show_done else "Không có việc nào đang làm."
                        st.info(msg)
                        return
                    n_cv = sum(len(m) for _, _, m in tasks_list)
                    st.caption(f"**{n_cv} việc con** trong **{len(tasks_list)} task**")
                    for hang_dict, _, my_sub in tasks_list:
                        tid    = hang_dict["ID"]
                        tt     = hang_dict.get("Trạng Thái", "")
                        icon   = _ICON_TT_NV.get(tt, "⚪")
                        cty    = hang_dict.get("Công Ty", "")
                        ten    = hang_dict.get("Tên Công Việc", "")
                        nv_chu = hang_dict.get("Nhân Viên", "")
                        dl     = hang_dict.get("Hạn Hoàn Thành", "") or hang_dict.get("Deadline", "")
                        with st.expander(
                            f"{icon} [{cty}] {ten}  —  {tt}",
                            expanded=not show_done
                        ):
                            st.markdown(f"**🏢 Công Ty:** {cty}  |  **👤** {nv_chu}  |  **📅** {dl}")
                            for _, cv in my_sub:
                                ten_cv_sub = cv.get("ten", cv.get("Tên", "—"))
                                ngay_ht    = cv.get("ngay_hoan_thanh", "")
                                if show_done:
                                    st.markdown(f"✅ ~~{ten_cv_sub}~~" + (f" &nbsp;`Hoàn thành: {ngay_ht}`" if ngay_ht else ""))
                                else:
                                    dl_sub = cv.get("deadline", cv.get("Deadline", ""))
                                    st.markdown(f"⏳ **{ten_cv_sub}**" + (f" &nbsp;`Hạn: {dl_sub}`" if dl_sub else ""))

                with tab_tc_d:
                    _render_cv_con_tab(tasks_dang_lam, show_done=False)

                with tab_st_d:
                    # ── Bộ lọc tháng/năm ──────────────────────────────
                    # Gom tất cả ngay_hoan_thanh có trong tasks_da_xong
                    _thang_nam_set = set()
                    for _hd, _, _msub in tasks_da_xong:
                        for _, _cv in _msub:
                            _nht = str(_cv.get("ngay_hoan_thanh") or "").strip()
                            if len(_nht) >= 7:   # "YYYY-MM..."
                                _thang_nam_set.add(_nht[:7])   # "YYYY-MM"
                    _thang_nam_sorted = sorted(_thang_nam_set, reverse=True)

                    if _thang_nam_sorted:
                        # Hiển thị dạng "Tháng MM/YYYY"
                        def _fmt_ym(ym):
                            try:
                                y, m = ym.split("-")
                                return f"Tháng {int(m):02d}/{y}"
                            except Exception:
                                return ym
                        _ym_options = ["Tất cả"] + [_fmt_ym(ym) for ym in _thang_nam_sorted]
                        _ym_map = {"Tất cả": None, **{_fmt_ym(ym): ym for ym in _thang_nam_sorted}}
                        _loc_ym = st.selectbox(
                            "📅 Lọc theo tháng hoàn thành",
                            options=_ym_options,
                            key=f"adm_nv_loc_ym_{adm_xem_nv}",
                        )
                        _filter_ym = _ym_map.get(_loc_ym)

                        # Lọc tasks_da_xong theo tháng được chọn
                        if _filter_ym:
                            _tasks_filtered = []
                            for _hd, _dcvf, _msub in tasks_da_xong:
                                _msub_f = [
                                    (_i, _cv) for _i, _cv in _msub
                                    if str(_cv.get("ngay_hoan_thanh") or "").startswith(_filter_ym)
                                ]
                                if _msub_f:
                                    _tasks_filtered.append((_hd, _dcvf, _msub_f))
                        else:
                            _tasks_filtered = tasks_da_xong
                    else:
                        _tasks_filtered = tasks_da_xong

                    _render_cv_con_tab(_tasks_filtered, show_done=True)

            else:
                # ── DANH SÁCH CARD NHÂN VIÊN ───────────────────────────
                for _idx_u, u in df_nv_users.iterrows():
                    ho_ten_u  = u["HoTen"]
                    uname     = u["Username"]
                    ns        = u["NgaySinh"]
                    ngay_tao  = u.get("NgayTao", "")[:10]
                    avatar_char = (ho_ten_u[0] if ho_ten_u else "?").upper()
                    tc, ts = _dem_task_cua(ho_ten_u)

                    st.markdown(f"""
                    <div class="nv-summary-card">
                        <div class="nv-top-row">
                            <div class="nv-avatar">{avatar_char}</div>
                            <div class="nv-name">{ho_ten_u}</div>
                        </div>
                        <div class="nv-meta">@{uname} &nbsp;·&nbsp; 🎂 {ns or '—'} &nbsp;·&nbsp; 📅 Tham gia {ngay_tao}</div>
                    </div>
                    <div class="nv-badge-row">📋 {tc} công việc chính &nbsp;|&nbsp; 🔹 {ts} việc con</div>
                    """, unsafe_allow_html=True)
                    if st.button("👁️ Xem chi tiết", key=f"adm_xem_{_idx_u}_{uname}", use_container_width=True):
                        lay_danh_sach_cong_viec.clear()
                        st.session_state["adm_xem_nv"] = ho_ten_u
                        st.rerun()
                    st.markdown("<div style='margin-bottom:6px'></div>", unsafe_allow_html=True)

    # ══════════════════════════════════════════════
    # TAB 3 — TẠO TASK MỚI
    # ══════════════════════════════════════════════
    with tab_tao_task:
        if st.session_state.get("_adm_task_success"):
            _dialog_tao_task_thanh_cong(st.session_state.pop("_adm_task_success"), "_adm_goto_board")

        @st.fragment
        def _fragment_tao_task_admin():
            _ADM_PREFIX   = "adm"
            ds_cong_ty    = lay_ten_cac_cong_ty()
            ds_nhan_vien  = lay_danh_sach_nhan_vien()
            ds_trang_thai = lay_ten_cac_trang_thai() or _DS_TRANG_THAI_MAC_DINH

            if not ds_cong_ty:
                st.warning("⚠️ Chưa có công ty nào! Hãy thêm ở tab **⚙️ Cài Đặt** trước.")

            adm_cong_ty = st.selectbox(
                "🏢 Công Ty Khách Hàng *",
                options=ds_cong_ty if ds_cong_ty else ["(Chưa có công ty)"],
                key="adm_cong_ty",
            )

            col_tt_top, col_nam = st.columns(2)
            with col_tt_top:
                adm_trang_thai = st.selectbox("📋 Trạng thái", options=ds_trang_thai, key="adm_trang_thai")
            with col_nam:
                adm_nam = st.text_input("Năm *", value=str(datetime.now().year), key="adm_nam")

            adm_nguoi_giao = ""
            adm_deadline = st.date_input("📅 Hạn hoàn thành", key="adm_deadline")

            col_lm_adm, col_tt_adm = st.columns(2)
            with col_lm_adm:
                adm_loai_may = st.selectbox(
                    "🔧 Loại Máy",
                    options=["-- Không chọn --"] + lay_ten_cac_loai_may(),
                    key="adm_loai_may",
                )
            with col_tt_adm:
                adm_tinh_trang = st.selectbox(
                    "🛠️ Tình Trạng",
                    options=["-- Không chọn --"] + lay_ten_cac_tinh_trang(),
                    key="adm_tinh_trang",
                )

            col_cs_adm, col_sc_adm = st.columns(2)
            with col_cs_adm:
                adm_cong_suat = st.text_input("⚡ Công Suất", placeholder="VD: 5.5kW", key="adm_cong_suat")
            with col_sc_adm:
                adm_so_cuc = st.text_input("🔩 Số Cực", placeholder="VD: 4P", key="adm_so_cuc")

            col_ms_adm, col_po_adm = st.columns(2)
            with col_ms_adm:
                adm_ma_so = st.text_input("🏷️ Mã Số", placeholder="VD: ABC-001", key="adm_ma_so")
            with col_po_adm:
                adm_so_po_noi_bo = st.text_input("📄 Số PO Nội Bộ", placeholder="VD: PO-2024-001", key="adm_so_po_noi_bo")

            col_kh_adm, col_bg_adm = st.columns(2)
            with col_kh_adm:
                adm_so_po_kh = st.text_input("📋 Số PO KH/HĐ", placeholder="VD: KH-2024-001", key="adm_so_po_kh")
            with col_bg_adm:
                adm_so_bao_gia = st.text_input("💰 Số Báo Giá", placeholder="VD: BG-2024-001", key="adm_so_bao_gia")

            adm_ten_task = st.text_input("📌 Tên công việc *", placeholder="Ví dụ: Sửa chữa động cơ bơm", key="adm_ten_task")
            adm_mo_ta    = st.text_area("📝 Mô tả chi tiết", placeholder="Nhập mô tả, yêu cầu kỹ thuật...", key="adm_mo_ta")

            st.divider()
            _fragment_checklist(_ADM_PREFIX, show_done=False, default_items=_MAC_DINH_CHECKLIST)
            st.divider()
            _fragment_cong_viec_con(_ADM_PREFIX, ds_nhan_vien, show_done=False)
            st.divider()

            if st.button("✅ Tạo Task", use_container_width=True, type="primary", key="adm_submit_task"):
                if not adm_ten_task.strip():
                    st.error("⛔ Vui lòng nhập tên công việc!")
                elif not ds_cong_ty:
                    st.error("⛔ Vui lòng thêm ít nhất một công ty trước!")
                elif adm_ma_so.strip() and not lay_danh_sach_cong_viec()[
                        lay_danh_sach_cong_viec()["Mã Số"].fillna("").str.strip() == adm_ma_so.strip()
                    ].empty:
                    _trung = lay_danh_sach_cong_viec()[
                        lay_danh_sach_cong_viec()["Mã Số"].fillna("").str.strip() == adm_ma_so.strip()
                    ].iloc[0]
                    st.error(f"⛔ Mã Số **{adm_ma_so.strip()}** đã tồn tại ở Task #{_trung.get('ID','')} — {_trung.get('Tên Công Việc','')} ({_trung.get('Công Ty','')})")
                else:
                    phe_duyet_luu = ""
                    with st.spinner("Đang lưu lên Google Sheets..."):
                        id_moi = them_cong_viec(
                            adm_ten_task.strip(), adm_mo_ta.strip(), adm_nguoi_giao,
                            adm_deadline.strftime("%Y-%m-%d"),
                            cong_ty=adm_cong_ty, cong_so="",
                            nam=adm_nam.strip(), trang_thai=adm_trang_thai,
                            nguoi_phe_duyet=phe_duyet_luu,
                            checklist=list(st.session_state.get(f"{_ADM_PREFIX}_checklist", [])),
                            cong_viec_con=list(st.session_state.get(f"{_ADM_PREFIX}_cong_viec_con")) or [{"ten": cd, "nhan_vien": "", "done": False} for cd in lay_ten_cac_cong_doan()],
                            cong_doan="",
                            loai_may=adm_loai_may if adm_loai_may != "-- Không chọn --" else "",
                            tinh_trang=adm_tinh_trang if adm_tinh_trang != "-- Không chọn --" else "",
                            cong_suat=adm_cong_suat.strip(),
                            so_cuc=adm_so_cuc.strip(),
                            ma_so=adm_ma_so.strip(),
                            so_po_noi_bo=adm_so_po_noi_bo.strip(),
                            so_po_kh=adm_so_po_kh.strip(),
                            so_bao_gia=adm_so_bao_gia.strip(),
                        )
                    st.session_state.pop(f"{_ADM_PREFIX}_checklist", None)
                    st.session_state.pop(f"{_ADM_PREFIX}_cong_viec_con", None)
                    st.session_state.pop(f"{_ADM_PREFIX}_cv_seeded", None)
                    # Chỉ xóa tên/mô tả/mã số (unique mỗi task), giữ lại các field còn lại
                    for _k in ["adm_ten_task", "adm_mo_ta", "adm_ma_so"]:
                        st.session_state.pop(_k, None)
                    st.session_state["_adm_task_success"] = (
                        f"✅ Đã tạo task #{id_moi} thành công! Công ty: **{adm_cong_ty}**"
                    )
                    st.rerun(scope="app")

        _fragment_tao_task_admin()

    # ========================================================
    # Tab 5: Bảng Quản Lý (Kanban board toàn bộ task)
    # ========================================================
    with tab_board:
        st.subheader("🗂️ Bảng Quản Lý Công Việc")
        col_adm_ref, col_adm_search = st.columns([1, 4])
        with col_adm_ref:
            if st.button("🔄 Làm mới", key="adm_board_refresh"):
                lay_danh_sach_cong_viec.clear()
                st.rerun()
        with col_adm_search:
            adm_q = st.text_input(
                "🔍",
                placeholder="Tìm kiếm theo tên công việc...",
                key="adm_board_search",
                label_visibility="collapsed",
            )

        if st.session_state.pop("_board_dirty", False):
            lay_danh_sach_cong_viec.clear()
        with st.spinner("Đang tải..."):
            df_board = lay_danh_sach_cong_viec()

        if adm_q.strip():
            _q = adm_q.strip().lower()
            df_board = df_board[
                df_board["Tên Công Việc"].fillna("").str.lower().str.contains(_q) |
                df_board["Mã Số"].fillna("").astype(str).str.lower().str.contains(_q)
            ]

        col_f1, col_f2 = st.columns(2)
        col_f1, col_f2, col_f3 = st.columns(3)
        with col_f1:
            ds_nv_f = sorted(df_board["Nhân Viên"].dropna().unique().tolist())
            loc_nv_b = st.multiselect("👤 Lọc nhân viên", ds_nv_f, key="adm_board_nv", placeholder="Tất cả")
            if loc_nv_b:
                df_board = df_board[df_board["Nhân Viên"].isin(loc_nv_b)]
        with col_f2:
            ds_ct_f = sorted(df_board["Công Ty"].dropna().unique().tolist())
            loc_ct_b = st.multiselect("🏢 Lọc công ty", ds_ct_f, key="adm_board_ct", placeholder="Tất cả")
            if loc_ct_b:
                df_board = df_board[df_board["Công Ty"].isin(loc_ct_b)]
        with col_f3:
            ds_nam_f = sorted(
                [str(y) for y in df_board["Năm"].dropna().unique() if str(y).strip() != ""],
                reverse=True,
            ) if "Năm" in df_board.columns else []
            loc_nam_b = st.selectbox(
                "📅 Lọc năm",
                options=["Tất cả"] + ds_nam_f,
                key="adm_board_nam",
            ) if ds_nam_f else "Tất cả"
            if loc_nam_b != "Tất cả":
                df_board = df_board[df_board["Năm"].astype(str) == loc_nam_b]

        ds_tt_board = lay_ten_cac_trang_thai() or _DS_TRANG_THAI_MAC_DINH

        # ── KANBAN BOARD ──────────────────────────────────────────
        df_board = df_board.sort_values("Ngày Tạo", ascending=False, na_position="last")
        _render_kanban_board(df_board, ds_tt_board, board_key="adm_kb",
                             force_open=bool(adm_q.strip()))

    # ========================================================
    # Tab 6: Quản Lý Công Việc Con
    # ========================================================
    with tab_cvc:
        st.subheader("📋 Quản Lý Công Việc Con")

        col_ref_cvc, _ = st.columns([1, 4])
        with col_ref_cvc:
            if st.button("🔄 Làm mới", key="adm_cvc_refresh"):
                lay_danh_sach_cong_viec.clear()
                st.rerun()

        with st.spinner("Đang tải dữ liệu..."):
            df_all_cvc = lay_danh_sach_cong_viec()

        _fragment_cvc_content(df_all_cvc, _aggrid_css)

    # ========================================================
    # Tab 7: Theo Dõi Tiến Độ Máy
    # ========================================================
    with tab_tdtdm:
        st.subheader("🔩 Theo Dõi Tiến Độ Máy")

        col_ref_tdm, _ = st.columns([1, 4])
        with col_ref_tdm:
            if st.button("🔄 Làm mới", key="adm_tdm_refresh"):
                lay_danh_sach_cong_viec.clear()
                st.rerun()

        with st.spinner("Đang tải dữ liệu..."):
            df_tdm_all = lay_danh_sach_cong_viec()

        _fragment_tdm_content(df_tdm_all, _aggrid_css)

    # ========================================================
    # Tab 8: Thùng Rác
    # ========================================================
    with tab_thung_rac:
        import datetime as _dt_tr
        st.subheader("🗑️ Thùng Rác")
        st.caption("Công việc bị xóa sẽ tự xóa vĩnh viễn sau **7 ngày**. Khôi phục trước khi hết hạn.")

        col_ref_tr, _ = st.columns([1, 4])
        with col_ref_tr:
            if st.button("🔄 Làm mới", key="adm_tr_refresh"):
                _fetch_tasks_cached.clear()
                st.rerun()

        with st.spinner("Đang tải..."):
            df_tr = lay_cong_viec_da_xoa()

        if df_tr.empty:
            st.info("✅ Thùng rác trống.")
        else:
            st.markdown(f"**{len(df_tr)} công việc** đang chờ xóa vĩnh viễn:")
            st.divider()
            for _, row_tr in df_tr.iterrows():
                tid_tr  = str(row_tr.get("ID", ""))
                ten_tr  = str(row_tr.get("Tên Công Việc", "") or "")
                cty_tr  = str(row_tr.get("Công Ty", "") or "")
                ngay_xoa_tr = str(row_tr.get("Ngày Xóa", "")).strip()[:10]
                # Tính số ngày còn lại
                con_lai = ""
                try:
                    ngay_xoa_d = _dt_tr.date.fromisoformat(ngay_xoa_tr)
                    het_han = ngay_xoa_d + _dt_tr.timedelta(days=7)
                    so_ngay = (het_han - _dt_tr.date.today()).days
                    if so_ngay <= 0:
                        con_lai = "⚠️ Hết hạn hôm nay"
                    elif so_ngay == 1:
                        con_lai = "⚠️ Còn 1 ngày"
                    else:
                        con_lai = f"🕐 Còn {so_ngay} ngày"
                except Exception:
                    pass

                with st.container(border=True):
                    c_info, c_restore, c_del = st.columns([5, 1, 1])
                    with c_info:
                        st.markdown(f"**{ten_tr}**")
                        meta_tr = []
                        if cty_tr:      meta_tr.append(f"🏢 {cty_tr}")
                        if ngay_xoa_tr: meta_tr.append(f"🗑️ Xóa: {ngay_xoa_tr}")
                        if con_lai:     meta_tr.append(con_lai)
                        st.caption(" · ".join(meta_tr))
                    with c_restore:
                        if st.button("♻️", key=f"tr_restore_{tid_tr}", help="Khôi phục"):
                            with st.spinner("Đang khôi phục..."):
                                khoi_phuc_cong_viec(int(tid_tr))
                            st.success(f"Đã khôi phục: {ten_tr}")
                            st.rerun()
                    with c_del:
                        if st.button("💀", key=f"tr_del_{tid_tr}", help="Xóa vĩnh viễn ngay"):
                            st.session_state[f"tr_confirm_{tid_tr}"] = True
                # Xác nhận xóa vĩnh viễn
                if st.session_state.pop(f"tr_confirm_{tid_tr}", False):
                    with st.spinner("Đang xóa vĩnh viễn..."):
                        xoa_vinh_vien_cong_viec(int(tid_tr))
                    st.rerun()

    # (pending_dlg được xử lý ở main() sau khi hàm này return)


def giao_dien_nhan_vien():
    """
    Giao diện dành cho nhân viên:
    - Nhân viên đã đăng nhập thấy công việc của chính mình (không cần chọn tên).
    - Admin khi vào chế độ NV có thể chọn bất kỳ nhân viên để xem.
    """
    st.header("👤 Trang Nhân Viên")

    vai_tro_hien_tai = st.session_state.get("vai_tro", "nhan_vien")
    ho_ten_hien_tai  = st.session_state.get("ho_ten", "")

    # Admin xem từ góc nhân viên → cho phép chọn
    if vai_tro_hien_tai == "admin":
        ten_nhan_vien = st.selectbox(
            "Chọn nhân viên để xem:",
            options=["-- Chọn nhân viên --"] + lay_danh_sach_nhan_vien()
        )
        if ten_nhan_vien == "-- Chọn nhân viên --":
            st.info("👆 Chọn nhân viên để xem công việc của họ.")
            return
    else:
        # Nhân viên đã đăng nhập → dùng tên từ session
        ten_nhan_vien = ho_ten_hien_tai
        st.markdown(f"##### Xin chào, **{ten_nhan_vien}** 👋")

    # Chuẩn hóa tên (trim khoảng trắng) để so sánh chính xác
    ten_nhan_vien = ten_nhan_vien.strip()

    # Nhân viên được cấp quyền xóa task (ngoài admin)
    _NV_CO_QUYEN_XOA = {"Nguyễn Anh Minh"}
    _co_quyen_xoa = ten_nhan_vien in _NV_CO_QUYEN_XOA

    # Mỗi lần tải trang luôn xóa cache để lấy dữ liệu mới nhất từ Google Sheets
    if "_last_nv_load" not in st.session_state or st.session_state["_last_nv_load"] != ten_nhan_vien:
        lay_danh_sach_cong_viec.clear()
        st.session_state["_last_nv_load"] = ten_nhan_vien

    # ---- Dialog thành công (phải đặt TRƯỚC tabs để luôn hiển thị) ----
    if st.session_state.get("_nv_task_success"):
        _dialog_nv_tao_task_xong(st.session_state.pop("_nv_task_success"))

    # ---- Scroll lên đầu + giữ tab Tạo Công Việc Mới sau khi đóng dialog ----
    if st.session_state.pop("_nv_scroll_top", False):
        components.html("""
        <script>
        window.parent.scrollTo({top: 0, behavior: 'smooth'});
        setTimeout(function() {
            var tabs = window.parent.document.querySelectorAll('button[role="tab"]');
            for (var i = 0; i < tabs.length; i++) {
                if (tabs[i].innerText.indexOf('T\u1ea1o C\u00f4ng Vi\u1ec7c M\u1edbi') !== -1) {
                    tabs[i].click(); break;
                }
            }
        }, 200);
        </script>
        """, height=0)

    # ---- Khôi phục tab Việc Cần Phê Duyệt sau khi dialog đóng ----
    if st.session_state.pop("_restore_tab_nv", None) == "phe_duyet":
        components.html("""
        <script>
        setTimeout(function() {
            // Tab "Việc Cần Phê Duyệt" là tab cuối cùng (index 2) trong nhóm tab NV
            var tabs = window.parent.document.querySelectorAll('button[role="tab"]');
            if (tabs.length > 0) {
                tabs[tabs.length - 1].click();
            }
        }, 150);
        </script>
        """, height=0)

    # ---- Tabs ----
    tab_cong_viec, tab_tao_task, tab_viec_cua_toi, tab_phe_duyet = st.tabs([
        "🗂️ Bảng Quản Lý Công Việc",
        "➕ Tạo Công Việc Mới",
        "📌 Việc Của Tôi",
        "✅ Việc Cần Phê Duyệt"
    ])

    # ========================================================
    # Tab 1: Công việc của tôi
    # ========================================================
    with tab_cong_viec:
        # ── Toolbar ────────────────────────────────────────────
        col_btn, col_sq = st.columns([1, 4])
        with col_btn:
            if st.button("🔄 Làm mới", key="nv_refresh_board"):
                lay_danh_sach_cong_viec.clear()
                st.session_state.pop("_last_nv_load", None)
                st.rerun()
        with col_sq:
            q_search = st.text_input(
                "🔍",
                placeholder="Tìm kiếm tên công việc...",
                key="nv_search_q",
                label_visibility="collapsed",
            )

        q_cty = st.text_input(
            "🏢",
            placeholder="Tìm kiếm theo công ty...",
            key="nv_search_cty",
            label_visibility="collapsed",
        )

        if st.session_state.pop("_board_dirty", False):
            lay_danh_sach_cong_viec.clear()
        with st.spinner("Đang tải công việc..."):
            df = lay_danh_sach_cong_viec()

        # Hiển thị tất cả công việc (không lọc theo nhân viên)
        df_cua_toi = df.copy()

        if q_search.strip():
            _q = q_search.strip().lower()
            df_cua_toi = df_cua_toi[
                df_cua_toi["Tên Công Việc"].fillna("").str.lower().str.contains(_q) |
                df_cua_toi["Mã Số"].fillna("").astype(str).str.lower().str.contains(_q)
            ]

        if q_cty.strip():
            df_cua_toi = df_cua_toi[
                df_cua_toi["Công Ty"].fillna("").str.lower().str.contains(q_cty.strip().lower())
            ]

        # Bộ lọc — mỗi cái 1 hàng full width
        ds_nam_nv = sorted(
            [str(y) for y in df_cua_toi["Năm"].dropna().unique() if str(y).strip() != ""],
            reverse=True,
        ) if "Năm" in df_cua_toi.columns else []
        loc_nam_nv = st.selectbox(
            "📅 Lọc theo năm",
            options=["Tất cả"] + ds_nam_nv,
            key="nv_loc_nam",
        ) if ds_nam_nv else "Tất cả"
        if loc_nam_nv != "Tất cả":
            df_cua_toi = df_cua_toi[df_cua_toi["Năm"].astype(str) == loc_nam_nv]

        ds_tt = lay_ten_cac_trang_thai() or _DS_TRANG_THAI_MAC_DINH

        if df_cua_toi.empty:
            st.info("Chưa có công việc nào trong hệ thống.")
        else:
            # ── KANBAN BOARD ────────────────────────────────────
            df_cua_toi = df_cua_toi.sort_values("Ngày Tạo", ascending=False, na_position="last")
            _render_kanban_board(df_cua_toi, ds_tt, board_key="nv_kb",
                                 force_open=bool(q_search.strip() or q_cty.strip()),
                                 show_delete=_co_quyen_xoa)

    # ========================================================
    # Tab 3: Việc Của Tôi (công việc con được giao)
    # ========================================================
    with tab_viec_cua_toi:
        if st.button("🔄 Làm mới", key="nv_refresh_vct"):
            lay_danh_sach_cong_viec.clear()
            st.session_state.pop("_last_nv_load", None)
            st.rerun()

        with st.spinner("Đang tải công việc..."):
            df_vct_all = lay_danh_sach_cong_viec()

        # Build danh sách tasks_dang_lam / tasks_da_xong cho user hiện tại
        _vct_dang_lam = []
        _vct_da_xong  = []
        _me_lower = ten_nhan_vien.strip().lower()
        for _, _row in df_vct_all.iterrows():
            try:
                _ds_cv = json.loads(str(_row.get("Công Việc Con") or "[]"))
            except Exception:
                _ds_cv = []
            _my_sub = [
                (_i, _cv) for _i, _cv in enumerate(_ds_cv)
                if isinstance(_cv, dict) and
                str(_cv.get("nhan_vien") or _cv.get("nguoi") or "").strip().lower() == _me_lower
            ]
            if not _my_sub:
                continue
            _hd = _row.to_dict()
            _undone = [(_i, _cv) for _i, _cv in _my_sub if not _cv.get("done", False)]
            _done   = [(_i, _cv) for _i, _cv in _my_sub if _cv.get("done", False)]
            if _undone:
                _vct_dang_lam.append((_hd, _ds_cv, _undone))
            if _done:
                _vct_da_xong.append((_hd, _ds_cv, _done))

        _n_dang_lam = sum(len(m) for _, _, m in _vct_dang_lam)
        _n_da_xong  = sum(len(m) for _, _, m in _vct_da_xong)

        _ICON_VCT = {
            "Đang Kiểm Tra": "🔵", "Đã Phê Duyệt": "🟢",
            "Đã Báo Giá": "🟠", "Có Đơn": "🟣", "Chờ Giao": "🟡",
            "Đã Hoàn Thành - Giao Máy": "✅", "Hoàn Thành": "✅",
        }

        _ds_tt_vct = lay_ten_cac_trang_thai() or _DS_TRANG_THAI_MAC_DINH

        def _render_vct_tab(tasks_list, show_done: bool):
            if not tasks_list:
                st.info("Không có việc nào đã hoàn thành." if show_done else "Không có việc nào đang làm.")
                return
            n_cv = sum(len(m) for _, _, m in tasks_list)
            st.caption(f"**{n_cv} việc con** trong **{len(tasks_list)} task**")
            for hang_dict, _, my_sub in tasks_list:
                tt     = hang_dict.get("Trạng Thái", "")
                icon   = _ICON_VCT.get(tt, "⚪")
                cty    = hang_dict.get("Công Ty", "")
                ten    = hang_dict.get("Tên Công Việc", "")
                dl     = hang_dict.get("Hạn Hoàn Thành", "") or hang_dict.get("Deadline", "")
                nv_chu = hang_dict.get("Nhân Viên", "")
                tid    = hang_dict.get("ID", "")
                with st.expander(f"{icon} [{cty}] {ten}  —  {tt}", expanded=not show_done):
                    st.markdown(f"**🏢 Công Ty:** {cty}  |  **👤** {nv_chu}  |  **📅** {dl}")
                    for _, cv in my_sub:
                        ten_cv_sub = cv.get("ten", cv.get("Tên", "—"))
                        ngay_ht    = cv.get("ngay_hoan_thanh", "")
                        if show_done:
                            st.markdown(f"✅ ~~{ten_cv_sub}~~" + (f" &nbsp;`Hoàn thành: {ngay_ht}`" if ngay_ht else ""))
                        else:
                            st.markdown(f"⏳ **{ten_cv_sub}**")
                    if st.button("📂 Xem & chỉnh sửa", key=f"vct_open_{tid}_{show_done}",
                                 use_container_width=True):
                        st.session_state.pop("_pending_dlg", None)
                        _task_dialog(hang_dict, _ds_tt_vct)

        _tab_dl, _tab_ht = st.tabs([
            f"⏳ Việc Đang Làm ({_n_dang_lam})",
            f"✅ Việc Đã Hoàn Thành ({_n_da_xong})",
        ])
        with _tab_dl:
            _render_vct_tab(_vct_dang_lam, show_done=False)
        with _tab_ht:
            _render_vct_tab(_vct_da_xong, show_done=True)

    # ========================================================
    # Tab 4: Việc Cần Phê Duyệt
    # ========================================================
    with tab_phe_duyet:
        col_btn_pd, _ = st.columns([1, 4])
        with col_btn_pd:
            if st.button("🔄 Làm mới", key="nv_refresh_pd"):
                lay_danh_sach_cong_viec.clear()
                st.rerun()

        with st.spinner("Đang tải..."):
            df_pd_all = lay_danh_sach_cong_viec()

        # Lọc: trạng thái "Đang Kiểm Tra" VÀ có Người Phê Duyệt
        df_pd = df_pd_all[
            (df_pd_all["Trạng Thái"].fillna("") == "Đang Kiểm Tra") &
            (df_pd_all["Người Phê Duyệt"].fillna("").str.strip() != "")
        ].copy()

        ds_tt_pd = lay_ten_cac_trang_thai() or _DS_TRANG_THAI_MAC_DINH

        if df_pd.empty:
            st.success("✅ Không có công việc nào đang chờ phê duyệt!")
        else:
            st.info(f"**{len(df_pd)}** công việc đang chờ phê duyệt.")
            for _, h in df_pd.iterrows():
                task_id   = h.get("ID", "")
                ten_cv    = str(h.get("Tên Công Việc", "") or "")
                cong_ty   = str(h.get("Công Ty", "") or "")
                nhan_vien = str(h.get("Nhân Viên", "") or "")
                ngay_tao  = str(h.get("Ngày Tạo", ""))[:10]
                deadline  = str(h.get("Ngày Kết Thúc", "") or "")[:10]

                today_d = datetime.now().date()
                is_over = False
                try:
                    if deadline:
                        is_over = datetime.strptime(deadline, "%Y-%m-%d").date() < today_d
                except Exception:
                    pass

                border_color = "#ef4444" if is_over else "#3b82f6"
                card_html = f"""
                <div style='border:1.5px solid {border_color};border-radius:10px;
                     padding:12px 16px;margin-bottom:10px;background:#f8faff;'>
                  <div style='font-weight:700;font-size:1rem;color:#1e3a8a;margin-bottom:4px;'>
                    #{task_id} · {ten_cv}
                  </div>
                  <div style='font-size:0.85rem;color:#475569;'>
                    🏢 {cong_ty}&nbsp;&nbsp;|&nbsp;&nbsp;👤 {nhan_vien}&nbsp;&nbsp;|&nbsp;&nbsp;
                    📅 Tạo: {ngay_tao}&nbsp;&nbsp;|&nbsp;&nbsp;
                    ⏰ Hạn: <span style='color:{"#ef4444" if is_over else "#16a34a"};font-weight:600;'>{deadline or "—"}</span>
                  </div>
                </div>"""
                st.markdown(card_html, unsafe_allow_html=True)
                if st.button("📂 Xem & Phê Duyệt", key=f"pd_open_{task_id}", use_container_width=False):
                    st.session_state["_restore_tab_nv"] = "phe_duyet"
                    _task_dialog(h.to_dict(), ds_tt_pd)

    # ========================================================
    # Tab 3: Tạo Công Việc Mới (nhân viên tự nhập)
    # ========================================================
    with tab_tao_task:
        st.subheader("➕ Tạo Công Việc Mới")
        st.info(f"Công việc sẽ được giao cho: **{ten_nhan_vien}** *(tự động)*")

        ds_cong_ty_nv = lay_ten_cac_cong_ty()
        if not ds_cong_ty_nv:
            st.warning("⚠️ Chưa có công ty nào trong hệ thống. Vui lòng liên hệ Admin để thêm công ty trước.")
        else:
            @st.fragment
            def _fragment_tao_task_nv():
                # Prefix cho session state (mỗi nhân viên dùng key riêng)
                _nv_prefix = f"nv_{ten_nhan_vien.replace(' ', '_')}"

                # Form version counter: thay đổi key → Streamlit tạo widget mới hoàn toàn
                _ver_key = f"{_nv_prefix}_form_ver"
                if _ver_key not in st.session_state:
                    st.session_state[_ver_key] = 0
                _v = st.session_state[_ver_key]

                ds_nv_nv         = lay_danh_sach_nhan_vien()
                ds_trang_thai_nv = lay_ten_cac_trang_thai() or _DS_TRANG_THAI_MAC_DINH

                # ── Hàng đầu: Công Ty ──
                # Các field sticky (giữ lại sau khi tạo task) dùng key không có version
                nv_cong_ty = st.selectbox("🏢 Công Ty *", options=["-- Chọn Công Ty --"] + ds_cong_ty_nv, key=f"{_nv_prefix}_ct")

                col_tt2, col_nam2 = st.columns(2)
                with col_tt2:
                    nv_trang_thai = st.selectbox("📋 Trạng thái", options=ds_trang_thai_nv, key=f"{_nv_prefix}_tt")
                with col_nam2:
                    nv_nam = st.text_input("📅 Năm", value=str(datetime.now().year), key=f"{_nv_prefix}_nam")

                nv_deadline = st.date_input("📅 Hạn Hoàn Thành", key=f"{_nv_prefix}_dl")

                col_lm_nv, col_tt_nv = st.columns(2)
                with col_lm_nv:
                    ds_loai_may_nv = lay_ten_cac_loai_may()
                    nv_loai_may = st.selectbox(
                        "🔧 Loại Máy",
                        options=["-- Không chọn --"] + ds_loai_may_nv,
                        key=f"{_nv_prefix}_loai_may",
                    )
                with col_tt_nv:
                    ds_tinh_trang_nv = lay_ten_cac_tinh_trang()
                    nv_tinh_trang = st.selectbox(
                        "🛠️ Tình Trạng",
                        options=["-- Không chọn --"] + ds_tinh_trang_nv,
                        key=f"{_nv_prefix}_tinh_trang",
                    )

                col_cs_nv, col_sc_nv = st.columns(2)
                with col_cs_nv:
                    nv_cong_suat = st.text_input("⚡ Công Suất", placeholder="VD: 5.5kW", key=f"{_nv_prefix}_cong_suat")
                with col_sc_nv:
                    nv_so_cuc = st.text_input("🔩 Số Cực", placeholder="VD: 4P", key=f"{_nv_prefix}_so_cuc")

                col_ms_nv, col_po_nv = st.columns(2)
                with col_ms_nv:
                    # Mã Số unique mỗi máy → reset theo version
                    nv_ma_so = st.text_input("🏷️ Mã Số", placeholder="VD: ABC-001", key=f"{_nv_prefix}_ma_so_{_v}")
                with col_po_nv:
                    nv_so_po_noi_bo = st.text_input("📄 Số PO Nội Bộ", placeholder="VD: PO-2024-001", key=f"{_nv_prefix}_so_po_noi_bo")

                col_kh_nv, col_bg_nv = st.columns(2)
                with col_kh_nv:
                    nv_so_po_kh = st.text_input("📋 Số PO KH/HĐ", placeholder="VD: KH-2024-001", key=f"{_nv_prefix}_so_po_kh")
                with col_bg_nv:
                    nv_so_bao_gia = st.text_input("💰 Số Báo Giá", placeholder="VD: BG-2024-001", key=f"{_nv_prefix}_so_bao_gia")

                # Tên & Mô tả unique mỗi task → reset theo version
                nv_ten_task = st.text_input("📌 Tên Công Việc *", placeholder="Mô tả ngắn công việc cần làm", key=f"{_nv_prefix}_ten_{_v}")
                nv_mo_ta    = st.text_area("📝 Mô Tả Chi Tiết", placeholder="Mô tả chi tiết về công việc...", key=f"{_nv_prefix}_mo_ta_{_v}")

                # Prefix có version cho checklist & công việc con → đổi key khi reset form
                _vp = f"{_nv_prefix}_v{_v}"
                _cl_key = f"{_vp}_checklist"
                _cv_key = f"{_vp}_cong_viec_con"

                st.divider()
                _fragment_checklist(_vp, show_done=False, default_items=_MAC_DINH_CHECKLIST)

                st.divider()
                _fragment_cong_viec_con(_vp, ds_nv_nv, show_done=False)

                st.divider()

                if st.button("✅ Tạo Task", use_container_width=True, type="primary", key=f"{_nv_prefix}_submit_{_v}"):
                    if nv_cong_ty == "-- Chọn Công Ty --":
                        st.error("❌ Vui lòng chọn Công Ty!")
                    elif not nv_ten_task.strip():
                        st.error("❌ Vui lòng nhập Tên Công Việc!")
                    elif nv_ma_so.strip() and not lay_danh_sach_cong_viec()[
                            lay_danh_sach_cong_viec()["Mã Số"].fillna("").str.strip() == nv_ma_so.strip()
                        ].empty:
                        _trung = lay_danh_sach_cong_viec()[
                            lay_danh_sach_cong_viec()["Mã Số"].fillna("").str.strip() == nv_ma_so.strip()
                        ].iloc[0]
                        st.error(f"⛔ Mã Số **{nv_ma_so.strip()}** đã tồn tại ở Task #{_trung.get('ID','')} — {_trung.get('Tên Task','')}")
                    else:
                        phe_duyet_nv = ""
                        with st.spinner("Đang lưu task mới..."):
                            _id_moi = them_cong_viec(
                                ten_task        = nv_ten_task.strip(),
                                mo_ta           = nv_mo_ta.strip(),
                                nguoi_duoc_giao = ten_nhan_vien,
                                deadline        = str(nv_deadline),
                                cong_ty         = nv_cong_ty if nv_cong_ty != "-- Chọn Công Ty --" else "",
                                cong_so         = "",
                                nam             = nv_nam.strip(),
                                trang_thai      = nv_trang_thai,
                                nguoi_phe_duyet = phe_duyet_nv,
                                checklist       = list(st.session_state.get(_cl_key, [])),
                                cong_viec_con   = list(st.session_state.get(_cv_key)) or [{"ten": cd, "nhan_vien": "", "done": False} for cd in lay_ten_cac_cong_doan()],
                                cong_doan       = "",
                                loai_may        = nv_loai_may if nv_loai_may != "-- Không chọn --" else "",
                                tinh_trang      = nv_tinh_trang if nv_tinh_trang != "-- Không chọn --" else "",
                                cong_suat       = nv_cong_suat.strip(),
                                so_cuc          = nv_so_cuc.strip(),
                                ma_so           = nv_ma_so.strip(),
                                so_po_noi_bo    = nv_so_po_noi_bo.strip(),
                                so_po_kh        = nv_so_po_kh.strip(),
                                so_bao_gia      = nv_so_bao_gia.strip(),
                            )
                            # Sync ảnh đo lường Nhận Máy (nếu có)
                            _nm_do_data = st.session_state.get(f"{_vp}_do_luong_creation", {})
                            if _nm_do_data and _id_moi:
                                cap_nhat_anh_do_luong(_id_moi, _nm_do_data)
                        # Tăng version → toàn bộ widget keys (kể cả checklist/cong_viec_con) đổi → reset sạch
                        st.session_state[_ver_key] = _v + 1
                        st.session_state["_nv_task_success"] = (
                            f"🎉 Đã tạo task **{nv_ten_task}** thành công! "
                            f"Chuyển sang tab **Công Việc Của Tôi** để xem."
                        )
                        st.session_state.pop("_last_nv_load", None)
                        st.rerun(scope="app")

            _fragment_tao_task_nv()


# ============================================================
# GIAO DIỆN ĐĂNG NHẬP / ĐĂNG KÝ
# ============================================================
def giao_dien_dang_nhap():
    """
    Trang đăng nhập và đăng ký.
    Lưu thông tin người dùng vào session_state sau khi xác thực thành công.
    """
    # CSS riêng cho trang auth
    st.markdown("""
    <style>
    /* Bỏ padding-top mặc định của Streamlit trên trang login */
    .block-container {
        padding-top: 1rem !important;
    }
    @media (max-width: 768px) {
        .block-container {
            padding-top: 0.5rem !important;
        }
    }
    .auth-container {
        max-width: 460px;
        margin: 1rem auto;
        background: white;
        border-radius: 16px;
        padding: 2.2rem 2rem 2rem;
        box-shadow: 0 4px 24px rgba(0,0,0,0.09);
    }
    .auth-title {
        text-align: center;
        font-size: 1.6rem;
        font-weight: 700;
        color: #2d3a5e;
        margin-bottom: 0.3rem;
    }
    .auth-sub {
        text-align: center;
        color: #7b8aa0;
        font-size: 0.9rem;
        margin-bottom: 1.5rem;
    }
    /* Fix input bị cắt đáy trên mobile */
    [data-testid="stForm"],
    [data-testid="stForm"] > div,
    [data-testid="stVerticalBlock"],
    [data-testid="stVerticalBlock"] > div,
    .element-container,
    [data-testid="stTextInput"],
    [data-testid="stTextInput"] > div {
        overflow: visible !important;
    }
    [data-testid="stTextInput"] input {
        padding-bottom: 10px !important;
    }
    </style>
    """, unsafe_allow_html=True)

    st.markdown("""
        <div class="auth-title">📋 Quản Lý Công Việc</div>
        <div class="auth-sub">Đăng nhập để tiếp tục</div>
    """, unsafe_allow_html=True)

    # JS: khôi phục session từ localStorage (nếu không phải logout thủ công)
    _is_manual_logout = st.session_state.get("manual_logout", False)
    _has_bad_token    = bool(st.query_params.get("s"))  # có token nhưng không hợp lệ
    if _is_manual_logout or _has_bad_token:
        # Đăng xuất thủ công hoặc token hết hạn → xóa localStorage
        st.components.v1.html(
            "<script>localStorage.removeItem('_qlcv_token');</script>",
            height=0,
        )
    else:
        # Thử khôi phục session từ localStorage
        st.components.v1.html("""
        <script>
        (function() {
            var t = localStorage.getItem('_qlcv_token');
            if (t) {
                var loc = window.parent.location;
                if (loc.search.indexOf('s=') === -1) {
                    loc.href = loc.pathname + '?s=' + encodeURIComponent(t);
                }
            }
        })();
        </script>""", height=0)

    # JS: đọc localStorage và tự điền username/password vào form
    st.components.v1.html("""
    <script>
    (function() {
        var UN_KEY = '_qlcv_un';
        var PW_KEY = '_qlcv_pw';
        var doc = window.parent.document;

        function fillInput(el, val) {
            var setter = Object.getOwnPropertyDescriptor(
                window.parent.HTMLInputElement.prototype, 'value'
            ).set;
            setter.call(el, val);
            el.dispatchEvent(new Event('input', { bubbles: true }));
            el.dispatchEvent(new Event('change', { bubbles: true }));
        }

        function tryAutoFill() {
            var un = localStorage.getItem(UN_KEY);
            var pw = localStorage.getItem(PW_KEY);
            if (!un || !pw) return;
            var inputs = doc.querySelectorAll('input[type="text"], input:not([type])');
            var pwInputs = doc.querySelectorAll('input[type="password"]');
            if (inputs.length > 0 && !inputs[0].value) fillInput(inputs[0], un);
            if (pwInputs.length > 0 && !pwInputs[0].value) fillInput(pwInputs[0], pw);
        }

        function hookSave() {
            // Hook vào nút Đăng Nhập để lưu credentials trước khi submit
            var btns = doc.querySelectorAll('button[kind="primaryFormSubmit"], [data-testid="baseButton-primaryFormSubmit"]');
            btns.forEach(function(btn) {
                if (btn._credHooked) return;
                btn._credHooked = true;
                btn.addEventListener('click', function() {
                    var unEl = doc.querySelector('input[type="text"], input:not([type])');
                    var pwEl = doc.querySelector('input[type="password"]');
                    if (unEl && unEl.value && pwEl && pwEl.value) {
                        localStorage.setItem(UN_KEY, unEl.value);
                        localStorage.setItem(PW_KEY, pwEl.value);
                    }
                });
            });
        }

        // Thử ngay và retry để chắc DOM đã render
        [100, 400, 900].forEach(function(d) {
            setTimeout(function() { tryAutoFill(); hookSave(); }, d);
        });
    })();
    </script>
    """, height=0)

    tab_dn, tab_dmk = st.tabs(["🔑  Đăng Nhập", "🔒  Đổi Mật Khẩu"])

    # ─────────────────────────── ĐĂNG NHẬP ───────────────────────────
    with tab_dn:
        with st.form("form_dang_nhap", clear_on_submit=False):
            _prefill_un = st.session_state.get("_saved_username", "")
            username_dn = st.text_input("Username", value=_prefill_un, placeholder="Nhập username của bạn")
            matkhau_dn  = st.text_input("Mật khẩu", type="password", placeholder="Nhập mật khẩu")
            btn_dn      = st.form_submit_button("Đăng Nhập", use_container_width=True,
                                                 type="primary")
        if btn_dn:
            if not username_dn or not matkhau_dn:
                st.error("Vui lòng nhập đầy đủ username và mật khẩu.")
            else:
                with st.spinner("Đang xác thực..."):
                    user = kiem_tra_dang_nhap(username_dn, matkhau_dn)
                if user:
                    st.session_state["dang_nhap"]     = True
                    st.session_state["user_id"]        = user["id"]
                    st.session_state["username"]       = user["username"]
                    st.session_state["ho_ten"]         = user["ho_ten"]
                    st.session_state["vai_tro"]        = user["vai_tro"]
                    st.session_state.pop("manual_logout", None)
                    st.session_state.pop("_open_task_data", None)
                    st.session_state.pop("_open_task_id", None)
                    token = _create_session_token({
                        "user_id":  str(user["id"]),
                        "username": user["username"],
                        "ho_ten":   user["ho_ten"],
                        "vai_tro":  user["vai_tro"],
                    })
                    st.session_state["session_token"] = token
                    st.query_params["s"] = token
                    st.success(f"Chào mừng, **{user['ho_ten']}**! 🎉")
                    st.rerun()
                else:
                    st.error("❌ Username hoặc mật khẩu không đúng.")

    # ─────────────────────────── ĐỔI MẬT KHẨU ───────────────────────────
    with tab_dmk:
        with st.form("form_doi_mat_khau", clear_on_submit=True):
            st.markdown("##### Đổi Mật Khẩu")
            username_dmk   = st.text_input("Username", placeholder="Nhập username của bạn")
            mk_hien_tai    = st.text_input("Mật khẩu hiện tại", type="password",
                                           placeholder="Nhập mật khẩu hiện tại")
            mk_moi         = st.text_input("Mật khẩu mới", type="password",
                                           placeholder="Tối thiểu 6 ký tự")
            xn_mk_moi      = st.text_input("Xác nhận mật khẩu mới", type="password",
                                           placeholder="Nhập lại mật khẩu mới")
            btn_dmk = st.form_submit_button("Đổi Mật Khẩu", use_container_width=True,
                                            type="primary")

        if btn_dmk:
            if not username_dmk or not mk_hien_tai or not mk_moi or not xn_mk_moi:
                st.error("❌ Vui lòng điền đầy đủ thông tin.")
            elif mk_moi != xn_mk_moi:
                st.error("❌ Mật khẩu mới xác nhận không khớp.")
            elif mk_moi == mk_hien_tai:
                st.error("❌ Mật khẩu mới phải khác mật khẩu hiện tại.")
            else:
                with st.spinner("Đang cập nhật mật khẩu..."):
                    ok, msg = doi_mat_khau(username_dmk, mk_hien_tai, mk_moi)
                if ok:
                    st.success("✅ Đổi mật khẩu thành công! Vui lòng đăng nhập lại.")
                else:
                    st.error(f"❌ {msg}")


# ============================================================
# HÀM CHÍNH
# ============================================================
def inject_css():
    st.markdown("""
    <style>
    /* ===== IMPORT GOOGLE FONT ===== */
    @import url('https://fonts.googleapis.com/css2?family=Be+Vietnam+Pro:wght@300;400;500;600;700&display=swap');

    /* ===== TOÀN CỤC ===== */
    html, body, [class*="css"] {
        font-family: 'Be Vietnam Pro', sans-serif;
    }
    /* Tắt pull-to-refresh trên Android Chrome để tránh bị reload khi kéo trong dialog */
    html, body {
        overscroll-behavior-y: none;
    }
    .stApp {
        background: linear-gradient(135deg, #f0f4ff 0%, #faf5ff 50%, #f0fdf4 100%);
        overscroll-behavior-y: none;
    }

    /* ===== st.dataframe header: purple + bold ===== */
    [data-testid="stDataFrame"] [class*="headerCell"],
    [data-testid="stDataFrame"] [class*="header-cell"],
    [data-testid="stDataFrame"] th {
        background-color: #7c3aed !important;
        color: white !important;
        font-weight: 700 !important;
        font-size: 0.82rem !important;
        letter-spacing: 0.02em !important;
        text-shadow: 0 1px 2px rgba(0,0,0,0.25) !important;
    }

    /* ===== TIÊU ĐỀ CHÍNH ===== */
    .main-header {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        padding: 0.65rem 1.4rem;
        border-radius: 14px;
        margin-bottom: 1.2rem;
        box-shadow: 0 3px 14px rgba(102,126,234,0.25);
        display: flex;
        align-items: center;
        justify-content: space-between;
        gap: 1rem;
    }
    .main-header h1 {
        color: white !important;
        font-size: 1.1rem !important;
        font-weight: 700 !important;
        margin: 0 !important;
        letter-spacing: 0.2px;
        white-space: nowrap;
    }
    .main-header-right {
        display: flex;
        align-items: center;
        gap: 0.8rem;
        flex-shrink: 0;
    }
    .main-header-user {
        color: rgba(255,255,255,0.9);
        font-size: 0.82rem;
        white-space: nowrap;
    }
    .main-header p {
        color: rgba(255,255,255,0.85) !important;
        margin: 0 !important;
        font-size: 0.8rem !important;
    }

    /* ===== ẨN SIDEBAR HOÀN TOÀN ===== */
    [data-testid="stSidebar"],
    [data-testid="stSidebarCollapsedControl"],
    [data-testid="collapsedControl"] {
        display: none !important;
    }

    /* Logout button bên trong header HTML */
    .header-logout-btn {
        background: rgba(255,255,255,0.15);
        color: white;
        border: 1.5px solid rgba(255,255,255,0.45);
        border-radius: 8px;
        font-weight: 600;
        font-size: 0.82rem;
        padding: 0.32rem 0.85rem;
        cursor: pointer;
        white-space: nowrap;
        flex-shrink: 0;
        transition: all 0.2s ease;
        font-family: inherit;
    }
    .header-logout-btn:hover {
        background: rgba(239,68,68,0.35);
        border-color: rgba(239,68,68,0.7);
    }

    /* ===== NÚT ĐĂNG XUẤT (st.button) ===== */
    .logout-btn-wrap {
        display: flex;
        align-items: center;
        justify-content: flex-end;
        height: 100%;
        padding-top: 0.35rem;
    }
    .logout-btn-wrap .stButton > button {
        background: rgba(239,68,68,0.12) !important;
        color: #ef4444 !important;
        border: 1.5px solid rgba(239,68,68,0.4) !important;
        border-radius: 8px !important;
        font-size: 1.1rem !important;
        padding: 0.3rem 0.7rem !important;
        width: 100% !important;
        transition: all 0.2s ease !important;
    }
    .logout-btn-wrap .stButton > button:hover {
        background: rgba(239,68,68,0.3) !important;
        border-color: rgba(239,68,68,0.75) !important;
    }

    /* ===== TOPBAR ACTIONS (bell + logout cột phải) ===== */
    .topbar-actions {
        display: flex;
        flex-direction: column;
        gap: 6px;
        align-items: stretch;
        justify-content: center;
        height: 100%;
        padding-top: 0.3rem;
    }
    .topbar-actions .stButton > button {
        border-radius: 8px !important;
        font-size: 1rem !important;
        padding: 0.3rem 0.4rem !important;
        width: 100% !important;
        white-space: nowrap !important;
        min-width: 0 !important;
        transition: all 0.2s ease !important;
    }
    /* Nút chuông */
    .topbar-actions .stButton:first-child > button {
        background: rgba(99,102,241,0.1) !important;
        color: #4f46e5 !important;
        border: 1.5px solid rgba(99,102,241,0.35) !important;
    }
    .topbar-actions .stButton:first-child > button:hover {
        background: rgba(99,102,241,0.22) !important;
    }
    /* Nút đăng xuất */
    .topbar-actions .stButton:last-child > button {
        background: rgba(239,68,68,0.1) !important;
        color: #ef4444 !important;
        border: 1.5px solid rgba(239,68,68,0.35) !important;
    }
    .topbar-actions .stButton:last-child > button:hover {
        background: rgba(239,68,68,0.25) !important;
    }

    /* Padding nội dung chính bình thường */
    .main .block-container {
        padding-bottom: 2rem !important;
    }

    /* Ẩn radio Streamlit gốc */
    div[data-testid="stRadio"] {
        display: none !important;
    }

    /* ===== TABS ===== */
    .stTabs [data-baseweb="tab-list"] {
        background: white;
        border-radius: 14px;
        padding: 6px;
        box-shadow: 0 2px 12px rgba(0,0,0,0.08);
        gap: 4px;
    }
    .stTabs [data-baseweb="tab"] {
        border-radius: 10px !important;
        font-weight: 600 !important;
        font-size: 0.9rem !important;
        color: #6b7280 !important;
        padding: 0.5rem 1.2rem !important;
        transition: all 0.2s ease;
    }
    .stTabs [aria-selected="true"] {
        background: linear-gradient(135deg, #667eea, #764ba2) !important;
        color: white !important;
        box-shadow: 0 4px 12px rgba(102,126,234,0.4) !important;
    }

    /* ===== METRIC CARDS ===== */
    [data-testid="stMetric"] {
        background: white;
        border-radius: 16px;
        padding: 1.2rem 1.5rem !important;
        box-shadow: 0 4px 20px rgba(0,0,0,0.07);
        border-left: 5px solid #667eea;
        transition: transform 0.2s ease;
    }
    [data-testid="stMetric"]:hover {
        transform: translateY(-3px);
        box-shadow: 0 8px 25px rgba(102,126,234,0.2);
    }
    [data-testid="stMetricLabel"] {
        font-weight: 600 !important;
        color: #6b7280 !important;
        font-size: 0.85rem !important;
        text-transform: uppercase;
        letter-spacing: 0.5px;
    }
    [data-testid="stMetricValue"] {
        font-size: 2rem !important;
        font-weight: 700 !important;
        color: #1e1b4b !important;
    }

    /* ===== BUTTONS ===== */
    .stButton > button {
        border-radius: 10px !important;
        font-weight: 600 !important;
        font-size: 0.9rem !important;
        padding: 0.5rem 1.2rem !important;
        border: none !important;
        transition: all 0.25s ease !important;
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%) !important;
        color: white !important;
        box-shadow: 0 4px 15px rgba(102,126,234,0.35) !important;
    }
    .stButton > button:hover {
        transform: translateY(-2px) !important;
        box-shadow: 0 6px 20px rgba(102,126,234,0.5) !important;
        opacity: 0.95 !important;
    }
    .stButton > button:active {
        transform: translateY(0px) !important;
    }
    .stButton > button:disabled,
    .stButton > button[disabled] {
        background: #c8c8d4 !important;
        color: #888 !important;
        box-shadow: none !important;
        cursor: not-allowed !important;
        transform: none !important;
        opacity: 0.7 !important;
    }

    /* ===== DOWNLOAD BUTTON ===== */
    .stDownloadButton > button {
        border-radius: 10px !important;
        font-weight: 600 !important;
        background: linear-gradient(135deg, #10b981, #059669) !important;
        color: white !important;
        border: none !important;
        box-shadow: 0 4px 15px rgba(16,185,129,0.35) !important;
        transition: all 0.25s ease !important;
    }
    .stDownloadButton > button:hover {
        transform: translateY(-2px) !important;
        box-shadow: 0 6px 20px rgba(16,185,129,0.5) !important;
    }

    /* ===== FORM & INPUT ===== */
    .stTextInput > div > div > input,
    .stTextArea > div > div > textarea,
    .stSelectbox > div > div {
        border-radius: 10px !important;
        border: 2px solid #e5e7eb !important;
        background: white !important;
        transition: border-color 0.2s ease;
        font-family: 'Be Vietnam Pro', sans-serif !important;
    }
    .stTextInput > div > div > input:focus,
    .stTextArea > div > div > textarea:focus {
        border-color: #667eea !important;
        box-shadow: 0 0 0 3px rgba(102,126,234,0.15) !important;
    }

    /* ===== FORM SUBMIT BUTTON ===== */
    [data-testid="stFormSubmitButton"] > button {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%) !important;
        color: white !important;
        border: none !important;
        border-radius: 12px !important;
        font-weight: 700 !important;
        font-size: 1rem !important;
        padding: 0.65rem !important;
        box-shadow: 0 4px 15px rgba(102,126,234,0.4) !important;
        transition: all 0.25s ease !important;
    }
    [data-testid="stFormSubmitButton"] > button:hover {
        transform: translateY(-2px) !important;
        box-shadow: 0 6px 22px rgba(102,126,234,0.55) !important;
    }

    /* ===== EXPANDER (TASK CARDS) ===== */
    [data-testid="stExpander"] {
        background: white !important;
        border-radius: 16px !important;
        border: 1.5px solid #e5e7eb !important;
        box-shadow: 0 2px 12px rgba(0,0,0,0.06) !important;
        margin-bottom: 0.8rem !important;
        overflow: hidden !important;
        transition: box-shadow 0.2s ease !important;
    }
    [data-testid="stExpander"]:hover {
        box-shadow: 0 6px 20px rgba(102,126,234,0.15) !important;
        border-color: #667eea !important;
    }
    [data-testid="stExpander"] summary {
        font-weight: 600 !important;
        font-size: 0.95rem !important;
        padding: 1rem 1.2rem !important;
        background: linear-gradient(90deg, #f8f7ff, #f3f4f6) !important;
        color: #1e1b4b !important;
    }

    /* ===== DATAFRAME ===== */
    [data-testid="stDataFrame"] {
        border-radius: 14px !important;
        overflow: hidden !important;
        box-shadow: 0 4px 20px rgba(0,0,0,0.07) !important;
        border: none !important;
    }

    /* ===== ALERT / INFO / SUCCESS / WARNING ===== */
    [data-testid="stAlert"] {
        border-radius: 12px !important;
        border: none !important;
        font-weight: 500 !important;
    }

    /* ===== SUBHEADER ===== */
    h2, h3 {
        color: #1e1b4b !important;
        font-weight: 700 !important;
    }

    /* ===== DIVIDER ===== */
    hr {
        border-color: #e5e7eb !important;
        margin: 1.2rem 0 !important;
    }

    /* ===== SPINNER ===== */
    .stSpinner > div {
        border-top-color: #667eea !important;
    }

    /* ===== FILE UPLOADER ===== */
    [data-testid="stFileUploader"] {
        border-radius: 12px !important;
        border: 2px dashed #667eea !important;
        background: #f8f7ff !important;
        padding: 0.5rem !important;
    }

    /* ===== MOBILE RESPONSIVE ===== */
    @media (max-width: 768px) {

        /* Main content full width, padding bottom cho bottom nav */
        .main .block-container {
            padding: 0.3rem 0.8rem 90px 0.8rem !important;
            max-width: 100% !important;
            margin-left: 0 !important;
            margin-top: 0 !important;
        }
        section[data-testid="stMain"] > div:first-child {
            padding-top: 0 !important;
        }

        /* --- Header nhỏ lại & stack dọc --- */
        .main-header {
            padding: 0.5rem 0.8rem !important;
            border-radius: 10px !important;
            flex-direction: column !important;
            align-items: flex-start !important;
            gap: 0.2rem !important;
        }
        .main-header h1 {
            font-size: 0.88rem !important;
            white-space: normal !important;
        }
        .main-header-user {
            font-size: 0.8rem !important;
            white-space: normal !important;
            word-break: break-word !important;
        }
        .main-header p {
            font-size: 0.75rem !important;
        }

        /* --- Form container full width --- */
        [data-testid="stForm"] {
            width: 100% !important;
            padding: 0 !important;
        }

        /* --- Metric cards full width --- */
        [data-testid="stMetric"] {
            padding: 0.9rem 1rem !important;
            margin-bottom: 0.5rem !important;
        }
        [data-testid="stMetricValue"] {
            font-size: 1.6rem !important;
        }

        /* --- Tabs: nhỏ hơn, scroll ngang --- */
        .stTabs [data-baseweb="tab-list"] {
            overflow-x: auto !important;
            flex-wrap: nowrap !important;
            padding: 4px !important;
            gap: 2px !important;
            -webkit-overflow-scrolling: touch;
        }
        .stTabs [data-baseweb="tab"] {
            font-size: 0.78rem !important;
            padding: 0.4rem 0.7rem !important;
            white-space: nowrap !important;
            flex-shrink: 0 !important;
        }

        /* --- Buttons: touch target lớn hơn --- */
        .stButton > button {
            font-size: 1rem !important;
            padding: 0.75rem 1rem !important;
            width: 100% !important;
            min-height: 48px !important;
        }
        [data-testid="stFormSubmitButton"] > button {
            min-height: 52px !important;
            font-size: 1rem !important;
        }
        .stDownloadButton > button {
            min-height: 48px !important;
            font-size: 1rem !important;
            width: 100% !important;
        }

        /* --- Input fields: lớn hơn cho mobile --- */
        .stTextInput > div > div > input,
        .stTextArea > div > div > textarea {
            font-size: 1rem !important;
            padding: 0.6rem 0.8rem !important;
            min-height: 44px !important;
        }
        .stSelectbox > div > div {
            min-height: 44px !important;
            height: auto !important;
        }
        [data-testid="stSelectbox"] [data-baseweb="select"] > div {
            height: auto !important;
            min-height: 44px !important;
        }
        [data-testid="stSelectbox"] [data-baseweb="select"] input,
        [data-testid="stSelectbox"] input {
            direction: ltr !important;
            text-align: left !important;
        }
        [data-testid="stSelectbox"] div[class*="singleValue"],
        [data-testid="stSelectbox"] div[class*="SingleValue"],
        [data-testid="stSelectbox"] div[class*="single-value"] {
            white-space: nowrap !important;
            overflow: hidden !important;
            text-overflow: ellipsis !important;
            font-size: 0.72rem !important;
            max-width: calc(100% - 28px) !important;
            direction: rtl !important;
            text-align: left !important;
        }

        /* --- Fix input bị cắt đáy trong column --- */
        [data-testid="stColumn"] > [data-testid="stVerticalBlock"] {
            overflow: visible !important;
            padding-bottom: 4px !important;
        }

        /* --- Expander: readable trên mobile --- */
        [data-testid="stExpander"] summary {
            font-size: 0.85rem !important;
            padding: 0.8rem 1rem !important;
            line-height: 1.4 !important;
        }

        /* --- File uploader: dễ chạm hơn --- */
        [data-testid="stFileUploader"] {
            padding: 1rem !important;
        }
        [data-testid="stFileUploader"] label {
            font-size: 0.9rem !important;
        }

        /* --- Ảnh full width --- */
        [data-testid="stImage"] img {
            width: 100% !important;
            height: auto !important;
            border-radius: 10px !important;
        }

        /* --- Dataframe: scroll ngang --- */
        [data-testid="stDataFrame"] {
            overflow-x: auto !important;
            -webkit-overflow-scrolling: touch;
        }

        /* --- Text size --- */
        p, li, label, .stMarkdown {
            font-size: 0.92rem !important;
        }
        h2 { font-size: 1.15rem !important; }
        h3 { font-size: 1rem !important; }

        /* --- Radio buttons: dễ chạm hơn --- */
        [data-testid="stRadio"] label {
            padding: 0.5rem 0 !important;
            font-size: 1rem !important;
        }
        [data-testid="stRadio"] > div {
            gap: 0.4rem !important;
        }

        /* --- Date input --- */
        input[type="date"] {
            min-height: 44px !important;
            font-size: 1rem !important;
        }

        /* --- Divider margin nhỏ hơn --- */
        hr {
            margin: 0.7rem 0 !important;
        }

        /* --- Alert --- */
        [data-testid="stAlert"] {
            font-size: 0.88rem !important;
            padding: 0.7rem 0.9rem !important;
        }

        /* --- Kanban cards: cuộn ngang thay vì xếp chồng --- */
        /* Chỉ áp dụng cho row có 3+ cột (kanban 4 cards), không ảnh hưởng filter 2 cột */
        [data-testid="stHorizontalBlock"]:has(> [data-testid="stColumn"]:nth-child(3)) {
            flex-wrap: nowrap !important;
            overflow-x: auto !important;
            -webkit-overflow-scrolling: touch !important;
            gap: 10px !important;
            padding-bottom: 10px !important;
            scroll-snap-type: x mandatory !important;
        }
        [data-testid="stHorizontalBlock"]:has(> [data-testid="stColumn"]:nth-child(3))
            > [data-testid="stColumn"] {
            min-width: 230px !important;
            flex: 0 0 230px !important;
            scroll-snap-align: start !important;
        }
        /* Ẩn scrollbar nhưng vẫn scroll được */
        [data-testid="stHorizontalBlock"]:has(> [data-testid="stColumn"]:nth-child(3))::-webkit-scrollbar {
            height: 4px !important;
        }
        [data-testid="stHorizontalBlock"]:has(> [data-testid="stColumn"]:nth-child(3))::-webkit-scrollbar-thumb {
            background: #c4b5fd !important;
            border-radius: 4px !important;
        }
    }

    /* ===== TABLET (768–1024px) ===== */
    @media (min-width: 769px) and (max-width: 1024px) {
        .main .block-container {
            padding: 1rem 1.5rem !important;
        }
        .main-header h1 {
            font-size: 1rem !important;
        }
        .stTabs [data-baseweb="tab"] {
            font-size: 0.85rem !important;
            padding: 0.45rem 0.9rem !important;
        }
    }
    </style>
    <script>
    (function() {
        var _VI_M = {
            'January':'Tháng 1','February':'Tháng 2','March':'Tháng 3',
            'April':'Tháng 4','May':'Tháng 5','June':'Tháng 6',
            'July':'Tháng 7','August':'Tháng 8','September':'Tháng 9',
            'October':'Tháng 10','November':'Tháng 11','December':'Tháng 12'
        };
        function replaceMonths(root) {
            try {
                // 1. select option
                (root.querySelectorAll ? root : document).querySelectorAll('select option').forEach(function(o){
                    var t=(o.text||'').trim(); if(_VI_M[t]) o.text=_VI_M[t];
                });
                // 2. TreeWalker: text node chứa đúng tên tháng
                var walk=document.createTreeWalker(root.nodeType===9?root.documentElement:root,4,null,false),n;
                while((n=walk.nextNode())){
                    var v=n.nodeValue; if(!v) continue;
                    var tr=v.trim(); if(_VI_M[tr]) n.nodeValue=v.replace(tr,_VI_M[tr]);
                }
            } catch(e){}
        }
        // Chạy ngay + sau load
        document.addEventListener('DOMContentLoaded',function(){replaceMonths(document);});
        setTimeout(function(){replaceMonths(document);},800);
        // MutationObserver: bắt calendar popup
        var _calObs=new MutationObserver(function(ms){
            ms.forEach(function(m){
                m.addedNodes.forEach(function(nd){
                    if(nd.nodeType===1) replaceMonths(nd);
                });
            });
        });
        _calObs.observe(document.documentElement,{childList:true,subtree:true});
    })();
    </script>
    <script>
    /* Reload trang sau 3 phút switch app để tránh app bị đơ/freeze */
    (function(){
        var THRESHOLD = 180000; // 3 phút
        var _hiddenAt = null;
        document.addEventListener('visibilitychange', function(){
            if(document.visibilityState === 'hidden'){
                _hiddenAt = Date.now();
            } else if(document.visibilityState === 'visible' && _hiddenAt !== null){
                var elapsed = Date.now() - _hiddenAt;
                _hiddenAt = null;
                if(elapsed >= THRESHOLD){
                    var token  = localStorage.getItem('_qlcv_token') || '';
                    var base   = window.location.origin + window.location.pathname;
                    window.location.href = base + (token ? '?s=' + token : '');
                }
            }
        });
    })();
    </script>
    """, unsafe_allow_html=True)


# ============================================================
# DIALOG THÔNG BÁO
# ============================================================
@st.dialog("🔔 Thông Báo", width="large")
def dialog_thong_bao(ho_ten: str):
    """Dialog hiển thị danh sách thông báo của user."""
    # Dọn thông báo cũ hơn 2 tuần (chạy 1 lần mỗi session)
    if not st.session_state.get("_da_xoa_tb_cu", False):
        xoa_thong_bao_cu()
        st.session_state["_da_xoa_tb_cu"] = True
    c_title, c_mark = st.columns([5, 2])
    with c_title:
        so_chua = dem_chua_doc(ho_ten)
        if so_chua > 0:
            st.markdown(f"**{so_chua} thông báo chưa đọc**")
    with c_mark:
        if st.button("✅ Đánh dấu tất cả đã đọc",
                     key="dlg_mark_all_read",
                     use_container_width=True):
            danh_dau_da_doc_tat_ca(ho_ten)
            st.rerun()

    st.divider()

    df_tb = lay_thong_bao_nguoi_dung(ho_ten)
    if df_tb.empty:
        st.info("🔔 Bạn chưa có thông báo nào.")
        return

    # CSS: style nút "Xem Chi Tiết" nhỏ như text link
    st.markdown("""<style>
    [data-testid="stDialog"] .tb-xem-btn button {
        background: none !important;
        border: none !important;
        box-shadow: none !important;
        color: #6366f1 !important;
        font-size: 0.82rem !important;
        padding: 0 0 8px 58px !important;
        min-height: unset !important;
        height: auto !important;
        text-decoration: underline !important;
        cursor: pointer !important;
        justify-content: flex-start !important;
    }
    [data-testid="stDialog"] .tb-xem-btn button:hover {
        color: #4338ca !important;
        background: none !important;
    }
    [data-testid="stDialog"] .tb-xem-btn {
        margin-top: -4px !important;
        margin-bottom: 6px !important;
    }
    </style>""", unsafe_allow_html=True)

    for _, tb_row in df_tb.head(50).iterrows():
        _da_doc   = str(tb_row.get("Da_Doc", "0")) == "1"
        _noi_dung = str(tb_row.get("Noi_Dung", ""))
        _tgian    = str(tb_row.get("Thoi_Gian", ""))
        _loai     = str(tb_row.get("Loai", ""))
        _task_id  = str(tb_row.get("Task_ID", ""))

        # Icon + màu avatar theo loại
        _icon_map = {
            "task_moi":     ("📋", "#6366f1"),
            "trang_thai":   ("🔄", "#0ea5e9"),
            "cong_viec_con":("🔧", "#f59e0b"),
        }
        _icon, _col_av = _icon_map.get(_loai, ("🔔", "#6b7280"))

        # Thời gian tương đối
        try:
            _dt = datetime.strptime(_tgian[:19], "%Y-%m-%d %H:%M:%S")
            _delta = datetime.now() - _dt
            if _delta.seconds < 60:              _tg_text = "Vừa xong"
            elif _delta.seconds < 3600:          _tg_text = f"{_delta.seconds // 60} phút trước"
            elif _delta.days == 0:               _tg_text = f"{_delta.seconds // 3600} giờ trước"
            elif _delta.days == 1:               _tg_text = "Hôm qua"
            else:                                _tg_text = _dt.strftime("%d/%m/%Y")
        except Exception:
            _tg_text = _tgian

        _bg      = "#eff6ff" if not _da_doc else "#f9fafb"
        _border  = "2px solid #bfdbfe" if not _da_doc else "1px solid #e5e7eb"
        _dot_html = "<span style='display:inline-block;width:8px;height:8px;border-radius:50%;background:#3b82f6;margin-left:6px;vertical-align:middle'></span>" if not _da_doc else ""
        _has_task = _task_id and _task_id != '0'
        _mk = f"tbmk{tb_row.name}"
        _cursor = "cursor:pointer;" if _has_task else ""

        # Card HTML gốc
        st.markdown(f"""
        <div style="display:flex;gap:12px;align-items:flex-start;
                    background:{_bg};border:{_border};
                    border-radius:12px;padding:12px 14px;margin-bottom:0;">
            <div style="flex-shrink:0;width:42px;height:42px;border-radius:50%;
                        background:{_col_av};display:flex;align-items:center;
                        justify-content:center;font-size:1.25rem">{_icon}</div>
            <div style="flex:1;min-width:0">
                <div style="font-size:0.93rem;color:#1e293b;line-height:1.45">{_noi_dung}{_dot_html}</div>
                <div style="font-size:0.78rem;color:#94a3b8;margin-top:4px">{_tg_text}
                    {f"&nbsp;&middot;&nbsp;Task #{_task_id}" if _has_task else ""}
                </div>
            </div>
        </div>""", unsafe_allow_html=True)

        # Nút "Xem Chi Tiết" nhỏ styled như text link
        if _has_task:
            st.markdown('<div class="tb-xem-btn">', unsafe_allow_html=True)
            if st.button("📂 Xem Chi Tiết", key=f"tb_card_{tb_row.name}"):
                st.session_state["_open_task_id_from_tb"] = _task_id
                st.rerun()
            st.markdown('</div>', unsafe_allow_html=True)
        else:
            st.markdown("<div style='margin-bottom:8px'></div>", unsafe_allow_html=True)


def main():
    """Hàm chính: khởi động và điều hướng ứng dụng."""
    inject_css()

    # ── Khôi phục session từ query param nếu chưa đăng nhập ──────────────
    if not st.session_state.get("dang_nhap"):
        if st.session_state.get("manual_logout"):
            giao_dien_dang_nhap()
            return

        token = st.query_params.get("s")
        sess  = _verify_session_token(token) if token else None

        if sess:
            uid            = sess["user_id"]
            ho_ten_sess    = sess["ho_ten"]
            vai_tro_sess   = sess["vai_tro"]
            try:
                df_users = lay_danh_sach_users()
                if not df_users.empty:
                    user_row = df_users[df_users["ID"].astype(str) == str(uid)]
                    if not user_row.empty:
                        ho_ten_sess  = user_row.iloc[0]["HoTen"]
                        vai_tro_sess = user_row.iloc[0]["VaiTro"]
            except Exception:
                pass
            st.session_state["dang_nhap"]      = True
            st.session_state["user_id"]        = uid
            st.session_state["username"]       = sess["username"]
            st.session_state["ho_ten"]         = ho_ten_sess
            st.session_state["vai_tro"]        = vai_tro_sess
            st.session_state["session_token"]  = token
            st.rerun()
        else:
            giao_dien_dang_nhap()
            return

    # Lưu token vào localStorage để khôi phục session khi reload
    _cur_token = st.session_state.get("session_token", "")
    if _cur_token:
        st.components.v1.html(
            f"<script>localStorage.setItem('_qlcv_token','{_cur_token}');</script>",
            height=0,
        )

    # Dọn thùng rác (chạy 1 lần mỗi session, không block UI)
    if not st.session_state.get("_da_don_thung_rac"):
        st.session_state["_da_don_thung_rac"] = True
        try:
            don_dep_thung_rac()
        except Exception:
            pass

    vai_tro   = st.session_state.get("vai_tro", "nhan_vien")
    ho_ten    = st.session_state.get("ho_ten", "")

    # ── Toast thông báo mới (tự động hiện ở góc màn hình) ─────────────────
    try:
        df_tb_all = lay_thong_bao_nguoi_dung(ho_ten)
        if not df_tb_all.empty:
            # Lấy ID lớn nhất hiện tại làm mốc
            _ids_hien_tai = pd.to_numeric(df_tb_all["ID"], errors="coerce").dropna()
            _max_id_hien_tai = int(_ids_hien_tai.max()) if not _ids_hien_tai.empty else 0

            _key_seen = f"_tb_seen_max_id_{ho_ten}"
            _last_seen_id = st.session_state.get(_key_seen, None)

            if _last_seen_id is None:
                # Lần đầu đăng nhập → ghi nhận mốc, không toast
                st.session_state[_key_seen] = _max_id_hien_tai
            elif _max_id_hien_tai > _last_seen_id:
                # Có thông báo mới → toast từng cái (tối đa 3)
                _moi = df_tb_all[
                    pd.to_numeric(df_tb_all["ID"], errors="coerce") > _last_seen_id
                ].head(3)
                _loai_icon = {
                    "task_moi":      "📋",
                    "trang_thai":    "🔄",
                    "cong_viec_con": "🔧",
                }
                for _, _r in _moi.iterrows():
                    _icon = _loai_icon.get(str(_r.get("Loai", "")), "🔔")
                    _nd   = str(_r.get("Noi_Dung", ""))
                    # Rút gọn nội dung ≤ 80 ký tự
                    _nd_short = (_nd[:77] + "…") if len(_nd) > 80 else _nd
                    st.toast(f"{_icon} {_nd_short}", icon="🔔")
                st.session_state[_key_seen] = _max_id_hien_tai
    except Exception:
        pass

    # ── Header + nút đăng xuất ──────────────────────────────────────────────
    role_badge   = "🛡️ Admin" if vai_tro == "admin" else "👤"
    so_chua_doc  = dem_chua_doc(ho_ten)
    _bell_label  = f"🔔 {so_chua_doc}" if so_chua_doc > 0 else "🔔"

    col_hdr, col_actions = st.columns([5, 1])
    with col_hdr:
        st.markdown(f"""
            <div class="main-header">
                <h1>📋 Quản Lý Công Việc</h1>
                <div class="main-header-user">{role_badge} &nbsp;<b>{ho_ten}</b></div>
            </div>
        """, unsafe_allow_html=True)
    with col_actions:
        st.markdown('<div class="topbar-actions">', unsafe_allow_html=True)
        if st.button(_bell_label, key="btn_bell", use_container_width=True,
                     help="Thông báo của bạn"):
            dialog_thong_bao(ho_ten)
        if st.button("🚪 Đăng Xuất", key="topbar_logout", use_container_width=True,
                     help="Đăng xuất"):
            token = st.session_state.get("session_token")
            if token:
                _session_store().pop(token, None)
            _saved_un = st.session_state.get("username", "")
            for k in ["dang_nhap", "user_id", "username", "ho_ten", "vai_tro", "session_token"]:
                st.session_state.pop(k, None)
            st.query_params.clear()
            st.session_state["manual_logout"] = True
            st.session_state["_saved_username"] = _saved_un
            st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)

    # ── Trạng thái dialog ─────────────────────────────────────────────────
    # _is_fresh_session = True chỉ khi page thực sự reload (session mới hoàn toàn)
    # Khác với st.rerun() — rerun trong cùng session vẫn giữ session_marker
    _is_fresh_session = "session_marker" not in st.session_state
    st.session_state["session_marker"] = True
    _dlg_was_active = st.session_state.get("_dlg_active_flag", False)
    st.session_state["_dlg_active_flag"] = False

    # ── Mở task dialog từ thông báo (sau rerun) ───────────────────────────
    if st.session_state.get("_open_task_id_from_tb"):
        _tid_tb = st.session_state.pop("_open_task_id_from_tb")
        st.session_state.pop("_open_task_id_from_tb_user", None)
        try:
            _df_all  = lay_danh_sach_cong_viec()
            _ds_tt   = lay_ten_cac_trang_thai()
            _matched = _df_all[_df_all["ID"].astype(str) == str(_tid_tb)]
            if not _matched.empty:
                _task_dialog(_matched.iloc[0].to_dict(), _ds_tt)
        except Exception:
            st.warning(f"Không tìm thấy Task #{_tid_tb}")

    # ── Điều hướng giao diện ───────────────────────────────────────────────
    if vai_tro == "admin":
        giao_dien_admin()
    else:
        giao_dien_nhan_vien()

    # ── Xử lý dialog: pending / khôi phục / phát hiện đóng bằng X ─────────
    _dlg_qp = st.query_params.get("dlg", "")
    _pdlg   = st.session_state.pop("_pending_dlg", None)

    _dlg_currently_active = st.session_state.get("_dlg_active_flag", False)

    if _pdlg:
        # Mở dialog từ click button / row selection
        _task_dialog(_pdlg[0], _pdlg[1])
    elif _dlg_qp and _is_fresh_session:
        # Session mới (page reload / reconnect sau khi background) → khôi phục dialog
        try:
            _df_r  = lay_danh_sach_cong_viec()
            _row_r = _df_r[_df_r["ID"].astype(str) == _dlg_qp]
            if not _row_r.empty:
                _task_dialog(_row_r.iloc[0].to_dict(),
                             lay_ten_cac_trang_thai() or _DS_TRANG_THAI_MAC_DINH)
            else:
                del st.query_params["dlg"]
        except Exception:
            try:
                del st.query_params["dlg"]
            except Exception:
                pass
    elif _dlg_qp and not _is_fresh_session and _dlg_was_active and not _dlg_currently_active:
        # Cùng session, dialog đóng bằng nút X → xóa query param
        try:
            del st.query_params["dlg"]
        except Exception:
            pass


# ============================================================
# ĐIỂM CHẠY CHƯƠNG TRÌNH
# ============================================================
if __name__ == "__main__":
    main()
